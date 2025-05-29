# 套件匯入
import openpyxl.workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
import pyautogui as pag
from selenium.webdriver.chrome.service import Service
import requests
from http import HTTPStatus
from selenium.webdriver.chrome.options import Options
from lxml import html
import csv
import numpy as np
import pandas as pd
import cv2
import pybi as pbi
import os
import sys
import aspose
import xlrd
import xlwt
import glob
import cx_Freeze
from cx_Freeze import setup
import setuptools
import jpype
jpype.startJVM()
from asposecells.api import Workbook, FileFormatType
import difflib
import openpyxl
from openpyxl import Workbook
from openpyxl import workbook
from openpyxl import worksheet
from openpyxl.styles import Font  # 導入字體模組
from openpyxl.styles import PatternFill  # 導入填充模組
from spire.xls import *
from collections import deque
from pandas.core.frame import DataFrame
# 导入openpyxl模块并将其重命名为pxl
import openpyxl as pxl
# 从openpyxl导入PatternFill类
from openpyxl.styles import PatternFill
# 从openpyxl导入colors类
from openpyxl.styles import colors
# 从openpyxl导入Font类
from openpyxl.styles import Font
import datetime
import pytesseract
from PIL import Image
# import ddddocr
import xlwings
import getpass
# import imagehash
from pathlib import Path
import matplotlib.pyplot as plt
from PIL import Image, ImageEnhance
import re
from bs4 import BeautifulSoup
import pykeyboard
from selenium.common.exceptions import NoAlertPresentException
import BeautifulReport
import unittest, doctest
import HTMLTestRunner

# ================================== 建立文件放置資料夾 ==================================

path = r'D:/'
p = Path(path)

# 定義子資料夾名稱
sub_folder = 'AutomotiveTest_QA'
# 連接主資料夾與子資料夾
p = p.joinpath(sub_folder)
p.mkdir(exist_ok=True)

# 定義子資料夾名稱
sub_folder_1 = 'Back_platform Test'
# 連接主資料夾與子資料夾
p = p.joinpath(sub_folder_1)
p.mkdir(exist_ok=True)

# 定義子資料夾名稱
sub_folder_2 = 'v 3.0.3.1756 250424'
# 連接主資料夾與子資料夾
p = p.joinpath(sub_folder_2)
p.mkdir(exist_ok=True)

# 定義子資料夾名稱
sub_folder_3 = 'Report'
# 連接主資料夾與子資料夾
p = p.joinpath(sub_folder_3)
p.mkdir(exist_ok=True)

# ============================= PS後台連接與登入 ============================= 

class PS_platform_testCase(unittest.TestCase):
    driver = webdriver.Chrome()

    def PS_platform_access(self):
        
        current_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
        print('測試起始時間: ', current_time, '\n')
        
        print('PS後台連線測試...', '\n')
        url = 'https://dev-admin-br-02.claretfox.com'
        # self.driver = webdriver.Chrome()
        self.driver.get(url)
        time.sleep(1)
        self.req = requests.get('https://dev-admin-br-02.claretfox.com')
        req = requests.post(url)
        state_code = req.status_code
        print('Http Response Code: ', state_code, '\n')
        if state_code == 200:
            print("HTTP回應成功!", '\n')
        else:
            print('Http Response Code:', state_code, '\n')
            print("HTTP回應失敗!", '\n')
        self.driver.find_element(By.XPATH, '/html/body/div[2]/div/form/label').click()
        self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[2]').click()
        print("語系已切換'繁體中文'!", '\n')
        time.sleep(1)
        
    def PS_platform_logging(self):
    
        self.driver.find_element(By.CLASS_NAME, 'content-group').click()  # 尋找登入介面元素位址

        self.driver.find_element(By.XPATH, '//*[@id="user_id"]').clear()  # 預設此欄位為null, 但仍先清除帳號欄位資訊
        self.driver.find_element(By.XPATH, '//*[@id="password"]').clear()  # 預設此欄位為null, 但仍先清除密碼欄位資訊
        self.driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div/form/div/div[4]/input')  # 驗證碼欄位
        self.driver.find_element(By.CSS_SELECTOR, '#captcha_img > img').screenshot('verificationCode.png')  # 驗證碼截圖存檔
        time.sleep(1)
        
        accountValue = getpass.getpass('請輸入後台帳號: ')
        self.driver.find_element(By.XPATH, '//*[@id="user_id"]').send_keys(accountValue)   # 輸入使用者帳號
        time.sleep(1)
        self.driver.find_element(By.XPATH, '//*[@id="user_id"]').send_keys(Keys.TAB)  # 切換至密碼輸入欄位
        
        pwdValue = getpass.getpass('請輸入後台密碼: ')
        self.driver.find_element(By.XPATH, '//*[@id="password"]').send_keys(pwdValue)  # 輸入個人密碼
        time.sleep(1)
        self.driver.find_element(By.XPATH, '//*[@id="password"]').send_keys(Keys.TAB)   # 切換至驗證碼輸入欄位
        
        verifiCodeValue = input('請輸入驗證碼: ')  # 輸入驗證碼
        print(verifiCodeValue, '\n')
        self.driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div/form/div/div[4]/input').send_keys(verifiCodeValue)
        self.driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div/form/div/div[6]/button').click()
        time.sleep(1)
        self.driver.find_element(By.XPATH, '/html/body/div[3]/div/div[4]/div[1]/div/div/h4/span')
        
        try:
            element = self.driver.find_element(By.XPATH, '/html/body/div[3]/div/div[4]/div[1]/div/div/h4/span')
            element.click()
            str_error = None
            print("登入成功!", '\n')
            time.sleep(1)
            os.remove('verificationCode.png')
        except Exception as e:
            print(e)
            str_error = True
            pass

        if str_error:
            print("登入資訊輸入錯誤，請重新輸入!", '\n')
            self.driver.refresh()
            

# ============================= 後台功能巡測【配置】============================= 
# *************************** 代理商管理 ***************************

    def Configuration_Agent_Management_zhTW(self):

        self.driver.find_element(By.ID, 'Configuration').click()
        print("進入配置功能選單!", '\n')
        time.sleep(1)

        self.driver.get('https://dev-admin-br-02.claretfox.com/Configuration/agent_managemen')
        print("切換代理商管理選單!", '\n')
        time.sleep(1)
        

#  ************************** 新增代理商/幣別 **************************

    # def Configuration_Agent_Management_append_agent(self):
            
        # self.driver.find_element(By.ID, f'agent_setting_btn').click()  # 新增
        # time.sleep(1)
        
        # show_currency = self.driver.find_element(By.ID, f'Add_currency')  # 顯示幣別
        # # show_currency.click()
        # show_currency_value = show_currency.get_attribute('value')
        # if show_currency_value == '':
        #     print('顯示幣別=', 'No value!', '\n')
        #     time.sleep(1)
        
        # account_currency = self.driver.find_element(By.ID, f'Add_account_currency')  # 帳戶幣別
        # # account_currency.click()
        # account_currency_value = account_currency.get_attribute('value')
        # if account_currency_value == '':
        #     print('帳戶幣別=', 'No value!', '\n')
        #     time.sleep(1)
        # time.sleep(1)
        
        # self.driver.find_element(By.ID, f'Add_agent_id').click()  # 代理商I.D
        # self.driver.find_element(By.ID, f'Add_agent_id').send_keys('ivanTest_admin')  # ivanTest_admin
        # time.sleep(1)
        
        # self.driver.find_element(By.ID, f'Add_host_name').click()  # 代理商名稱
        # self.driver.find_element(By.ID, f'Add_host_name').send_keys('ivanTest_admin')  # ivanTest_admin
        # time.sleep(1)
        
        # self.driver.find_element(By.ID, f'Add_domain').click()  # 網域
        # self.driver.find_element(By.XPATH, f' //*[@id="Add_domain"]/option[@value="74118c56985fce12d2edd07a6262e50b"]').click()  # iplaystar.net
        # time.sleep(1)
        
        # self.driver.find_element(By.ID, f'Add_prefix').click()  # 前綴
        # self.driver.find_element(By.ID, f'Add_prefix').send_keys('DEV-API')  # DEV-API
        # time.sleep(1)
        
        # self.driver.find_element(By.ID, f'Add_timezone').click()  # 時區
        # self.driver.find_element(By.ID, f'Add_timezone').send_keys('+08:00')  # +08:00
        # time.sleep(1)
        
        # self.driver.find_element(By.ID, f'Add_domain_group_id').click()  # Domain Group ID
        # self.driver.find_element(By.XPATH, f'//*[@id="Add_domain_group_id"]/option[@value="1"]').click()  # 1
        # time.sleep(1)
        
        # show_currency_1 = self.driver.find_element(By.ID, f'Add_currency')  # 顯示幣別
        # show_currency_1.click()
        # show_currency_slt = self.driver.find_element(By.XPATH, f'//*[@id="Add_currency"]/option[@value="USD"]')  # USD
        # show_currency_slt.click()
        # show_currency_value_1 = show_currency_slt.get_attribute('value')
        # print('顯示幣別=', show_currency_value_1, '\n')       
        # time.sleep(2)
            
        # account_currency_1 = self.driver.find_element(By.ID, f'Add_account_currency')  # 帳戶幣別
        # account_currency_1.click()
        # account_currency_slt = self.driver.find_element(By.XPATH, f'//*[@id="Add_account_currency"]/option[@value="USD"]')  # USD
        # account_currency_slt.click()
        # account_currency_value_1 = account_currency_slt.get_attribute('value')
        # print('帳戶幣別=', account_currency_value_1, '\n')       
        # time.sleep(2)
                    
        # self.driver.find_element(By.ID, f'a_confirm_submit').click()  # 送出
        # time.sleep(2)
        
        # show_info = self.driver.find_element(By.XPATH, f'/html/body/div[2]')  # 成功
        # print('代理商新增:', show_info.text, '\n')
        # time.sleep(2)
        
        # self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0_filter"]/label/input').click()  # 篩選
        # self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0_filter"]/label/input').send_keys('ivantest')  # ivanTest_admin
        # time.sleep(1)
        
        # self.driver.find_element(By.NAME, f'update').click()  # 編輯
        # time.sleep(1)
        
        # agent_id = self.driver.find_element(By.ID, f'Add_agent_id')  # 代理商I.D
        # self.driver.execute_script("arguments[0].scrollIntoView();", agent_id)
        # print('代理商I.D:', agent_id.get_attribute('value'), '\n')
        # time.sleep(1)
        
        # agent_name = self.driver.find_element(By.ID, f'Add_host_name')  # 代理商名稱
        # self.driver.execute_script("arguments[0].scrollIntoView();", agent_name)
        # print('代理商名稱:', agent_name.get_attribute('value'), '\n')
        # time.sleep(1)
        
        # show_currency_2 = self.driver.find_element(By.ID, f'Add_currency')  # 顯示幣別
        # show_currency_2.click()
        # show_currency_value_2 = show_currency_2.get_attribute('value')
        # print('顯示幣別=', show_currency_value_2, '\n')
        # time.sleep(2)
        
        # account_currency_2 = self.driver.find_element(By.ID, f'Add_account_currency')  # 帳戶幣別
        # account_currency_2.click()
        # show_currency_value_2 = account_currency_2.get_attribute('value')
        # print('帳戶幣別=', show_currency_value_2, '\n')
        # time.sleep(2)
        
        # self.driver.refresh() 
    
    
    # #  ************************** 批量新增代理商/幣別 **************************
    
    # def Configuration_Agent_Management_batch_append_agent(self):
            
    #     self.driver.find_element(By.ID, f'agent_setting_batch_btn').click()  # 批量新增
    #     time.sleep(1)
        
    #     for _ in range(3):    
    #         self.driver.find_element(By.CLASS_NAME, f'icon-plus3').click()  # "+"  批量新增4名代理商
    #         time.sleep(1)
        
    #     agent_id_values = ['ivanTest_001_', 'ivanTest_002_', 'ivanTest_003_', 'ivanTest_004_']
              
    #     for id_value in range(4):        
    #         agent_id_box = self.driver.find_elements(By.XPATH, '//*[@id="Add_agent_id"]')  # 代理商I.D
    #         agent_id_box[id_value].click()
    #         time.sleep(1)
             
    #         agent_id_box[id_value].send_keys(agent_id_values[id_value])
    #         print('代理商I.D:', agent_id_box[id_value].get_attribute('value'), '\n')
    #         time.sleep(1)
    
    #     agent_name_values = ['ivanTest_001_', 'ivanTest_002_', 'ivanTest_003_', 'ivanTest_004_']
                 
    #     for name_value in range(4):    
    #         agent_name_box = self.driver.find_elements(By.XPATH, '//*[@id="Add_host_name"]')  # 代理商名稱
    #         agent_name_box[name_value].click()
    #         time.sleep(1)
            
    #         agent_name_box[name_value].send_keys(agent_name_values[name_value])
    #         print('代理商名稱:', agent_name_box[name_value].get_attribute('value'), '\n')
    #         time.sleep(1)

    #     for domain_value in range(4):
    #         domain_values = self.driver.find_elements(By.ID, f'Add_domain')  # 網域
    #         domain_values[domain_value].click()
    #         time.sleep(1)
            
    #         domain_value_slt = self.driver.find_elements(By.XPATH, f'//*[@id="Add_domain"]/option[@value="74118c56985fce12d2edd07a6262e50b"]')  # iplaystar.net
    #         domain_value_slt[domain_value].click()  
    #         print('網域:', domain_value_slt[domain_value].get_attribute('value'), '\n')
    #         time.sleep(1)
            
    #     for prefix_value in range(4):     
    #         prefix_values = self.driver.find_elements(By.ID, f'Add_prefix')  # 前綴
    #         prefix_values[prefix_value].click()
    #         time.sleep(1)
            
    #         prefix_values[prefix_value].send_keys('DEV-API')  # DEV-API
    #         print('前綴:', prefix_values[prefix_value].get_attribute('value'), '\n')
    #         time.sleep(1)
        
    #     for timezone_value in range(4): 
    #         timezone_values = self.driver.find_elements(By.ID, f'Add_timezone')  # 時區
    #         timezone_values[timezone_value].click() 
    #         time.sleep(1)
            
    #         timezone_values[timezone_value].send_keys('+08:00')  # +08:00
    #         print('時區:', timezone_values[timezone_value].get_attribute('value'), '\n')
    #         time.sleep(1)     
            
    #     for domain_group_id_value in range(4):     
    #         domain_group_id_values = self.driver.find_elements(By.ID, f'Add_domain_group_id')  # Domain Group ID
    #         domain_group_id_values[domain_group_id_value].click()
    #         time.sleep(1)
            
    #         domain_group_id_slt = self.driver.find_elements(By.XPATH, f'//*[@id="Add_domain_group_id"]/option[@value="1"]')  # 1
    #         domain_group_id_slt[domain_group_id_value].click()
    #         print('Domain Group ID:', domain_group_id_slt[domain_group_id_value].get_attribute('value'), '\n')
    #         time.sleep(1)   
            
    #     self.driver.find_element(By.ID, f'Add_currency').click()  # 顯示幣別
    #     time.sleep(1)
        
    #     show_currency_slt = self.driver.find_element(By.XPATH, f'//*[@id="Add_currency"]/option[@value="USD"]')  # USD
    #     show_currency_slt.click()
        
    #     show_currency_value_1 = show_currency_slt.get_attribute('value')
    #     print('顯示幣別=', show_currency_value_1, '\n')       
    #     time.sleep(1)
        
    #     self.driver.find_element(By.ID, f'Add_account_currency').click()  # 帳戶幣別
    #     time.sleep(1)
        
    #     account_currency_slt = self.driver.find_element(By.XPATH, f'//*[@id="Add_account_currency"]/option[@value="USD"]')  # USD
    #     account_currency_slt.click()
        
    #     account_currency_value_1 = account_currency_slt.get_attribute('value')
    #     print('帳戶幣別=', account_currency_value_1, '\n')       
    #     time.sleep(1)
        
    #     self.driver.find_element(By.ID, f'a_confirm_submit').click()  # 送出
    #     time.sleep(2)
        
    #     self.driver.find_element(By.XPATH, f'/html/body/div[2]')  # 成功
    #     print('代理商批量新增:', self.driver.find_element(By.XPATH, f'/html/body/div[2]').text, '\n')
    #     time.sleep(2)
        
        # self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0_filter"]/label/input').click()  # 篩選
        # self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0_filter"]/label/input').send_keys('ivanTest')  # ivanTest
        # time.sleep(1)
        
        # self.driver.find_element(By.NAME, f'update').click()  # 編輯
        # time.sleep(1)
                
        # show_currency_2 = self.driver.find_element(By.ID, f'Add_currency')  # 顯示幣別
        # show_currency_2.click()
        # show_currency_value_2 = show_currency_2.get_attribute('value')
        # print('顯示幣別=', show_currency_value_2, '\n')
        # time.sleep(1)
        
        # account_currency_2 = self.driver.find_element(By.ID, f'Add_account_currency')  # 帳戶幣別
        # account_currency_2.click()
        # show_currency_value_2 = account_currency_2.get_attribute('value')
        # print('帳戶幣別=', show_currency_value_2, '\n')
        # time.sleep(1)
        
        self.driver.refresh()
    



        self.driver.close()
        
        end_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
        print('測試結束時間: ', end_time, '\n')
        





if __name__ == '__main__':
    
    report_path = r'D:\AutomotiveTest_QA\Back_platform Test\v 3.0.3.1756 250424\Report'
                            
    suite = unittest.TestSuite()
    suite.addTest(PS_platform_testCase('PS_platform_access'))  # PS後台連接測試
    suite.addTest(PS_platform_testCase('PS_platform_logging'))  # PS後台登入測試
    suite.addTest(PS_platform_testCase('Configuration_Agent_Management_zhTW'))  # 代理商管理
    suite.addTest(PS_platform_testCase('Configuration_Agent_Management_batch_append_agent'))  # 批量新增代理商
    
            
    
    

    now_time = time.strftime("%Y-%m-%d_%H_%M_%S", time.localtime(time.time())) 
    report_title = now_time + "_PS後台代理商管理(新增代理商)功能巡測測試報告.html"
    result_path = os.path.join(report_path, report_title)

    desc = "PS後台代理商管理(新增代理商)功能巡測測試報告"

    file_name = open(result_path, 'wb')
    runners = HTMLTestRunner.HTMLTestRunner(stream = file_name, title = '測試資訊', description = desc)
    runners.run(suite)
    file_name.close()
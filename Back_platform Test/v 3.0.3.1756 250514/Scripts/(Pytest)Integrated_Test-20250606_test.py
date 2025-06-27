# 套件匯入
import openpyxl.workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
import pyautogui as pag
from selenium.webdriver.chrome.service import Service
import requests
from http import HTTPStatus
from selenium.webdriver.chrome.options import Options
from lxml import html
import csv
import numpy as np
import pandas as pd
import cv2
import pybi as pbi
import os
import sys
import aspose
import xlrd
import xlwt
import glob
import setuptools
import jpype
import difflib
import openpyxl
from openpyxl import Workbook
from openpyxl import workbook
from openpyxl import worksheet
from openpyxl.styles import Font  # 導入字體模組
from openpyxl.styles import PatternFill  # 導入填充模組
from spire.xls import *
from collections import deque
from pandas.core.frame import DataFrame
# 导入openpyxl模块并将其重命名为pxl
import openpyxl as pxl
# 从openpyxl导入PatternFill类
from openpyxl.styles import PatternFill
# 从openpyxl导入colors类
from openpyxl.styles import colors
# 从openpyxl导入Font类
from openpyxl.styles import Font
import datetime
import pytesseract
from PIL import Image
import xlwings
import getpass
from pathlib import Path
import matplotlib.pyplot as plt
from PIL import Image, ImageEnhance
import re
from bs4 import BeautifulSoup
import pykeyboard
from selenium.common.exceptions import NoAlertPresentException
from selenium.common.exceptions import UnexpectedAlertPresentException
import unittest, doctest
import HTMLTestRunner
from BeautifulReport import BeautifulReport
import pytest



# # ================================== 建立文件放置資料夾 ==================================

# path = r'D:/'
# p = Path(path)

# # 定義子資料夾名稱
# sub_folder = 'AutomotiveTest_QA'
# # 連接主資料夾與子資料夾
# p = p.joinpath(sub_folder)
# p.mkdir(exist_ok=True)

# # 定義子資料夾名稱
# sub_folder_1 = 'Back_platform Test'
# # 連接主資料夾與子資料夾
# p = p.joinpath(sub_folder_1)
# p.mkdir(exist_ok=True)

# # 定義子資料夾名稱
# sub_folder_2 = 'v 3.0.3.1756 250514'
# # 連接主資料夾與子資料夾
# p = p.joinpath(sub_folder_2)
# p.mkdir(exist_ok=True)

# # 定義子資料夾名稱
# sub_folder_3 = 'Report'
# # 連接主資料夾與子資料夾
# p = p.joinpath(sub_folder_3)
# p.mkdir(exist_ok=True)


# ============================= PS後台連接與登入 ============================= 

# for img_no in range(1, 100):

# class PS_platform_testCase(unittest.TestCase):
class Test_PS_platform_testCase():
    '''PS Platform Test'''
    
# driver = webdriver.Chrome()

    options = webdriver.ChromeOptions()
    options.add_argument('--ignore-ssl-errors=yes')
    options.add_argument('--ignore-certificate-errors')
    driver = webdriver.Chrome(options=options)

    def test_PSweb_driver_opened(self):
        '''開啟PS後台自動化測試瀏覽器'''
        print('測試開始! PS後台瀏覽器已開啟。')
        time.sleep(1)
        
        current_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
        print('測試起始時間: ', current_time, '\n')
        
        print('PS後台連線測試...', '\n')
        url = 'https://dev-admin-br-02.iplaystar.net'
        # self.driver = webdriver.Chrome()
        self.driver.get(url)
        time.sleep(1)
        self.req = requests.get('https://dev-admin-br-02.iplaystar.net')
        req = requests.post(url)
        state_code = req.status_code
        print('Http Response Code: ', state_code, '\n')
        if state_code == 200:
            print("HTTP回應成功!", '\n')
        else:
            print('Http Response Code:', state_code, '\n')
            print("HTTP回應失敗!", '\n')   
            
        self.driver.find_element(By.XPATH, '/html/body/div[2]/div/form/label').click()
        self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[2]').click()
        print("語系已切換'繁體中文'!", '\n')
        time.sleep(1)
        
    # def test_tearDown(self):
    #     '''截圖並保存'''
        
    #     self.driver.save_screenshot('screenshot' + {img_no} + '.png')
    #     img_no += 1
        
    def test_PS_platform_logging(self):
        '''PS後台連線測試'''

        ws = self.driver.window_handles[0]
        self.driver.switch_to.window(ws)
        time.sleep(1)
        
        self.driver.find_element(By.CLASS_NAME, 'content-group').click()  # 尋找登入介面元素位址

        self.driver.find_element(By.XPATH, '//*[@id="user_id"]').clear()  # 預設此欄位為null, 但仍先清除帳號欄位資訊
        self.driver.find_element(By.XPATH, '//*[@id="password"]').clear()  # 預設此欄位為null, 但仍先清除密碼欄位資訊
        self.driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div/form/div/div[4]/input')  # 驗證碼欄位
        self.driver.find_element(By.CSS_SELECTOR, '#captcha_img > img').screenshot('verificationCode.png')  # 驗證碼截圖存檔
        time.sleep(1)
        
        accountValue = getpass.getpass('請輸入後台帳號: ')
        self.driver.find_element(By.XPATH, '//*[@id="user_id"]').send_keys(accountValue)   # 輸入使用者帳號
        time.sleep(1)
        self.driver.find_element(By.XPATH, '//*[@id="user_id"]').send_keys(Keys.TAB)  # 切換至密碼輸入欄位
        
        pwdValue = getpass.getpass('請輸入後台密碼: ')
        self.driver.find_element(By.XPATH, '//*[@id="password"]').send_keys(pwdValue)  # 輸入個人密碼
        time.sleep(1)
        self.driver.find_element(By.XPATH, '//*[@id="password"]').send_keys(Keys.TAB)   # 切換至驗證碼輸入欄位
        
        verifiCodeValue = input('請輸入驗證碼: ')  # 輸入驗證碼
        # print(verifiCodeValue, '\n')
        self.driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div/form/div/div[4]/input').send_keys(verifiCodeValue)
        self.driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div/form/div/div[6]/button').click()
        time.sleep(1)
        self.driver.find_element(By.XPATH, '/html/body/div[3]/div/div[4]/div[1]/div/div/h4/span')
        
        try:
            element = self.driver.find_element(By.XPATH, '/html/body/div[3]/div/div[4]/div[1]/div/div/h4/span')
            element.click()
            str_error = None
            print("登入成功!", '\n')
            time.sleep(1)
            os.remove('verificationCode.png')
        except Exception as e:
            print(e)
            str_error = True
            pass

        if str_error:
            print("登入資訊輸入錯誤，請重新輸入!", '\n')
            self.driver.refresh()
                
        # def test_tearDown(self):
        #     '''截圖並保存'''
            
        #     self.driver.save_screenshot('screenshot' + {img_no} + '.png')
        #     img_no += 1


# # ============================= 後台功能巡測【帳務】============================= 
# # *************************** 遊戲績效 ***************************

#     def test_Accounting_Game_Performance_zhCN(self):
#         '''【帳務】遊戲績效功能頁切換'''

#         self.driver.find_element(By.ID, 'Accounting').click()
#         print("進入帳務功能選單!", '\n')
#         time.sleep(2)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Accounting/game_performance')
#         print("切換遊戲績效選單!", '\n')
#         time.sleep(1)
        
#         self.driver.refresh()
        
    
# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_gp_zhCN(self):
#         '''搜尋列功能驗證'''
                

# # ************************** 起始時間 **************************
    
#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button/i').click()  # 起始時間
#         time.sleep(1)
        
#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[6]/td[1]/a').click()  # 2024/12/30

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('起始時間:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)   
        

# # ************************** 結束時間 **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i').click()  # 結束時間
#         time.sleep(1)

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[5]/a').click()  # 2025/01/31
     
#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('結束時間:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)   

#         self.driver.refresh()


# #  ************************** 搜尋類別(遊戲) **************************

#         self.driver.find_element(By.ID, 'search_class').click()  # 搜尋類別
#         time.sleep(1)
        
#         for serachType in range(1, 5):
#             element = self.driver.find_element(By.XPATH, f'//*[@id="search_class"]/option[{serachType}]')  # 遊戲/玩家/代理商/每日   
#             element.click()
#             print('搜尋類別:', element.text, '\n')
#             time.sleep(1)

#         self.driver.refresh()


# #  ************************** 代理商類別 **************************

#         agent_class_list = ['All', 'PS', 'Test']
        
#         for agentType_slt in agent_class_list:
#             agentType = self.driver.find_element(By.ID, 'agent_attr')  # 代理商類別
#             agentType.click()
#             time.sleep(1)
            
#             agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="{agentType_slt}"]')  # -- 全選 -- / PS / Test
#             agentTypeSelect.click()
#             print('代理商類別:', agentTypeSelect.text, '\n')
#             time.sleep(1)

#         self.driver.refresh()
        

# #  ************************** 帳戶幣別 **************************

#         self.driver.find_element(By.ID, f'currency_search').click()  # 帳戶幣別
#         time.sleep(1)

#         account_type_list = ['CNY', 'IDR', 'TWD']
        
#         for account_type_slt in account_type_list:    
#             accountTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="currency_search"]/option[@value="{account_type_slt}"]')  # CNY / IDR / TWD
#             accountTypeSelect.click()
#             print('帳戶幣別:', accountTypeSelect.text, '\n')
#             time.sleep(1)

#         self.driver.refresh()


# #  ************************** 遊戲績效 **************************

#     def test_game_performance_search_func_identify_zhCN(self):
#         '''遊戲績效搜尋功能驗證'''

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button/i').click()  # 起始時間
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[6]/td[1]/a').click()  # 2024/12/30
#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('起始時間:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)   

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i').click()  # 結束時間
#         time.sleep(1) 

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[5]/a').click()  # 2025/01/31
#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('結束時間:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent_attr').click()  # 代理商類別
#         time.sleep(1)
        
#         agentTypeSelect = self.driver.find_element(By.XPATH, '//*[@id="agent_attr"]/option[@value="Test"]')  # 代理商類別 Test
#         agentTypeSelect.click()
#         print('代理商類別:', agentTypeSelect.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'currency_search').click()  # 帳戶幣別
#         time.sleep(1)

#         accountTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="currency_search"]/option[@value="CNY"]')  # CNY
#         accountTypeSelect.click()
#         print('帳戶幣別:', accountTypeSelect.text, '\n')
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="select_all"]')  # 代理商 全選
#         agent_select.click()
#         print('代理商:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="sh_btn"]').click()  # 送出
#         time.sleep(1)

#         slot_game = self.driver.find_element(By.XPATH, '//*[@id="tb_game"]/div[11]/div/div[1]/h5')
#         self.driver.execute_script("arguments[0].scrollIntoView();", slot_game)  # 老虎機
#         time.sleep(2)
            
#         cardGame = self.driver.find_element(By.XPATH, '//*[@id="tb_game"]/div[12]/div/div[1]/h5')
#         self.driver.execute_script("arguments[0].scrollIntoView();", cardGame)  # 棋牌遊戲
#         time.sleep(2)

#         exportIcon = self.driver.find_element(By.XPATH, '//*[@id="export"]') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", exportIcon) # 頁首
#         time.sleep(1)
        
#         self.driver.refresh()


# #  ************************** 營運狀態 **************************

#     def test_operational_status_search_func_identify_zhCN(self):
#         '''營運狀態搜尋功能驗證'''

#         self.driver.find_element(By.ID, 'agent_attr').click()  # 代理商類別
#         time.sleep(1)
        
#         agentTypeSelect = self.driver.find_element(By.XPATH, '//*[@id="agent_attr"]/option[@value="Test"]')  # 代理商類別 Test
#         agentTypeSelect.click()
#         print('代理商類別:', agentTypeSelect.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'status_search').click()  # 營運狀態
#         time.sleep(1)

#         for open_states_code in (1, 3):
#             openStates = self.driver.find_element(By.XPATH, f'//*[@id="status_search"]/option[@value="{open_states_code}"]')  # 啟用(已營運) / 啟用(未營運)
#             openStates.click()
#             print('營運狀態:', openStates.text, '\n')
#             time.sleep(1)

#             agent_select = self.driver.find_element(By.XPATH, f'//*[@id="select_all"]')  # 代理商 全選
#             agent_select.click()
#             print('代理商:', agent_select.get_attribute('value'), '\n')

#             self.driver.find_element(By.XPATH, f'//*[@id="sh_btn"]').click()  # 送出
#             time.sleep(1)

#             slot_game = self.driver.find_element(By.XPATH, '//*[@id="tb_game"]/div[11]/div/div[1]/h5')
#             self.driver.execute_script("arguments[0].scrollIntoView();", slot_game)  # 老虎機
#             time.sleep(2)
                
#             cardGame = self.driver.find_element(By.XPATH, '//*[@id="tb_game"]/div[12]/div/div[1]/h5')
#             self.driver.execute_script("arguments[0].scrollIntoView();", cardGame)  # 棋牌遊戲
#             time.sleep(2)
                
#             exportIcon = self.driver.find_element(By.XPATH, '//*[@id="export"]') 
#             self.driver.execute_script("arguments[0].scrollIntoView();", exportIcon) # 頁首
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)
        

# # --------------------------- 遊戲績效(EN) ---------------------------

#     def test_Accounting_Game_Performance_EN(self):
#         '''【帳務】遊戲績效語系切換(英)'''
    
#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()  
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[1]').click()  # EN
#         print("語系已切換'English'!", '\n')
#         time.sleep(1)


# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_gp_EN(self):
#         '''搜尋列功能驗證(英)'''
        

# # ************************** 起始時間 **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button/i').click()  # Start Time

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[2]/a').click()  # 2024/12/30
#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('Start Time:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)   
        

# # ************************** 結束時間 **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i').click()  # End Time

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[6]/a').click()  # 2025/01/31
#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('End Time:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)   

#         self.driver.refresh()


# #  ************************** 搜尋類別 **************************

#         self.driver.find_element(By.ID, 'search_class').click()  # search_class
#         time.sleep(1)  

#         for serachType in range(1, 5):
#             element = self.driver.find_element(By.XPATH, f'//*[@id="search_class"]/option[{serachType}]')  # search_class Game/Player/Agent/Daily 
#             element.click()
#             print('Search Class:', element.text, '\n')
#             time.sleep(2)

#         self.driver.refresh()


# #  ************************** 代理商類別 **************************

#         agent_class_list = ['All', 'PS', 'Test']
#         for agentType_slt in agent_class_list:
#             self.driver.find_element(By.ID, 'agent_attr').click()  # Agent Class
#             time.sleep(1)
            
#             agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="{agentType_slt}"]')  # -- Select All -- / PS / Test
#             agentTypeSelect.click()
#             print('Agent Class:', agentTypeSelect.text, '\n')
#             time.sleep(1)

#         self.driver.refresh()


# #  ************************** 帳戶幣別 **************************

#         self.driver.find_element(By.ID, f'currency_search').click()  # Account currency
#         time.sleep(1)

#         account_type_list = ['CNY', 'IDR', 'TWD']
        
#         for account_type_slt in account_type_list:
#             accountTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="currency_search"]/option[@value="{account_type_slt}"]')  # CNY / IDR / TWD
#             accountTypeSelect.click()
#             print('Account currency:', accountTypeSelect.text, '\n')
#             time.sleep(1)

#         self.driver.refresh()


# #  ************************** 遊戲績效 **************************

#     def test_game_performance_search_func_identify_EN(self):
#         '''遊戲績效搜尋功能驗證(英)'''

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button/i').click()  # Start Time
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[2]/a').click()  # 2024/12/30
#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('Start Time:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)    

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i').click()  # End Time
#         time.sleep(1)

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[6]/a').click()  # 2025/01/31
#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('End Time:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)  

#         self.driver.find_element(By.ID, 'agent_attr').click()  # Agent Class
#         time.sleep(1)
        
#         agentTypeSelect = self.driver.find_element(By.XPATH, '//*[@id="agent_attr"]/option[@value="Test"]')  # Test
#         agentTypeSelect.click()
#         print('Agent Class:', agentTypeSelect.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'currency_search').click()  # Account currency
#         time.sleep(1)

#         accountTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="currency_search"]/option[@value="CNY"]')  # CNY
#         accountTypeSelect.click()
#         print('Account currency:', accountTypeSelect.text, '\n')
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="select_all"]')  # All
#         agent_select.click()
#         print('Agent:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="sh_btn"]').click()  # Submit    
#         time.sleep(1)

#         slot_game = self.driver.find_element(By.XPATH, '//*[@id="tb_game"]/div[11]/div/div[1]/h5')
#         self.driver.execute_script("arguments[0].scrollIntoView();", slot_game)  # 老虎機
#         time.sleep(2)
            
#         cardGame = self.driver.find_element(By.XPATH, '//*[@id="tb_game"]/div[12]/div/div[1]/h5')
#         self.driver.execute_script("arguments[0].scrollIntoView();", cardGame)  # 棋牌遊戲
#         time.sleep(2)
            
#         exportIcon = self.driver.find_element(By.XPATH, '//*[@id="export"]') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", exportIcon) # 頁首
#         time.sleep(1)
        
#         self.driver.refresh()


# #  ************************** 營運狀態 **************************

#     def test_operational_status_search_func_identify_EN(self):
#         '''營運狀態搜尋功能驗證(英)'''

#         self.driver.find_element(By.ID, 'agent_attr').click()  # Agent Class
#         time.sleep(1)
        
#         agentTypeSelect = self.driver.find_element(By.XPATH, '//*[@id="agent_attr"]/option[@value="Test"]')  # Test
#         agentTypeSelect.click()
#         print('Agent Class:', agentTypeSelect.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'status_search').click()  # Operational Status
#         time.sleep(1)

#         for open_states_code in (1, 3):
#             openStates = self.driver.find_element(By.XPATH, f'//*[@id="status_search"]/option[@value="{open_states_code}"]')  # Enabled (Operating) / Enabled (Non-Operating)
#             openStates.click()
#             print('Operational Status:', openStates.text, '\n')
#             time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="select_all"]')  # Select All
#         agent_select.click()
#         print('Agent:', agent_select.get_attribute('value'), '\n')

#         self.driver.find_element(By.XPATH, f'//*[@id="sh_btn"]').click()  # Submit 
#         time.sleep(1)

#         slot_game = self.driver.find_element(By.XPATH, '//*[@id="tb_game"]/div[11]/div/div[1]/h5')
#         self.driver.execute_script("arguments[0].scrollIntoView();", slot_game)  # 老虎機
#         time.sleep(2)
            
#         cardGame = self.driver.find_element(By.XPATH, '//*[@id="tb_game"]/div[12]/div/div[1]/h5')
#         self.driver.execute_script("arguments[0].scrollIntoView();", cardGame)  # 棋牌遊戲
#         time.sleep(2)
            
#         exportIcon = self.driver.find_element(By.XPATH, '//*[@id="export"]') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", exportIcon) # 頁首
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)
        

# # --------------------------- 遊戲績效(Tai) ---------------------------

#     def test_Accounting_Game_Performance_Tai(self):
#         '''【帳務】遊戲績效語系切換(泰)'''
    
#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[4]').click()
#         time.sleep(1)
#         print("語系已切換'ไทย'!", '\n')
        

# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_gp_Tai(self):
#         '''搜尋列功能驗證(泰)'''


# # ************************** 起始時間 **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button/i').click()  # เวลาเริ่มต้น
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[2]/a').click()  # 2024/12/30
#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('เวลาเริ่มต้น:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)
           

# # ************************** 結束時間 **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i').click()  # เวลาสิ้นสุด
#         time.sleep(1)

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[6]/a').click()  # 2025/01/31
#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('เวลาสิ้นสุด:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)   

#         self.driver.refresh()


# #  ************************** 搜尋類別 **************************

#         self.driver.find_element(By.ID, 'search_class').click()  # หมวดหมู่การค้นหา
#         time.sleep(1)

#         for serachType in range(1, 5):
#             element = self.driver.find_element(By.XPATH, f'//*[@id="search_class"]/option[{serachType}]')  # เกม/ผู้เล่น/ตัวแทน/ทุกวัน   
#             element.click()
#             print('หมวดหมู่การค้นหา:', element.text, '\n')
#             time.sleep(1)

#         self.driver.refresh()


# #  ************************** 代理商類別 **************************

#         agent_class_list = ['All', 'PS', 'Test']

#         for agentType_slt in agent_class_list:
#             self.driver.find_element(By.ID, 'agent_attr').click()  # หมวดตัวแทน
            
#             time.sleep(1)
#             agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="{agentType_slt}"]')  # -- เลือกทั้งหมด -- / PS / Test
#             agentTypeSelect.click()
#             print('หมวดตัวแทน:', agentTypeSelect.text, '\n')
#             time.sleep(1)

#         self.driver.refresh()


# #  ************************** 帳戶幣別 **************************

#         self.driver.find_element(By.ID, f'currency_search').click()  # สกุลเงินในบัญชี
#         time.sleep(1)

#         account_type_list = ['CNY', 'IDR', 'TWD']
        
#         for account_type_slt in account_type_list:
#             accountTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="currency_search"]/option[@value="{account_type_slt}"]')  # CNY / IDR / TWD
#             accountTypeSelect.click()
#             print('สกุลเงินในบัญชี:', accountTypeSelect.text, '\n')
#             time.sleep(2)

#         self.driver.refresh()


# #  ************************** 遊戲績效 **************************

#     def test_game_performance_search_func_identify_Tai(self):
#         '''遊戲績效搜尋功能驗證(泰)'''

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button/i').click()  # เวลาเริ่มต้น
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[2]/a').click()  # 2024/12/30
#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('เวลาเริ่มต้น:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)    

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i').click()  # เวลาสิ้นสุด
#         time.sleep(1)

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[6]/a').click()  # 2025/01/31
#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('เวลาสิ้นสุด:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)  

#         self.driver.find_element(By.ID, 'agent_attr').click()  # หมวดตัวแทน
#         time.sleep(1)
        
#         agentTypeSelect = self.driver.find_element(By.XPATH, '//*[@id="agent_attr"]/option[@value="Test"]')  # Test
#         agentTypeSelect.click()
#         print('หมวดตัวแทน:', agentTypeSelect.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'currency_search').click()  # สกุลเงินในบัญชี
#         time.sleep(1)

#         accountTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="currency_search"]/option[@value="CNY"]')  # CNY
#         accountTypeSelect.click()
#         print('สกุลเงินในบัญชี:', accountTypeSelect.text, '\n')
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="select_all"]')  # All
#         agent_select.click()
#         print('ตัวแทน:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="sh_btn"]').click()  # ส่ง 
#         time.sleep(1)

#         slot_game = self.driver.find_element(By.XPATH, '//*[@id="tb_game"]/div[11]/div/div[1]/h5')
#         self.driver.execute_script("arguments[0].scrollIntoView();", slot_game)  # 老虎機
#         time.sleep(2)
            
#         cardGame = self.driver.find_element(By.XPATH, '//*[@id="tb_game"]/div[12]/div/div[1]/h5')
#         self.driver.execute_script("arguments[0].scrollIntoView();", cardGame)  # 棋牌遊戲
#         time.sleep(2)
            
#         exportIcon = self.driver.find_element(By.XPATH, '//*[@id="export"]') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", exportIcon) # 頁首
#         time.sleep(1)
        
#         self.driver.refresh()


# #  ************************** 營運狀態 **************************

#     def test_operational_status_search_func_identify_Tai(self):
#         '''營運狀態搜尋功能驗證(泰)'''

#         self.driver.find_element(By.ID, 'agent_attr').click()  # หมวดตัวแทน
#         time.sleep(1)
        
#         agentTypeSelect = self.driver.find_element(By.XPATH, '//*[@id="agent_attr"]/option[@value="Test"]')  # Test
#         agentTypeSelect.click()
#         print('หมวดตัวแทน:', agentTypeSelect.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'status_search').click()  # สถานะการบริการ
#         time.sleep(1)

#         for open_states_code in (1, 3):
#             openStates = self.driver.find_element(By.XPATH, f'//*[@id="status_search"]/option[@value="{open_states_code}"]')  # เปิดใช้งาน (ให้เปิดบริการแล้ว) / เปิดใช้งาน (ไม่ได้เปิดให้บริการ)
#             openStates.click()
#             print('สถานะการบริการ:', openStates.text, '\n')
#             time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="select_all"]')  # เลือกทั้งหมด
#         agent_select.click()
#         print('ตัวแทน:', agent_select.get_attribute('value'), '\n')

#         self.driver.find_element(By.XPATH, f'//*[@id="sh_btn"]').click()  # ส่ง  
#         time.sleep(1)

#         slot_game = self.driver.find_element(By.XPATH, '//*[@id="tb_game"]/div[11]/div/div[1]/h5')
#         self.driver.execute_script("arguments[0].scrollIntoView();", slot_game)  # 老虎機
#         time.sleep(2)
            
#         cardGame = self.driver.find_element(By.XPATH, '//*[@id="tb_game"]/div[12]/div/div[1]/h5')
#         self.driver.execute_script("arguments[0].scrollIntoView();", cardGame)  # 棋牌遊戲
#         time.sleep(2)
            
#         exportIcon = self.driver.find_element(By.XPATH, '//*[@id="export"]') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", exportIcon) # 頁首
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# # ============================= 後台功能巡測【帳務】============================= 
# # *************************** 營運報表 ***************************

#     def test_Accounting_Operating_Statement_zhCN(self):
#         '''【帳務】營運報表功能頁切換''' 

#         # self.driver.get('https://dev-admin-br-02.iplaystar.net/Accounting/operating_statement')
#         # print("切換營運報表選單!", '\n')
#         # time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[3]').click()
#         print("語系已切換'简体中文'!", '\n')
        
#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Accounting/operating_statement')
#         print("切換營運報表選單!", '\n')
#         time.sleep(1)
        

# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_os_zhCN(self):
#         '''搜尋列功能驗證'''
                

# # ************************** 起始時間 **************************

#         self.driver.find_element(By.XPATH, '//*[@id="queryarea"]/div/div/div[2]/div[2]/div[1]/div/div/span/button/i').click()  # 起始時間
#         time.sleep(1)

#         for _ in range(7):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[6]/a').click()  # 2024/11/30
#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('起始時間:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)  
 

# # ************************** 結束時間 **************************

#         self.driver.find_element(By.XPATH, '//*[@id="queryarea"]/div/div/div[2]/div[2]/div[2]/div/div/span/button/i').click()  # 結束時間
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[6]/td[1]/a').click()  # 2024/12/30
#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('結束時間:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)   

#         self.driver.refresh()


# #  ************************** 帳戶幣別 **************************

#         self.driver.find_element(By.ID, f'currency_search').click()  # 帳戶幣別
#         time.sleep(1)        
        
#         currency_list = ['KRW', 'THB', 'USD']
        
#         for account_id in currency_list:
#             accountTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="currency_search"]/option[@value="{account_id}"]')  # KRW / THB / USD
#             accountTypeSelect.click()
#             print('帳戶幣別:', accountTypeSelect.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()


# #  ************************** 代理商類別 **************************

#         self.driver.find_element(By.ID, 'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_list = ['All', 'C66', 'Platform']
        
#         for agent_id in agent_class_list:    
#             agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="{agent_id}"]')  # -- 全選 -- / C66 / Platform
#             agentTypeSelect.click()
#             print('代理商類別:', agentTypeSelect.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 營運報表 **************************

#     def test_operating_statement_search_func_identify(self):
#         '''營運報表搜尋功能驗證'''      

#         self.driver.find_element(By.XPATH, '//*[@id="queryarea"]/div/div/div[2]/div[2]/div[1]/div/div/span/button/i').click()  # 起始時間
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[6]/td[1]/a').click()  # 2024/12/30
#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('起始時間:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="queryarea"]/div/div/div[2]/div[2]/div[2]/div/div/span/button/i').click()  # 結束時間
#         time.sleep(1)

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[4]/a').click()  # 2025/01/30
#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('結束時間:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'currency_search').click()  # 帳戶幣別
#         time.sleep(1)

#         accountTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="currency_search"]/option[@value="IDR"]')  # IDR
#         accountTypeSelect.click()
#         print('帳戶幣別:', accountTypeSelect.text, '\n')
#         time.sleep(1)   

#         self.driver.find_element(By.ID, 'agent_attr').click()  # 代理商類別
#         time.sleep(1)
        
#         agentTypeSelect = self.driver.find_element(By.XPATH, '//*[@id="agent_attr"]/option[@value="All"]')  # 代理商類別 -- 全選 --
#         agentTypeSelect.click()
#         print('代理商類別:', agentTypeSelect.text, '\n')
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="79713763-selectable"]')  # 代理商 Test-5                                                                                               
#         agent_select.click()
#         print('代理商:', agent_select.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="game_class"]').click()  #  遊戲類型                                                                                                 
#         time.sleep(1)

#         game_class_slt = self.driver.find_element(By.XPATH, f'//*[@id="game_class"]/option[@value="slot"]')  # slot
#         game_class_slt.click()
#         print('遊戲類型:', game_class_slt.get_attribute('value'), '\n')
#         time.sleep(1)

#         game_slt = self.driver.find_element(By.XPATH, f'//*[@id="1839777187-selectable"]')  # 遊戲 PSS-ON-00158_賽博魔方
#         game_slt.click()
#         print('遊戲:', game_slt.text, '\n')
#         time.sleep(1)
        
#         self.driver.find_element(By.XPATH, f'//*[@id="sh_btn"]').click()  # 送出    
#         time.sleep(3)

#         page_end = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[7]')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#         time.sleep(2)

#         exportIcon = self.driver.find_element(By.XPATH, '//*[@id="export_all"]/i') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", exportIcon) # 頁首 (匯出)
#         exportIcon.click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="allarea"]/div/div[1]/h5/span[1]').click()
#         time.sleep(2)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 錯誤訊息驗證 **************************

#     def test_error_info_identify(self):
#         '''錯誤訊息驗證'''
   
#         self.driver.find_element(By.XPATH, '//*[@id="queryarea"]/div/div/div[2]/div[2]/div[1]/div/div/span/button/i').click()  # 起始時間
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[6]/td[1]/a').click()  # 2024/12/30
#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('起始時間:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="queryarea"]/div/div/div[2]/div[2]/div[2]/div/div/span/button/i').click()  # 結束時間
#         time.sleep(1)

#         for _ in range(4):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[6]/a').click()  # 2025/02/01
#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('結束時間:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="sh_btn"]').click()   
#         time.sleep(2)

#         try:
#             alertt = self.driver.switch_to.alert
#             print(alertt.text, '\n')
#             time.sleep(3)
#             alertt.accept()
                
#         except NoAlertPresentException:
#             pass
        
#         time.sleep(2)
#         self.driver.refresh()
#         time.sleep(1)


# # --------------------------- 營運報表(EN) ---------------------------

#     def test_Accounting_Operating_Statement_EN(self):
#         '''【帳務】營運報表語系切換(英)''' 

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[1]').click()
#         print("語系已切換'English'!", '\n')
#         time.sleep(1)


# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_os_EN(self):
#         '''搜尋列功能驗證(英)'''


# # ************************** 起始時間 **************************

#         self.driver.find_element(By.XPATH, '//*[@id="queryarea"]/div/div/div[2]/div[2]/div[1]/div/div/span/button/i').click()  # Start Time
#         time.sleep(1)

#         for _ in range(7):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[7]/a').click()  # 2024/11/30
#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('Start Time:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)  
        

# # ************************** 結束時間 **************************

#         self.driver.find_element(By.XPATH, '//*[@id="queryarea"]/div/div/div[2]/div[2]/div[2]/div/div/span/button/i').click()  # End Time
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[2]/a').click()  # 2024/12/30
#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('End Time:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)   

#         self.driver.refresh()


# #  ************************** 帳戶幣別 **************************

#         self.driver.find_element(By.ID, f'currency_search').click()  # Account currency
#         time.sleep(1)
        
        
#         currency_list = ['KRW', 'THB', 'USD']
        
#         for account_id in currency_list:
#             accountTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="currency_search"]/option[@value="{account_id}"]')  # KRW / THB / USD
#             accountTypeSelect.click()
#             print('Account currency:', accountTypeSelect.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()


# #  ************************** 代理商類別 **************************

#         self.driver.find_element(By.ID, 'agent_attr').click()  # Agent Class
#         time.sleep(1)
        
#         agent_class_list = ['All', 'C66', 'Platform']
        
#         for agent_id in agent_class_list:    
#             agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="{agent_id}"]')  # -- Select All -- / C66 / Platform
#             agentTypeSelect.click()
#             print('Agent Class:', agentTypeSelect.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()


# #  ************************** 營運報表 **************************

#     def test_operating_statement_search_func_identify_EN(self):
#         '''營運報表搜尋功能驗證(英)'''  

#         self.driver.find_element(By.XPATH, '//*[@id="queryarea"]/div/div/div[2]/div[2]/div[1]/div/div/span/button/i').click()  # Start Time
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[2]/a').click()  # 2024/12/30
#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('Start Time:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="queryarea"]/div/div/div[2]/div[2]/div[2]/div/div/span/button/i').click()  # End Time
#         time.sleep(1)

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[5]/a').click()  # 2025/01/30
#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('End Time:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'currency_search').click()  # Account currency
#         time.sleep(1)

#         accountTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="currency_search"]/option[@value="IDR"]')  # IDR
#         accountTypeSelect.click()
#         print('Account currency:', accountTypeSelect.text, '\n')
#         time.sleep(1)   

#         self.driver.find_element(By.ID, 'agent_attr').click()  # Agent Class
#         time.sleep(1)
        
#         agentTypeSelect = self.driver.find_element(By.XPATH, '//*[@id="agent_attr"]/option[@value="All"]')  # -- Select All --
#         agentTypeSelect.click()
#         print('Agent Class:', agentTypeSelect.text, '\n')
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="79713763-selectable"]')  # Test-5                                                                                               
#         agent_select.click()
#         print('Agent:', agent_select.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="game_class"]').click  #  Game Class                                                                                                 
#         time.sleep(1)

#         game_class_slt = self.driver.find_element(By.XPATH, f'//*[@id="game_class"]/option[@value="slot"]')  # slot
#         game_class_slt.click()
#         print('Game Class:', game_class_slt.get_attribute('value'), '\n')
#         time.sleep(1)

#         game_slt = self.driver.find_element(By.XPATH, f'//*[@id="1839777187-selectable"]')  # PSS-ON-00158_CYBER CUBE
#         game_slt.click()
#         print('Game:', game_slt.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="sh_btn"]').click()  # Submit 
#         time.sleep(3)

#         page_end = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[7]')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#         time.sleep(2)

#         exportIcon = self.driver.find_element(By.XPATH, '//*[@id="export_all"]/i') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", exportIcon) # 頁首 (匯出)
#         exportIcon.click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="allarea"]/div/div[1]/h5/span[1]').click()
#         time.sleep(2)

#         self.driver.refresh()


# #  ************************** Error_info identify **************************

#     def test_error_info_identify_EN(self):
#         '''錯誤訊息驗證(英)'''
   
#         self.driver.find_element(By.XPATH, '//*[@id="queryarea"]/div/div/div[2]/div[2]/div[1]/div/div/span/button/i').click()  # Start Time
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[2]/a').click()  # 2024/12/30
#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('Start Time:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="queryarea"]/div/div/div[2]/div[2]/div[2]/div/div/span/button/i').click()  # End Time
#         time.sleep(1)

#         for _ in range(4):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[7]/a').click()  # 2025/02/01
#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('End Time:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="sh_btn"]').click()    
#         time.sleep(2)

#         try:
#             alertt = self.driver.switch_to.alert
#             print(alertt.text, '\n')
#             time.sleep(3)
#             alertt.accept()
                
#         except NoAlertPresentException:
#             pass

#         time.sleep(2)
#         self.driver.refresh()
#         time.sleep(1)
        

# # --------------------------- 營運報表(Tai) ---------------------------

#     def test_Accounting_Operating_Statement_Tai(self):
#         '''【帳務】營運報表語系切換(泰)'''     

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[4]').click()
#         print("語系已切換'ไทย'!", '\n')
        
        
# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_os_Tai(self):
#         '''搜尋列功能驗證(泰)'''


# # ************************** 起始時間 ************************** 

#         self.driver.find_element(By.XPATH, '//*[@id="queryarea"]/div/div/div[2]/div[2]/div[1]/div/div/span/button/i').click()  # เวลาเริ่มต้น
#         time.sleep(1)

#         for _ in range(7):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[7]/a').click()  # 2024/11/30
#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('เวลาเริ่มต้น:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)   


# # ************************** 結束時間 **************************

#         self.driver.find_element(By.XPATH, '//*[@id="queryarea"]/div/div/div[2]/div[2]/div[2]/div/div/span/button/i').click()  # เวลาสิ้นสุด
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[2]/a').click()  # 2024/12/30
#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('เวลาสิ้นสุด:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)   

#         self.driver.refresh()


# #  ************************** 帳戶幣別 **************************

#         self.driver.find_element(By.ID, f'currency_search').click()  # สกุลเงินในบัญชี
#         time.sleep(1)

#         currency_list = ['KRW', 'THB', 'USD']
        
#         for account_id in currency_list:
#             accountTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="currency_search"]/option[@value="{account_id}"]')  # KRW / THB / USD
#             accountTypeSelect.click()
#             print('สกุลเงินในบัญชี:', accountTypeSelect.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()


# #  ************************** 代理商類別 **************************

#         self.driver.find_element(By.ID, 'agent_attr').click()  # หมวดตัวแทน
#         time.sleep(1)

#         agent_class_list = ['All', 'C66', 'Platform']
        
#         for agent_id in agent_class_list:    
#             agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="{agent_id}"]')  # -- เลือกทั้งหมด -- / C66 / Platform
#             agentTypeSelect.click()
#             print('หมวดตัวแทน:', agentTypeSelect.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()


# #  ************************** คำชี้แจงการดำเนินงาน **************************

#     def test_operating_statement_search_func_identify_Tai(self):
#         '''營運報表搜尋功能驗證(泰)'''  

#         self.driver.find_element(By.XPATH, '//*[@id="queryarea"]/div/div/div[2]/div[2]/div[1]/div/div/span/button/i').click()  # เวลาเริ่มต้น
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[2]/a').click()  # 2024/12/30  
#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('เวลาเริ่มต้น:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="queryarea"]/div/div/div[2]/div[2]/div[2]/div/div/span/button/i').click()  # เวลาสิ้นสุด
#         time.sleep(1)

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[5]/a').click()  # 2025/01/30
#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('เวลาสิ้นสุด:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'currency_search').click()  # สกุลเงินในบัญชี
#         time.sleep(1)

#         accountTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="currency_search"]/option[@value="IDR"]')  # IDR
#         accountTypeSelect.click()
#         print('สกุลเงินในบัญชี:', accountTypeSelect.text, '\n')
#         time.sleep(1)   

#         self.driver.find_element(By.ID, 'agent_attr').click()  # หมวดตัวแทน
#         time.sleep(1)
        
#         agentTypeSelect = self.driver.find_element(By.XPATH, '//*[@id="agent_attr"]/option[@value="All"]')  # -- เลือกทั้งหมด --                                                                                                                            
#         agentTypeSelect.click()
#         print('หมวดตัวแทน:', agentTypeSelect.text, '\n')
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="79713763-selectable"]')  # Test-5                                                                                                 
#         agent_select.click()
#         print('ตัวแทน:', agent_select.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="game_class"]').click()  #  ประเภทเกม                                                                                                 
#         time.sleep(1)

#         game_class_slt = self.driver.find_element(By.XPATH, f'//*[@id="game_class"]/option[2]')  # ประเภทเกม สล็อตแมชชีน
#         game_class_slt.click()
#         print('ประเภทเกม:', game_class_slt.text, '\n')
#         time.sleep(1)

#         game_slt = self.driver.find_element(By.XPATH, f'//*[@id="1839777187-selectable"]')  # PSS-ON-00158_ไซเบอร์คิวบ์
#         game_slt.click()
#         print('เลือกเกม:', game_slt.text, '\n')
#         time.sleep(1)
        
#         self.driver.find_element(By.XPATH, f'//*[@id="sh_btn"]').click()  # ส่ง 
#         time.sleep(3)

#         page_end = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[7]')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#         time.sleep(2)

#         exportIcon = self.driver.find_element(By.XPATH, '//*[@id="export_all"]/i') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", exportIcon) # 頁首 (ส่งออก)
#         exportIcon.click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="allarea"]/div/div[1]/h5/span[1]').click()
#         time.sleep(2)
            
#         self.driver.refresh()


# #  ************************** Error_info identify **************************

#     def test_error_info_identify_Tai(self):
#         '''錯誤訊息驗證(泰)'''
   
#         self.driver.find_element(By.XPATH, '//*[@id="queryarea"]/div/div/div[2]/div[2]/div[1]/div/div/span/button/i').click()  # เวลาเริ่มต้น
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[2]/a').click()  # 2024/12/30
#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('เวลาเริ่มต้น:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="queryarea"]/div/div/div[2]/div[2]/div[2]/div/div/span/button/i').click()  # เวลาสิ้นสุด
#         time.sleep(1)

#         for _ in range(4):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[7]/a').click()  # 2025/02/01
#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('เวลาสิ้นสุด:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="sh_btn"]').click()    
#         time.sleep(2)

#         try:
#             alertt = self.driver.switch_to.alert
#             print(alertt.text, '\n')
#             time.sleep(3)
#             alertt.accept()
                
#         except NoAlertPresentException:
#             pass

#         time.sleep(2)
#         self.driver.refresh()
#         time.sleep(1)


# # ============================= 後台功能巡測【玩家】============================= 
# # *************************** 玩家資訊 ***************************

#     def test_Player_Player_Info_zhCN(self):
#         '''【玩家】玩家資訊功能頁切換'''  
    
#         self.driver.find_element(By.ID, 'Player').click()
#         print("進入玩家功能選單!", '\n')
#         time.sleep(1)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Player/player_info')
#         print("切換玩家資訊選單!", '\n')
#         time.sleep(1)
        
#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[3]').click()
#         print("語系已切換'简体中文'!", '\n')
#         time.sleep(1)
        

# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_pi_zhCN(self):
#         '''搜尋列功能驗證'''
        
        
# #  ************************** 代理商類別 **************************

#         self.driver.find_element(By.ID, f'agent_attr_black').click()  # 代理商類別
#         time.sleep(1)
        
#         agent_id_list = ['All', 'PS', 'Test']    
#         for agent_id in agent_id_list:
#             agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr_black"]/option[@value="{agent_id}"]')  # -- 全選 -- / PS / Test
#             agentTypeSelect.click()
#             print('代理商類別:', agentTypeSelect.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)
        

# #  ************************** 代理商 **************************

#         self.driver.find_element(By.ID, f'agent_attr_black').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_slt = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr_black"]/option[@value="All"]')  # -- 全選 --
#         agent_class_slt.click()
#         print('代理商類別:', agent_class_slt.get_attribute('value'), '\n')
#         time.sleep(1)

#         agent_id_list = ['platform', 'TEST', 'TEST2']
#         for agent_id_1 in agent_id_list:
#             self.driver.find_element(By.ID, f'black_list_agent').click()  # 代理商
#             time.sleep(1)
            
#             agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="black_list_agent"]/option[@value="{agent_id_1}"]')  # Platform / Test / Test-2
#             agentTypeSelect.click()
#             print('代理商:', agentTypeSelect.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 遊戲類型 **************************

#         for game_type in (1, 3):
            
#             self.driver.find_element(By.ID, f'player_info_game_type').click()  # 遊戲類型
#             time.sleep(1)
            
#             agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="player_info_game_type"]/option[{game_type}]')  # 老虎機 / 捕魚機 / 棋牌遊戲
#             agentSelect.click()
#             print('遊戲類型:', agentSelect.text, '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 玩家資訊搜尋 **************************

#     def test_player_info_search_func_identify(self):
#         '''玩家資訊搜尋功能驗證'''      

#         self.driver.find_element(By.ID, 'agent_attr_black').click()  # 代理商類別
#         time.sleep(1)
        
#         agentTypeSelect = self.driver.find_element(By.XPATH, '//*[@id="agent_attr_black"]/option[@value="All"]')  # 代理商類別 -- 全選 --
#         agentTypeSelect.click()
#         print('代理商類別:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'black_list_agent').click()  # 代理商
#         time.sleep(1)

#         agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="black_list_agent"]/option[@value="PLAYSTAR"]')  # PLAYSTAR
#         agentSelect.click()
#         print('代理商:', agentSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'player_info_game_type').click()  # 遊戲類型
#         time.sleep(1)
            
#         agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="player_info_game_type"]/option[1]')  # 老虎機 
#         agentSelect.click()
#         print('遊戲類型:', agentSelect.text, '\n')
#         time.sleep(1)

#         player_input = self.driver.find_element(By.XPATH, f'//*[@id="black_list_sh_player"]')  # 玩家I.D輸入框
#         player_input.click()
#         player_input.send_keys('ivan_li', ',', ' ', 'ivanTester_01', ',', ' ', 'ivanTester_02', ',', ' ', 'ivanTester_03', ',', ' ', 'ivanTester_04', ',', ' ', 'ivanTester_05')
#         time.sleep(1)
        
#         self.driver.find_element(By.ID, 'black_list_sh_btn').click() 
#         time.sleep(1)

#         select_table = self.driver.find_element(By.XPATH, '//*[@id="DataTables_Table_0_filter"]/label/input')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", select_table)  # 篩選

#         page_end = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[2]')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#         time.sleep(2)

#         page_head = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[1]/div/div/h4/span') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head) # 頁首
#         page_head.click()
#         time.sleep(1)

#         self.driver.refresh()


# # --------------------------- 玩家資訊(EN) ---------------------------

#     def test_Player_Player_Info_EN(self):
#         '''【玩家】玩家資訊語系切換(英)'''

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[1]').click()
#         print("語系已切換:", 'English', '\n')
#         time.sleep(1)
        

# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_pi_EN(self):
#         '''搜尋列功能驗證(英)'''


# #  ************************** Agent Class **************************

#         self.driver.find_element(By.ID, f'agent_attr_black').click()  # Agent Class
#         time.sleep(1)
            
#         agent_id_list = ['All', 'PS', 'Test']    
#         for agent_id in agent_id_list:
#             agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr_black"]/option[@value="{agent_id}"]')  # -- Select All -- / PS / Test
#             agentTypeSelect.click()
#             print('Agent Class:', agentTypeSelect.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** Agent **************************

#         self.driver.find_element(By.ID, f'agent_attr_black').click()  # Agent Class
#         time.sleep(1)

#         agent_class_slt = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr_black"]/option[@value="All"]')  # -- Select All --
#         agent_class_slt.click()
#         print('Agent Class:', agent_class_slt.get_attribute('value'), '\n')
#         time.sleep(1)
        
        
#         agent_id_list = ['platform', 'TEST', 'TEST2']
#         for agent_id_1 in agent_id_list:
#             self.driver.find_element(By.ID, f'black_list_agent').click()  # Agent
#             time.sleep(1)
            
#             agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="black_list_agent"]/option[@value="{agent_id_1}"]')  # Platform / Test / Test-2
#             agentTypeSelect.click()
#             print('Agent:', agentTypeSelect.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** Game Type **************************

#         for game_type in (1, 3):
            
#             self.driver.find_element(By.ID, f'player_info_game_type').click()  # Game Type
#             time.sleep(1)
            
#             agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="player_info_game_type"]/option[{game_type}]')  # Slot / Fishing Game / Card Game
#             agentSelect.click()
#             print('Game Type:', agentSelect.text, '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** Player Info search **************************

#     def test_player_info_search_func_identify_EN(self):
#         '''玩家資訊搜尋功能驗證(英)''' 

#         self.driver.find_element(By.ID, 'agent_attr_black').click()  # Agent Class
#         time.sleep(1)
        
#         agentTypeSelect = self.driver.find_element(By.XPATH, '//*[@id="agent_attr_black"]/option[@value="All"]')  # 代理商類別 -- Select All --
#         agentTypeSelect.click()
#         print('Agent Class:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'black_list_agent').click()  # Agent
#         time.sleep(1)

#         agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="black_list_agent"]/option[@value="PLAYSTAR"]')  # PLAYSTAR
#         agentSelect.click()
#         print('Agent:', agentSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'player_info_game_type').click()  # Game Type
#         time.sleep(1)
            
#         agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="player_info_game_type"]/option[1]')  # Slot 
#         agentSelect.click()
#         print('Game Type:', agentSelect.text, '\n')
#         time.sleep(1)

#         player_input = self.driver.find_element(By.XPATH, f'//*[@id="black_list_sh_player"]')  # 玩家I.D輸入框
#         player_input.click()
#         player_input.send_keys('ivan_li', ',', ' ', 'ivanTester_01', ',', ' ', 'ivanTester_02', ',', ' ', 'ivanTester_03', ',', ' ', 'ivanTester_04', ',', ' ', 'ivanTester_05')
#         time.sleep(1)
        
#         self.driver.find_element(By.ID, 'black_list_sh_btn').click()   
#         time.sleep(1)

#         select_table = self.driver.find_element(By.XPATH, '//*[@id="DataTables_Table_0_filter"]/label/input')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", select_table)  # 篩選

#         page_end = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[2]')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#         time.sleep(2)

#         page_head = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[1]/div/div/h4/span') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head) # 頁首
#         page_head.click()
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# # --------------------------- 玩家資訊(Tai) ---------------------------

#     def test_Player_Player_Info_Tai(self):
#         '''【玩家】玩家資訊語系切換(泰)'''
        

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[4]').click()
#         print("語系已切換:", 'ไทย', '\n')
#         time.sleep(1)


# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_pi_Tai(self):
#         '''搜尋列功能驗證(泰)'''
        

# #  ************************** หมวดตัวแทน **************************

#         self.driver.find_element(By.ID, f'agent_attr_black').click()  # หมวดตัวแทน
#         time.sleep(1)
        
#         agent_id_list = ['All', 'PS', 'Test']    
#         for agent_id in agent_id_list:
#            agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr_black"]/option[@value="{agent_id}"]')  # -- เลือกทั้งหมด -- / PS / Test
#            agentTypeSelect.click()
#            print('หมวดตัวแทน:', agentTypeSelect.get_attribute('value'), '\n')
#            time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** ตัวแทน **************************

#         self.driver.find_element(By.ID, f'agent_attr_black').click()  # หมวดตัวแทน
#         time.sleep(1)

#         agent_class_slt = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr_black"]/option[@value="All"]')  # -- เลือกทั้งหมด --
#         agent_class_slt.click()
#         print('หมวดตัวแทน:', agent_class_slt.get_attribute('value'), '\n')
#         time.sleep(1)
                
#         agent_id_list = ['platform', 'TEST', 'TEST2']
#         for agent_id_1 in agent_id_list:
#             self.driver.find_element(By.ID, f'black_list_agent').click()  # ตัวแทน
#             time.sleep(1)
            
#             agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="black_list_agent"]/option[@value="{agent_id_1}"]')  # Platform / Test / Test-2
#             agentTypeSelect.click()
#             print('ตัวแทน:', agentTypeSelect.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** ประเภทเกม **************************

#         for game_type in (1, 3):
            
#             self.driver.find_element(By.ID, f'player_info_game_type').click()  # ประเภทเกม
#             time.sleep(1)
            
#             agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="player_info_game_type"]/option[{game_type}]')  # สล็อตแมชชีน / รายการการใช้เหรียญทอง PS
#             agentSelect.click()
#             print('ประเภทเกม:', agentSelect.text, '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)
        

# #  ************************** ข้อมูลผู้เล่น search **************************

#     def test_player_info_search_func_identify_Tai(self):
#         '''玩家資訊搜尋功能驗證(泰)''' 

#         self.driver.find_element(By.ID, 'agent_attr_black').click()  # หมวดตัวแทน
#         time.sleep(1)
        
#         agentTypeSelect = self.driver.find_element(By.XPATH, '//*[@id="agent_attr_black"]/option[@value="All"]')  # -- เลือกทั้งหมด --
#         agentTypeSelect.click()
#         print('หมวดตัวแทน:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'black_list_agent').click()  # ตัวแทน
#         time.sleep(1)

#         agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="black_list_agent"]/option[@value="PLAYSTAR"]')  # PLAYSTAR
#         agentSelect.click()
#         print('ตัวแทน:', agentSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'player_info_game_type').click()  # ประเภทเกม
#         time.sleep(1)
            
#         game_type_Select = self.driver.find_element(By.XPATH, f'//*[@id="player_info_game_type"]/option[1]')  # สล็อตแมชชีน 
#         game_type_Select.click()
#         print('ประเภทเกม:', game_type_Select.text, '\n')
#         time.sleep(1)

#         player_input = self.driver.find_element(By.XPATH, f'//*[@id="black_list_sh_player"]')  # 玩家I.D輸入框
#         player_input.click()
#         player_input.send_keys('ivan_li', ',', ' ', 'ivanTester_01', ',', ' ', 'ivanTester_02', ',', ' ', 'ivanTester_03', ',', ' ', 'ivanTester_04', ',', ' ', 'ivanTester_05')
#         time.sleep(1)
        
#         self.driver.find_element(By.ID, 'black_list_sh_btn').click()  
#         time.sleep(1)

#         select_table = self.driver.find_element(By.XPATH, '//*[@id="DataTables_Table_0_filter"]/label/input')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", select_table)  # 篩選

#         page_end = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[2]')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#         time.sleep(2)

#         page_head = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[1]/div/div/h4/span') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head) # 頁首
#         page_head.click()
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# # ============================= 後台功能巡測【玩家】============================= 
# # *************************** 玩家查詢 ***************************

#     def test_Player_Player_Query_zhCN(self):
#         '''【玩家】玩家查詢功能頁切換''' 
    
#         self.driver.find_element(By.ID, 'Player').click()
#         print("進入玩家功能選單!", '\n')
#         time.sleep(1)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Player/player_query')
#         print("切換玩家查詢選單!", '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[3]').click()
#         print("語系已切換'简体中文'!", '\n')
#         time.sleep(1)
        

# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_pq_zhCN(self):
#         '''搜尋列功能驗證'''
        
        
# # ************************** 起始時間 **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button/i').click()  # 起始時間
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[6]/td[1]/a').click()  # 2024/12/30
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('起始時間:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]').click()  # 關閉
#         time.sleep(1)
        
  
# # ************************** 結束時間 **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i').click()  # 結束時間
#         time.sleep(1)

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[5]/a').click()  # 2025/01/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('結束時間:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]').click()  # 關閉
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 代理商類別 **************************

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)
    
#         agent_class_list = ['PS', 'Test']
#         for agent_class_id in agent_class_list:            
#             agent_class_slt = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="{agent_class_id}"]')  # PS / Test
#             agent_class_slt.click()
#             print('代理商類別:', agent_class_slt.get_attribute('value'), '\n')
#             time.sleep(1)
        
#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 代理商 **************************
        
#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="All"]')  # -- 全選 -- 
#         agentTypeSelect.click()
#         print('代理商類別:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         agent_list = ['PLAYSTAR', 'PS-APP', 'TEST2']
#         for agent_id in agent_list:
#             self.driver.find_element(By.ID, 'agent').click()  # 代理商
#             time.sleep(1)
            
#             agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="{agent_id}"]')  # PLAYSTAR / PS-APP / Test-2
#             agentSelect.click()
#             print('代理商:', agentSelect.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)
        

# #  ************************** 玩家查詢搜尋 **************************

#     def test_player_query_search_func_identify_zhCN(self):
#         '''玩家查詢搜尋功能驗證'''    

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button/i').click()  # 起始時間
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[6]/td[1]/a').click()  # 2024/12/30
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('起始時間:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         button = self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]')  # 確定
#         self.driver.execute_script("$(arguments[0]).click()",button)
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i').click()  # 結束時間
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[6]/td[1]/a').click()  # 2024/12/30
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('結束時間:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]').click()  # 確定
#         time.sleep(1) 

#         self.driver.find_element(By.ID, 'agent_attr').click()  # 代理商類別
#         time.sleep(1)
        
#         agentTypeSelect = self.driver.find_element(By.XPATH, '//*[@id="agent_attr"]/option[@value="All"]')  # 代理商類別 -- 全選 --
#         agentTypeSelect.click()
#         print('代理商類別:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent').click()  # 代理商
#         time.sleep(1)

#         agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="PLAYSTAR"]')  # PLAYSTAR
#         agentSelect.click()
#         print('代理商:', agentSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'game_select').click()  # 選擇遊戲
#         time.sleep(1)

#         game_select = self.driver.find_element(By.XPATH, f'//*[@id="game_select"]/option[@value="All"]')  # -- 全選 --
#         game_select.click()
#         print('遊戲:', game_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'sh_btn').click()   
#         time.sleep(1)

#         page_end = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[4]')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#         time.sleep(2)

#         page_head = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[1]/div/div/h4/i') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head) # 頁首
#         page_head.click()
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)
        

# #  ************************** 錯誤訊息驗證 **************************

#     def test_error_info_identify_pq(self):
#         '''錯誤訊息驗證'''

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button/i').click()  # 起始時間
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[6]/td[1]/a').click()  # 2024/12/30
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]').click()  # 確定
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('起始時間:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i').click()  # 結束時間
#         time.sleep(1)

#         for _ in range(4):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[2]/td[5]/a').click()  # 2025/02/07
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]').click()  # 確定
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('結束時間:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'sh_btn').click()   
#         time.sleep(1)

#         error_info = self.driver.find_element(By.XPATH, f'/html/body/div[2]') 
#         print(error_info.text, '\n')

#         self.driver.refresh()
#         time.sleep(1)


# # --------------------------- 玩家查詢(EN) ---------------------------

#     def test_Player_Player_Query_EN(self):
#         '''【玩家】玩家查詢語系切換(英)'''
             
#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[1]').click()
#         print("語系已切換:", 'English', '\n')
#         time.sleep(1)
        
        
# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_pq_EN(self):
#         '''搜尋列功能驗證(英)'''
        

# # ************************** Start Time **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button/i').click()  # Start Time
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[2]/a').click()  # 2024/12/30
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('Start Time:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]').click()  # Close
#         time.sleep(1)

  
# # ************************** End Time **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i').click()  # End Time
#         time.sleep(1)

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[6]/a').click()  # 2025/01/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('End Time:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]')  # Close
#         time.sleep(1)   

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** Agent Class **************************

#         self.driver.find_element(By.ID, f'agent_attr').click()  # Agent Class
#         time.sleep(1)
        
#         agent_class_list = ['PS', 'Test']
#         for agent_class_id in agent_class_list:            
#             agent_class_slt = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="{agent_class_id}"]')  # PS / Test
#             agent_class_slt.click()
#             print('Agent Class:', agent_class_slt.get_attribute('value'), '\n')
#             time.sleep(1)
        
#         self.driver.refresh()
#         time.sleep(1)
    

# #  ************************** Agent **************************

#         self.driver.find_element(By.ID, f'agent_attr').click()  # Agent Class
#         time.sleep(1)

#         agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="All"]')  # -- Select All -- 
#         agentTypeSelect.click()
#         print('Agent Class:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         agent_list = ['PLAYSTAR', 'PS-APP', 'TEST2']
#         for agent_id in agent_list:
#             self.driver.find_element(By.ID, 'agent').click()  # Agent
#             time.sleep(1)
            
#             agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="{agent_id}"]')  # PLAYSTAR / PS-APP / Test-2
#             agentSelect.click()
#             print('Agent:', agentSelect.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** Player Query search **************************

#     def test_player_query_search_func_identify_EN(self):
#         '''玩家查詢搜尋功能驗證(英)'''  

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button/i').click()  # Start Time
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[2]/a').click()  # 2024/12/30
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('Start Time:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]').click()  # Close
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i').click()  # End Time
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[2]/a').click()  # 2024/12/30
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('End Time:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]').click()  # Close
#         time.sleep(1)  

#         self.driver.find_element(By.ID, 'agent_attr').click()  # Agent Class
#         time.sleep(1)
        
#         agentTypeSelect = self.driver.find_element(By.XPATH, '//*[@id="agent_attr"]/option[@value="All"]')  # -- Select All --
#         agentTypeSelect.click()
#         print('Agent Class:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent').click()  # Agent
#         time.sleep(1)

#         agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="PLAYSTAR"]')  # PLAYSTAR
#         agentSelect.click()
#         print('Agent:', agentSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'game_select').click()  # Select Game
#         time.sleep(1)

#         game_select = self.driver.find_element(By.XPATH, f'//*[@id="game_select"]/option[@value="All"]')  # -- Select All --
#         game_select.click()
#         print('Game:', game_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'sh_btn').click()   
#         time.sleep(1)

#         page_end = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[4]')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#         time.sleep(2)

#         page_head = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[1]/div/div/h4/i') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head) # 頁首
#         page_head.click()
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** error info 驗證 **************************

#     def test_error_info_identify_pq_EN(self):
#         '''錯誤訊息驗證(英)'''

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button/i').click()  # Start Time
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[2]/a').click()  # 2024/12/30
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]').click()  # Close
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('Start Time:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i').click()  # End Time
#         time.sleep(1)

#         for _ in range(4):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[2]/td[6]/a').click()  # 2025/02/07
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]').click()  # Close
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('End Time:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'sh_btn').click()    
#         time.sleep(1)

#         error_info = self.driver.find_element(By.XPATH, f'/html/body/div[2]')
#         print(error_info.text, '\n')

#         self.driver.refresh()
#         time.sleep(1)


# # --------------------------- 玩家查詢(Tai) ---------------------------

#     def test_Player_Player_Query_Tai(self):
#         '''【玩家】玩家查詢語系切換(泰)'''
        
#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1) 

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[4]').click()
#         print("語系已切換:", 'ไทย', '\n')
#         time.sleep(1)
        

# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_pq_Tai(self):
#         '''搜尋列功能驗證(泰)'''


# # ************************** เวลาเริ่มต้น **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button/i').click()  # เวลาเริ่มต้น
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[2]/a').click()  # 2024/12/30
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('เวลาเริ่มต้น:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2] ').click()  # Done
#         time.sleep(1)

  
# # ************************** เวลาสิ้นสุด **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i').click()  # เวลาสิ้นสุด
#         time.sleep(1)

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[6]/a').click()  # 2025/01/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('เวลาสิ้นสุด:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]').click()  # Done
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** หมวดตัวแทน **************************

#         self.driver.find_element(By.ID, f'agent_attr').click()  # หมวดตัวแทน
#         time.sleep(1)
        
#         agent_class_list = ['PS', 'Test']
#         for agent_class_id in agent_class_list:            
#             agent_class_slt = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="{agent_class_id}"]')  # PS / Test
#             agent_class_slt.click()
#             print('หมวดตัวแทน:', agent_class_slt.get_attribute('value'), '\n')
#             time.sleep(1)
        
#         self.driver.refresh()
#         time.sleep(1)
        

# #  ************************** หมวดตัวแทน **************************

#         self.driver.find_element(By.ID, f'agent_attr').click()  # หมวดตัวแทน
#         time.sleep(1)

#         agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="All"]')  # -- เลือกทั้งหมด -- 
#         agentTypeSelect.click()
#         print('หมวดตัวแทน:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         agent_list = ['PLAYSTAR', 'PS-APP', 'TEST2']
#         for agent_id in agent_list:
#             self.driver.find_element(By.ID, 'agent').click()  # ตัวแทน
#             time.sleep(1)
            
#             agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="{agent_id}"]')  # PLAYSTAR / PS-APP / Test-2
#             agentSelect.click()
#             print('ตัวแทน:', agentSelect.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** แบบสอบถามผู้เล่น search **************************

#     def test_player_query_search_func_identify_Tai(self):
#         '''玩家查詢搜尋功能驗證(泰)''' 

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button/i').click()  # เวลาเริ่มต้น
#         time.sleep(1)

#         for _ in range(6):
#            self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#            time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[2]/a').click()  # 2024/12/30
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('เวลาเริ่มต้น:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]').click()  # Done
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i').click()  # เวลาสิ้นสุด
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[2]/a').click()  # 2024/12/30
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('เวลาสิ้นสุด:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]')  # Done
#         time.sleep(1)   

#         self.driver.find_element(By.ID, 'agent_attr').click()  # หมวดตัวแทน
#         time.sleep(1)
        
#         agentTypeSelect = self.driver.find_element(By.XPATH, '//*[@id="agent_attr"]/option[@value="All"]')  # -- เลือกทั้งหมด --
#         agentTypeSelect.click()
#         print('หมวดตัวแทน:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent').click()  # ตัวแทน
#         time.sleep(1)

#         agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="PLAYSTAR"]')  # PLAYSTAR
#         agentSelect.click()
#         print('ตัวแทน:', agentSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'game_select').click()  # เลือกเกม
#         time.sleep(1)

#         game_select = self.driver.find_element(By.XPATH, f'//*[@id="game_select"]/option[@value="All"]')  # -- เลือกทั้งหมด --
#         game_select.click()
#         print('เลือกเกม:', game_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'sh_btn').click()   
#         time.sleep(1)

#         page_end = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[4]')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#         time.sleep(2)

#         page_head = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[1]/div/div/h4/i') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head) # 頁首
#         page_head.click()
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)
        

# #  ************************** error info 驗證 **************************

#     def test_error_info_identify_pq_Tai(self):
#         '''錯誤訊息驗證(泰)'''

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button/i').click()  # เวลาเริ่มต้น
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[2]/a').click()  # 2024/12/30
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]').click()  # Close
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('เวลาเริ่มต้น:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i').click()  # เวลาสิ้นสุด
#         time.sleep(1)

#         for _ in range(4):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[2]/td[6]/a').click()  # 2025/02/07
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]').click()  # Close
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('เวลาสิ้นสุด:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'sh_btn').click()  
#         time.sleep(1)

#         error_info = self.driver.find_element(By.XPATH, f'/html/body/div[2]')
#         print(error_info.text, '\n')
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)
        
        
# # ============================= 後台功能巡測【玩家】============================= 
# # *************************** 遊戲紀錄 ***************************

#     def test_Player_Game_History_zhCN(self):
#         '''【玩家】遊戲紀錄功能頁切換'''    

#         self.driver.find_element(By.ID, 'Player').click()
#         print("進入玩家功能選單!", '\n')
#         time.sleep(2)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Player/game_history')
#         print("切換遊戲紀錄選單!", '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[3]').click()
#         time.sleep(1)
        
#         print("語系已切換'简体中文'!", '\n')
#         time.sleep(1)
        

# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_gh_zhCN(self):
#         '''搜尋列功能驗證'''
        

# # --------------------------- 遊戲紀錄(zh_cn) ---------------------------
# # ************************** 起始時間 **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button/i').click()  # 起始時間
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[6]/td[1]/a').click()  # 2024/12/30
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('起始時間:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         button = self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]')  # 關閉
#         self.driver.execute_script("$(arguments[0]).click()",button)
#         time.sleep(1)

  
# # ************************** 結束時間 **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i').click()  # 結束時間
#         time.sleep(1)

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[5]/a').click()  # 2025/01/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('結束時間:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]').click()  # 關閉
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 顯示筆數 **************************

#         self.driver.find_element(By.ID, 'count').click()  # 顯示筆數
#         time.sleep(1)
        
#         count_list = ['20', '50', '500']
#         for counts in count_list:
#             agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="count"]/option[@value="{counts}"]')  #  20 / 50 / 500
#             agentTypeSelect.click()
#             print('顯示筆數:', agentTypeSelect.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 代理商類別 **************************

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)
        
#         agent_class_list = ['All', 'PS', 'Test']
#         for agent_class_id in agent_class_list:
#             agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="{agent_class_id}"]')  # -- 全選 -- / PS / Test
#             agentTypeSelect.click()
#             print('代理商類別:', agentTypeSelect.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 代理商 **************************

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="RK5"]')  # RK5
#         agentTypeSelect.click()
#         print('代理商類別:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent').click()  # 代理商
#         time.sleep(1)

#         agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="RK5-THB"]')  # RK5-THB
#         agentSelect.click()
#         print('代理商:', agentSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test
#         agentTypeSelect.click()
#         print('代理商類別:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent').click()  # 代理商
#         time.sleep(1)

#         agentSelect_1 = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="TEST3"]')  # Test-3
#         agentSelect_1.click()
#         print('代理商:', agentSelect_1.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)
        

# #  ************************** 遊戲類型 **************************

#         for game_type in range(1, 4):
#             self.driver.find_element(By.ID, 'game_type').click()  # 遊戲類型
#             time.sleep(1)
            
#             agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="game_type"]/option[{game_type}]')  # -- 全選 -- / 老虎機 / 棋牌遊戲 
#             agentSelect.click()
#             print('遊戲類型:', agentSelect.text, '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 遊戲紀錄搜尋 **************************

#     def test_game_history_search_func_identify_zhCN(self):
#         '''遊戲紀錄搜尋功能驗證'''    

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button/i').click()  # 起始時間
#         time.sleep(1)

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[4]/a').click()  # 2025/01/30
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('起始時間:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]').click()  # 關閉
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i').click()  # 結束時間
#         time.sleep(1)

#         for _ in range(4):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[2]/td[5]/a').click()  # 2025/02/07
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('結束時間:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]').click()  # 關閉
#         time.sleep(1) 

#         self.driver.find_element(By.ID, 'agent_attr').click()  # 代理商類別
#         time.sleep(1)
        
#         agentTypeSelect = self.driver.find_element(By.XPATH, '//*[@id="agent_attr"]/option[@value="All"]')  # 代理商類別 -- 全選 --
#         agentTypeSelect.click()
#         print('代理商類別:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent').click()  # 代理商
#         time.sleep(1)

#         agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="PLAYSTAR"]')  # PLAYSTAR
#         agentSelect.click()
#         print('代理商:', agentSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'game_select').click()  # 選擇遊戲
#         time.sleep(1)

#         game_select =self.driver.find_element(By.XPATH, f'//*[@id="game_type"]/option[@value="PSS%"]')  # 老虎機
#         game_select.click()
#         print('遊戲類型:', game_select.text, '\n')
#         time.sleep(1)

#         player_input = self.driver.find_element(By.XPATH, f'//*[@id="player"]')  # 玩家名稱輸入欄位
#         player_input.click()
#         player_input.send_keys('ivan_li')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'sh_btn').click()  
#         time.sleep(1)

#         page_end = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[4]')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#         time.sleep(1)

#         page_head = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[1]/div/div/h4/i') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head) # 頁首
#         page_head.click()
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# # --------------------------- 遊戲紀錄(EN) ---------------------------

#     def test_Player_Game_History_EN(self):
#         '''【玩家】遊戲紀錄語系切換(英)'''   

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[1]').click()
#         time.sleep(1)
        
#         print("語系已切換:", 'English', '\n')
#         time.sleep(1)
        
        
# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_gh_EN(self):
#         '''搜尋列功能驗證(英)'''        
        

# # ************************** Start Time **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button/i').click()  # Start Time
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[2]/a').click()  # 2024/12/30
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('Start Time:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2] ').click()  # 關閉
#         time.sleep(1)

  
# # ************************** End Time **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i').click()  # End Time
#         time.sleep(1)

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[6]/a').click()  # 2025/01/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('End Time:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2] ').click()  # 關閉
#         time.sleep(1)  

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** Show Count **************************

#         self.driver.find_element(By.ID, 'count').click()  # Show Count
#         time.sleep(1)
        
#         count_list = ['20', '50', '500']
#         for counts in count_list:
#             agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="count"]/option[@value="{counts}"]')  #  20 / 50 / 500
#             agentTypeSelect.click()
#             print('Show Count:', agentTypeSelect.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** Agent Class **************************

#         self.driver.find_element(By.ID, f'agent_attr').click()  # Agent Class
#         time.sleep(1)

#         agent_class_list = ['All', 'PS', 'Test']
#         for agent_class_id in agent_class_list:
#             agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="{agent_class_id}"]')  # -- Select All -- / PS / Test
#             agentTypeSelect.click()
#             print('Agent Class:', agentTypeSelect.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** Agent **************************

#         self.driver.find_element(By.ID, f'agent_attr').click()  # Agent Class
#         time.sleep(1)

#         agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="RK5"]')  # RK5
#         agentTypeSelect.click()
#         print('Agent Class:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent').click()  # Agent
#         time.sleep(1)

#         agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="RK5-THB"]')  # RK5-THB
#         agentSelect.click()
#         print('Agent:', agentSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'agent_attr').click()  # Agent Class
#         time.sleep(1)

#         agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test
#         agentTypeSelect.click()
#         print('Agent Class:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent').click()  # Agent
#         time.sleep(1)

#         agentSelect_1 = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="TEST3"]')  # Test-3
#         agentSelect_1.click()
#         print('Agent:', agentSelect_1.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** Game Class **************************

#         for game_type in range(1, 4):
#             self.driver.find_element(By.ID, 'game_type').click()  # Game Class
#             time.sleep(1)
            
#             agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="game_type"]/option[{game_type}]')  # -- Select All -- / Slot / Card Game  
#             agentSelect.click()
#             print('Game Class:', agentSelect.text, '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** Game History搜尋 **************************

#     def test_game_history_search_func_identify_EN(self):
#         '''遊戲紀錄搜尋功能驗證(英)'''   

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button/i').click()  # Start Time
#         time.sleep(1)

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[5]/a').click()  # 2025/01/30
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('Start Time:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2] ').click()  # 關閉
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i').click()  # End Time
#         time.sleep(1)

#         for _ in range(4):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[2]/td[6]/a').click()  # 2025/02/07
#         time.sleep(1)
        
#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('End Time:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2] ').click()  # 關閉
#         time.sleep(1)   

#         self.driver.find_element(By.ID, 'agent_attr').click()  # Agent Class
#         time.sleep(1)
        
#         agentTypeSelect = self.driver.find_element(By.XPATH, '//*[@id="agent_attr"]/option[@value="All"]')  # -- Select All --
#         agentTypeSelect.click()
#         print('Agent Class:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent').click()  # Agent
#         time.sleep(1)

#         agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="PLAYSTAR"]')  # PLAYSTAR
#         agentSelect.click()
#         print('Agent:', agentSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'game_select').click()  # Select Game
#         time.sleep(1)

#         game_select = self.driver.find_element(By.XPATH, f'//*[@id="game_type"]/option[@value="PSS%"]')  # SLOT
#         game_select.click()
#         print('Select Game:', game_select.text, '\n')
#         time.sleep(1)

#         player_input = self.driver.find_element(By.XPATH, f'//*[@id="player"]')  # 玩家名稱輸入欄位
#         player_input.click()
#         player_input.send_keys('ivan_li')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'sh_btn').click()    
#         time.sleep(1)

#         page_end = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[4]')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#         time.sleep(2)

#         page_head = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[1]/div/div/h4/i') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head) # 頁首
#         page_head.click()
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# # --------------------------- 遊戲紀錄(Tai) ---------------------------

#     def test_Player_Game_History_Tai(self):
#         '''【玩家】遊戲紀錄語系切換(泰)'''    

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[4]').click()
#         time.sleep(1)
        
#         print("語系已切換:", 'ไทย', '\n')
#         time.sleep(1)


# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_gh_Tai(self):
#         '''搜尋列功能驗證(泰)'''   


# # ************************** เวลาเริ่มต้น **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button/i').click()  # เวลาเริ่มต้น
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[2]/a').click()  # 2024/12/30
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('เวลาเริ่มต้น:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2] ').click()  # 關閉
#         time.sleep(1)

  
# # ************************** เวลาสิ้นสุด **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i').click()  # เวลาสิ้นสุด
#         time.sleep(1)

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[6]/a').click()  # 2025/01/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('เวลาสิ้นสุด:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2] ').click()  # 關閉
#         time.sleep(1)   

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** จำนวนครังที่แสดง **************************

#         self.driver.find_element(By.ID, 'count').click()  # จำนวนครังที่แสดง
#         time.sleep(1)
        
#         count_list = ['20', '50', '500']
#         for counts in count_list:
#             agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="count"]/option[@value="{counts}"]')  #  20 / 50 / 500
#             agentTypeSelect.click()
#             print('จำนวนครังที่แสดง:', agentTypeSelect.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** หมวดตัวแทน **************************

#         self.driver.find_element(By.ID, f'agent_attr').click()  # หมวดตัวแทน
#         time.sleep(1)
        
#         agent_class_list = ['All', 'PS', 'Test']
#         for agent_class_id in agent_class_list:
#             agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="{agent_class_id}"]')  # -- เลือกทั้งหมด -- / PS / Test
#             agentTypeSelect.click()
#             print('หมวดตัวแทน:', agentTypeSelect.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** ตัวแทน **************************

#         self.driver.find_element(By.ID, f'agent_attr').click()  # หมวดตัวแทน
#         time.sleep(1)

#         agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="RK5"]')  # RK5
#         agentTypeSelect.click()
#         print('หมวดตัวแทน:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent').click()  # ตัวแทน
#         time.sleep(1)

#         agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="RK5-THB"]')  # RK5-THB
#         agentSelect.click()
#         print('ตัวแทน:', agentSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'agent_attr').click()  # หมวดตัวแทน
#         time.sleep(1)

#         agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test
#         agentTypeSelect.click()
#         print('หมวดตัวแทน:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent').click()  # ตัวแทน
#         time.sleep(1)

#         agentSelect_1 = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="TEST3"]')  # Test-3
#         agentSelect_1.click()
#         print('ตัวแทน:', agentSelect_1.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)
        

# #  ************************** ประเภทเกม **************************

#         for game_type in range(1, 4):
            
#             self.driver.find_element(By.ID, 'game_type').click()  # ประเภทเกม
#             time.sleep(1)
            
#             agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="game_type"]/option[{game_type}]')  # -- เลือกทั้งหมด -- / สล็อตแมชชีน / รายการการใช้เหรียญทอง PS   
#             agentSelect.click()
#             print('ประเภทเกม:', agentSelect.text, '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** Game History搜尋 **************************

#     def test_game_history_search_func_identify_Tai(self):
#         '''遊戲紀錄搜尋功能驗證(泰)'''   

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button/i').click()  # เวลาเริ่มต้น
#         time.sleep(1)

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[5]/a').click()  # 2025/01/30
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('เวลาเริ่มต้น:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2] ').click()  # 關閉
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i').click()  # เวลาสิ้นสุด
#         time.sleep(1)

#         for _ in range(4):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[2]/td[6]/a').click()  # 2025/02/07
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('เวลาสิ้นสุด:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2] ').click()  # 關閉
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent_attr').click()  # หมวดตัวแทน
#         time.sleep(1)
        
#         agentTypeSelect = self.driver.find_element(By.XPATH, '//*[@id="agent_attr"]/option[@value="All"]')  # -- Select All --
#         agentTypeSelect.click()
#         print('หมวดตัวแทน:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent').click()  # ตัวแทน
#         time.sleep(1)

#         agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="PLAYSTAR"]')  # PLAYSTAR
#         agentSelect.click()
#         print('ตัวแทน:', agentSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'game_select').click()  # Sประเภทเกม
#         time.sleep(1)

#         game_select = self.driver.find_element(By.XPATH, f'//*[@id="game_type"]/option[2]')  # สล็อตแมชชีน
#         game_select.click()
#         print('ประเภทเกม:', game_select.text, '\n')
#         time.sleep(1)

#         player_input = self.driver.find_element(By.XPATH, f'//*[@id="player"]')  # 玩家名稱輸入欄位
#         player_input.click()
#         player_input.send_keys('ivan_li')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'sh_btn').click()    
#         time.sleep(1)

#         page_end = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[4]')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#         time.sleep(2)

#         page_head = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[1]/div/div/h4/i') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head) # 頁首
#         page_head.click()
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# # ============================= 後台功能巡測【玩家】============================= 
# # *************************** 交易紀錄 ***************************

#     def test_Player_Transaction_zhCN(self):
#         '''【玩家】交易紀錄功能頁切換'''       

#         self.driver.find_element(By.ID, 'Player').click()
#         print("進入玩家功能選單!", '\n')
#         time.sleep(1)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Player/transaction')
#         print("切換交易紀錄選單!", '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[3]').click()
#         time.sleep(1)
        
#         print("語系已切換'简体中文'!", '\n')
#         time.sleep(1)
        
        
# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_tr_zhCN(self):
#         '''搜尋列功能驗證'''


# # --------------------------- 交易紀錄(zh_cn) ---------------------------
# # ************************** 起始時間 **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button/i').click()  # 起始時間
#         time.sleep(1)

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[3]/a').click()  # 2025/01/01
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('起始時間:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

  
# # ************************** 結束時間 **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i').click()  # 結束時間
#         time.sleep(1)

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[5]/a').click()  # 2025/01/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('結束時間:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 代理商類別 **************************

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)
        
#         agent_id_list = ['Platform', 'PS', 'Test']
#         for agent_class_id in agent_id_list:
#             agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="{agent_class_id}"]')  # Platform / PS / Test
#             agentTypeSelect.click()
#             print('代理商類別:', agentTypeSelect.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)
        

# #  ************************** 代理商 **************************

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="RK5"]')  # RK5
#         agentTypeSelect.click()
#         print('代理商類別:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent').click()  # 代理商
#         time.sleep(1)

#         agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="RK5-THB"]')  # RK5-THB
#         agentSelect.click()
#         print('代理商:', agentSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test
#         agentTypeSelect.click()
#         print('代理商類別:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent').click()  # 代理商
#         time.sleep(1)

#         agentSelect_1 = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="TEST3"]')  # Test-3
#         agentSelect_1.click()
#         print('代理商:', agentSelect_1.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 類別 **************************

#         for game_type in range(1, 5):
#             self.driver.find_element(By.ID, 'game_type').click()  # 類別
#             time.sleep(1)
            
#             agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="game_type"]/option[{game_type}]')  # 老虎機 / 活動 / 棋牌遊戲 / Coin 
#             agentSelect.click()
#             print('類別:', agentSelect.text, '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 是否開啟內轉紀錄 **************************

#         self.driver.find_element(By.ID, 'internal_transfer').click()  # 內轉紀錄
#         time.sleep(1)

#         for transfer in (0, 1):
#             transfer_record = self.driver.find_element(By.XPATH, f'//*[@id="internal_transfer"]/option[@value="{transfer}"]')  # 否 / 是 
#             transfer_record.click()
#             print('是否開啟內轉紀錄:', transfer_record.text, '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 交易紀錄搜尋 **************************

#     def test_transaction_search_func_identify_zhCN(self):
#         '''交易紀錄搜尋功能驗證'''    

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button/i').click()  # 起始時間
#         time.sleep(1)

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[3]/a').click()  # 2025/01/01
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('起始時間:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i').click()  # 結束時間
#         time.sleep(1)

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[5]/a').click()  # 2025/01/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('結束時間:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent_attr')  # 代理商類別
#         time.sleep(1)
        
#         agentTypeSelect = self.driver.find_element(By.XPATH, '//*[@id="agent_attr"]/option[@value="All"]')  # 代理商類別 -- 全選 --
#         agentTypeSelect.click()
#         print('代理商類別:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent').click()  # 代理商
#         time.sleep(1)

#         agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="PLAYSTAR"]')  # PLAYSTAR
#         agentSelect.click()
#         print('代理商:', agentSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         for transfer in (0, 1):
#             self.driver.find_element(By.ID, 'internal_transfer').click()  # 內轉紀錄
#             time.sleep(1)
            
#             transfer_record = self.driver.find_element(By.XPATH, f'//*[@id="internal_transfer"]/option[@value="{transfer}"]')  #  否 / 是
#             transfer_record.click()
#             print('是否開啟內轉紀錄:', transfer_record.text, '\n')
#             time.sleep(1)
            
#             self.driver.find_element(By.ID, 'sh_btn').click()  
#             time.sleep(1)

#             page_end = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[3]')  
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#             time.sleep(1)

#             page_head = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[1]/div/div/h4/span') 
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head) # 頁首
#             page_head.click()
#             time.sleep(1)
            
#             exportIcon = self.driver.find_element(By.XPATH, '//*[@id="export"]/i') 
#             self.driver.execute_script("arguments[0].scrollIntoView();", exportIcon)  # 匯出
#             exportIcon.click()
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# # --------------------------- 交易紀錄(EN) ---------------------------

#     def test_Player_Transaction_EN(self):
#         '''【玩家】交易紀錄語系切換(英)'''
          
#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[1]').click()
#         time.sleep(1)
        
#         print("語系已切換:", 'English', '\n')
#         time.sleep(1)
        

# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_tr_EN(self):
#         '''搜尋列功能驗證(英)'''


# # ************************** Start Time **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button/i').click()  # Start Time
#         time.sleep(1)

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[4]/a').click()  # 2025/01/01
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('Start Time:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

  
# # ************************** End Time **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i').click()  # End Time
#         time.sleep(1)

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[6]/a').click()  # 2025/01/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('End Time:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)
        

# #  ************************** Agent Class **************************

#         self.driver.find_element(By.ID, f'agent_attr').click()  # Agent Class
#         time.sleep(1)

#         agent_id_list = ['Platform', 'PS', 'Test']        
#         for agent_class_id in agent_id_list:
#             agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="{agent_class_id}"]')  # Platform / PS / Test
#             agentTypeSelect.click()
            
#             print('Agent Class:', agentTypeSelect.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** Agent **************************

#         self.driver.find_element(By.ID, f'agent_attr').click()  # Agent Class
#         time.sleep(1)

#         agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="RK5"]')  # RK5
#         agentTypeSelect.click()
#         print('Agent Class:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent').click()  # Agent
#         time.sleep(1)

#         agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="RK5-THB"]')  # RK5-THB
#         agentSelect.click()
#         print('Agent:', agentSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'agent_attr').click()  # Agent Class
#         time.sleep(1)

#         agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test
#         agentTypeSelect.click()
#         print('Agent Class:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent').click()  # Agent
#         time.sleep(1)

#         agentSelect_1 = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="TEST3"]')  # Test-3
#         agentSelect_1.click()
#         print('Agent:', agentSelect_1.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)
        

# #  ************************** Type **************************

#         for game_type in range(1, 5):
#             self.driver.find_element(By.ID, 'game_type').click()  # Type
#             time.sleep(1)
            
#             agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="game_type"]/option[{game_type}]')  # Slot / Event / Card Game / Coin 
#             agentSelect.click()
#             print('Type:', agentSelect.text, '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)
        

# #  ************************** Do you want to open the transfer record between the wallets **************************

#         for transfer in (0, 1):
#             self.driver.find_element(By.ID, 'internal_transfer').click()  # transfer record
#             time.sleep(1)
            
#             transfer_record = self.driver.find_element(By.XPATH, f'//*[@id="internal_transfer"]/option[@value="{transfer}"]')  # No / Yes 
#             transfer_record.click()
            
#             print('Do you want to open the transfer record between the wallets:', transfer_record.text, '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** Transaction搜尋 **************************

#     def test_transaction_search_func_identify_EN(self):
#         '''交易紀錄搜尋功能驗證(英)'''    

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button/i').click()  # Start Time
#         time.sleep(1)

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[4]/a').click()  # 2025/01/01
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('Start Time:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i').click()  # End Time
#         time.sleep(1)

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)
            
#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[6]/a').click()  # 2025/01/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('End Time:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent_attr').click()  # Agent Class
#         time.sleep(1)
        
#         agentTypeSelect = self.driver.find_element(By.XPATH, '//*[@id="agent_attr"]/option[@value="All"]')  # -- Select All --
#         agentTypeSelect.click()
#         print('Agent Class:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent').click()  # Agent
#         time.sleep(1)

#         agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="PLAYSTAR"]')  # PLAYSTAR
#         agentSelect.click()
#         print('Agent:', agentSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         for transfer in (0, 1):
#             self.driver.find_element(By.ID, 'internal_transfer').click()  # transfer record
#             time.sleep(1)
            
#             transfer_record = self.driver.find_element(By.XPATH, f'//*[@id="internal_transfer"]/option[@value="{transfer}"]')  #  No / Yes
#             transfer_record.click()
#             print('Do you want to open the transfer record between the wallets:', transfer_record.text, '\n')
#             time.sleep(1)
            
#             self.driver.find_element(By.ID, 'sh_btn').click()  # Submit
#             time.sleep(1)

#             page_end = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[3]')  
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#             time.sleep(2)

#             page_head = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[1]/div/div/h4/span') 
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head) # 頁首
#             page_head.click()
#             time.sleep(1)
            
#             exportIcon = self.driver.find_element(By.XPATH, '//*[@id="export"]/i') 
#             self.driver.execute_script("arguments[0].scrollIntoView();", exportIcon)  # 匯出
#             exportIcon.click()
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# # --------------------------- 交易紀錄(Tai) ---------------------------

#     def test_Player_Transaction_Tai(self):
#         '''【玩家】交易紀錄語系切換(泰)''' 

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[4]').click()
#         time.sleep(1)
#         print("語系已切換:", 'ไทย', '\n')
#         time.sleep(1)


# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_tr_Tai(self):
#         '''搜尋列功能驗證(泰)'''
        

# # ************************** เวลาเริ่มต้น **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button/i').click()  # เวลาเริ่มต้น
#         time.sleep(1)

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[4]/a').click()  # 2025/01/01
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('เวลาเริ่มต้น:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

  
# # ************************** เวลาสิ้นสุด **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i').click()  # เวลาสิ้นสุด
#         time.sleep(1)

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[6]/a').click()  # 2025/01/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('เวลาสิ้นสุด:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** หมวดตัวแทน **************************

#         self.driver.find_element(By.ID, f'agent_attr').click()  # หมวดตัวแทน
#         time.sleep(1)

#         agent_id_list = ['Platform', 'PS', 'Test'] 
#         for agent_class_id in agent_id_list:
#             agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="{agent_class_id}"]')  # Platform / PS / Test
#             agentTypeSelect.click()
#             print('หมวดตัวแทน:', agentTypeSelect.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)
        

# #  ************************** ตัวแทน **************************

#         self.driver.find_element(By.ID, f'agent_attr').click()  # หมวดตัวแทน
#         time.sleep(1)

#         agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="RK5"]')  # RK5
#         agentTypeSelect.click()
#         print('หมวดตัวแทน:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent').click()  # ตัวแทน
#         time.sleep(1)

#         agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="RK5-THB"]')  # RK5-THB
#         agentSelect.click()
#         print('ตัวแทน:', agentSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'agent_attr').click()  # หมวดตัวแทน
#         time.sleep(1)

#         agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test
#         agentTypeSelect.click()
#         print('หมวดตัวแทน:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent').click()  # ตัวแทน
#         time.sleep(1)

#         agentSelect_1 = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="TEST3"]')  # Test-3
#         agentSelect_1.click()
#         print('ตัวแทน:', agentSelect_1.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)
        

# #  ************************** ประเภท **************************

#         for game_type in range(1, 5):
#             self.driver.find_element(By.ID, 'game_type').click()  # ประเภท
#             time.sleep(1)
            
#             agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="game_type"]/option[{game_type}]')  # สล็อตแมชชีน / กิจกรรม / รายการการใช้เหรียญทอง PS / Coin
#             agentSelect.click()
#             print('ประเภท:', agentSelect.text, '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** เปิดบันทึกการโอนภายในหรือไม่ **************************

#         for transfer in (0, 1):
#             self.driver.find_element(By.ID, 'internal_transfer').click()  # เปิดบันทึกการโอนภายในหรือไม่
#             time.sleep(1)
            
#             transfer_record = self.driver.find_element(By.XPATH, f'//*[@id="internal_transfer"]/option[@value="{transfer}"]')  # ไม่ / ใช่ 
#             transfer_record.click()
            
#             print('เปิดบันทึกการโอนภายในหรือไม่:', transfer_record.text, '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** บันทึกธุรกรรม搜尋 **************************

#     def test_transaction_search_func_identify_Tai(self):
#         '''交易紀錄搜尋功能驗證(泰)'''  

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button/i').click()  # เวลาเริ่มต้น
#         time.sleep(1)

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[4]/a').click()  # 2025/01/01
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('เวลาเริ่มต้น:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i').click()  # เวลาสิ้นสุด
#         time.sleep(1)

#         for _ in range(5):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[6]/a').click()  # 2025/01/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('เวลาสิ้นสุด:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent_attr').click()  # หมวดตัวแทน
#         time.sleep(1)
        
#         agentTypeSelect = self.driver.find_element(By.XPATH, '//*[@id="agent_attr"]/option[@value="All"]')  # -- เลือกทั้งหมด --
#         agentTypeSelect.click()
#         print('หมวดตัวแทน:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent').click()  # ตัวแทน
#         time.sleep(1)

#         agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="PLAYSTAR"]')  # PLAYSTAR
#         agentSelect.click()
#         print('ตัวแทน:', agentSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         for transfer in (0, 1):
#             self.driver.find_element(By.ID, 'internal_transfer').click()  # เปิดบันทึกการโอนภายในหรือไม่
#             time.sleep(1)
            
#             transfer_record = self.driver.find_element(By.XPATH, f'//*[@id="internal_transfer"]/option[@value="{transfer}"]')  # ไม่ / ใช่
#             transfer_record.click()
#             print('เปิดบันทึกการโอนภายในหรือไม่:', transfer_record.text, '\n')
#             time.sleep(1)
            
#             self.driver.find_element(By.ID, 'sh_btn').click()    
#             time.sleep(1)

#             page_end = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[3]')  
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#             time.sleep(2)

#             page_head = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[1]/div/div/h4/span') 
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head) # 頁首
#             page_head.click()
#             time.sleep(1)
            
#             exportIcon = self.driver.find_element(By.XPATH, '//*[@id="export"]/i') 
#             self.driver.execute_script("arguments[0].scrollIntoView();", exportIcon)  # 匯出
#             exportIcon.click()
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# # ============================= 後台功能巡測【玩家】============================= 
# # *************************** 錢包查詢 ***************************

#     def test_Player_Wallet_query_zhCN(self):
#         '''【玩家】錢包查詢功能頁切換'''    

#         self.driver.find_element(By.ID, 'Player').click()
#         print("進入玩家功能選單!", '\n')
#         time.sleep(1)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Player/wallet_search')
#         print("切換錢包查詢選單!", '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[3]').click()
#         print("語系已切換'简体中文'!", '\n')
#         time.sleep(1)
        
        
# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_wq_zhCN(self):
#         '''搜尋列功能驗證'''    

#         self.driver.find_element(By.ID, 'agent_attr').click()  # 代理商類別
#         time.sleep(1)
        
#         agentTypeSelect = self.driver.find_element(By.XPATH, '//*[@id="agent_attr"]/option[@value="All"]')  #  全選
#         agentTypeSelect.click()
#         print('代理商類別:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent').click()  # 代理商
#         time.sleep(1)

#         agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="PLAYSTAR"]')  # PLAYSTAR
#         agentSelect.click()
#         print('代理商:', agentSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent_attr').click()  # 選擇遊戲
#         time.sleep(1)

#         for game_id in (1, 4, 5):
#             game_select = self.driver.find_element(By.XPATH, f'//*[@id="game_select"]/option[{game_id}]')  # 博八博九 / 決戰52張 / U4
#             game_select.click()
#             print('選擇遊戲:', game_select.text, '\n')
#             time.sleep(1)

#             self.driver.find_element(By.ID, 'sh_btn').click()   
#             time.sleep(1)

#             page_end = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[3]')  
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#             time.sleep(2)

#             page_head = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[1]/div/div/h4/span') 
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head) # 頁首
#             page_head.click()
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# # --------------------------- 錢包查詢(EN) ---------------------------

#     def test_Player_Wallet_query_EN(self): 
#         '''【玩家】錢包查詢語系切換(英)'''
        
#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         language_en_slt = self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[1]')
#         language_en_slt.click()
#         print("語系已切換:", 'English', '\n')
#         time.sleep(1)
        
        
#     # ************************** 搜尋功能 **************************

#     def test_search_func_identify_wq_EN(self):
#         '''搜尋列功能驗證(英)'''  

#         self.driver.find_element(By.ID, 'agent_attr').click()  # Agent Class
#         time.sleep(1)
        
#         agentTypeSelect = self.driver.find_element(By.XPATH, '//*[@id="agent_attr"]/option[@value="All"]')  #  -- Select All --
#         agentTypeSelect.click()
#         print('Agent Class:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent').click()  # Agent
#         time.sleep(1)

#         agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="PLAYSTAR"]')  # PLAYSTAR
#         agentSelect.click()
#         print('Agent:', agentSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent_attr').click()  # Select Game
#         time.sleep(1)

#         for game_id in (1, 4, 5):
#             game_select = self.driver.find_element(By.XPATH, f'//*[@id="game_select"]/option[{game_id}]')  # Pok Deng / PAI KANG / U4
#             game_select.click()
#             print('Game:', game_select.text, '\n')
#             time.sleep(1)

#             self.driver.find_element(By.ID, 'sh_btn').click()    
#             time.sleep(1)

#             page_end = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[3]')  
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#             time.sleep(2)

#             page_head = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[1]/div/div/h4/span') 
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head) # 頁首
#             page_head.click()
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# # --------------------------- 錢包查詢(Tai) ---------------------------

#     def test_Player_Wallet_query_Tai(self):
#         '''【玩家】錢包查詢語系切換(泰)''' 

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[4]').click()
#         print("語系已切換:", 'ไทย', '\n')
#         time.sleep(1)
        
        
# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_wq_Tai(self):
#         '''搜尋列功能驗證(泰)'''  

#         self.driver.find_element(By.ID, 'agent_attr').click()  # หมวดตัวแทน
#         time.sleep(1)
        
#         agentTypeSelect = self.driver.find_element(By.XPATH, '//*[@id="agent_attr"]/option[@value="All"]')  #  -- เลือกทั้งหมด --
#         agentTypeSelect.click()
#         print('หมวดตัวแทน:', agentTypeSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent').click()  # ตัวแทน
#         time.sleep(1)

#         agentSelect = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="PLAYSTAR"]')  # PLAYSTAR
#         agentSelect.click()
#         print('ตัวแทน:', agentSelect.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'agent_attr').click()  # เลือกเกม
#         time.sleep(1)

#         for game_id in (1, 4, 5):
#             game_select = self.driver.find_element(By.XPATH, f'//*[@id="game_select"]/option[{game_id}]')  # ป๊อกเด้ง / ไพ่แคง / U4
#             game_select.click()
#             print('เลือกเกม:', game_select.text, '\n')
#             time.sleep(1)

#             self.driver.find_element(By.ID, 'sh_btn').click()    
#             time.sleep(1)

#             page_end = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[3]')  
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#             time.sleep(2)

#             page_head = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[1]/div/div/h4/span') 
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head) # 頁首
#             page_head.click()
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# # ============================= 後台功能巡測【活動】============================= 
# # *************************** 中獎名單(抽獎) ***************************

#     def test_Event_Winner_List_Lottery_zhCN(self):
#         '''【活動】中獎名單(抽獎)功能頁切換'''    

#         self.driver.find_element(By.ID, 'Event2').click()
#         print("進入活動功能選單!", '\n')
#         time.sleep(1)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Scratch/scratch_hit_list')
#         print("切換中獎名單(抽獎)選單!", '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[3]').click()
#         print("語系已切換'简体中文'!", '\n')
#         time.sleep(1)
        
        
# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_wl_zhCN(self):
#         '''搜尋列功能驗證'''


# # ************************** 起始時間 **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div[1]/div/div/span/button/i').click()  # 起始時間
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[7]/a').click()  # 2024/12/01
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('起始時間:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2] ').click()  # 關閉
#         time.sleep(1)

  
# # ************************** 結束時間 **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/span/button/i').click()  # 結束時間
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[6]/td[2]/a').click()  # 2024/12/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('結束時間:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2] ').click()  # 關閉
#         time.sleep(1)  

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 搜尋類別 **************************

#         for search_class in range(1, 3):
#             self.driver.find_element(By.ID, f'search_time_type').click()  # 搜尋類別
#             time.sleep(1) 
            
#             agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="search_time_type"]/option[{search_class}]')  # 派彩時間 / 結帳時間
#             agentTypeSelect.click()
#             print('搜尋類別:', agentTypeSelect.text, '\n')
#             time.sleep(1)
            
#         self.driver.refresh()
#         time.sleep(1)
        

# #  ************************** 顯示筆數 **************************

#         show_count_list = [2, 3, 6]
#         for show_count in show_count_list:
            
#             self.driver.find_element(By.ID, 'count').click()  # 顯示筆數
#             time.sleep(1)
            
#             show_count_num = self.driver.find_element(By.XPATH, f'//*[@id="count"]/option[{show_count}]')  # 50 / 100 / 1000
#             show_count_num.click() 
#             print('顯示筆數:', show_count_num.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 代理商類別 **************************

#         agent_class_list = ["All", 'Platform', 'PS', 'Test']
#         for agent_class in agent_class_list:
            
#             self.driver.find_element(By.ID, 'agent_attr').click()  # 代理商類別
#             time.sleep(1)
            
#             agent_class_name = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="{agent_class}"]')  # -- 全選 -- / Platform / PS / Test
#             agent_class_name.click() 
#             print('代理商類別:', agent_class_name.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 代理商 **************************

#         agent_list = ['PLAYSTAR', 'platform', 'QATEST', 'TEST', 'TEST-IDR', 'TEST-MYR']
#         for agent_ in agent_list:
#             self.driver.find_element(By.ID, 'agent').click()  # 代理商
#             time.sleep(1)
            
#             agent_name = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="{agent_}"]')  # PLAYSTAR / Platform / QATEST / Test / Test-IDR / Test-MYR
#             agent_name.click() 
#             print('代理商:', agent_name.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 篩選 **************************

#         filter_key_list = ['allWin', '2stPrize', '5stPrize', '8stPrize', 'GoodLuck']
#         for filter_ in filter_key_list:
#             self.driver.find_element(By.ID, 'filter_key').click()  # 篩選
#             time.sleep(1)
            
#             filter_name = self.driver.find_element(By.XPATH, f'//*[@id="filter_key"]/option[@value="{filter_}"]')  # All 奖 / 2stPrize / 5stPrize / 8stPrize / GoodLuck
#             filter_name.click() 
#             print('篩選:', filter_name.text, '\n')
#             time.sleep(1)

#         self.driver.refresh()


# #  ************************** 中獎名單(抽獎)搜尋(派彩時間) **************************

#     def test_winner_list_search_func_identify_pt_zhCN(self):
#         '''中獎名單(抽獎)搜尋(派彩時間)功能驗證'''    

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div[1]/div/div/span/button/i').click()  # 起始時間
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[7]/a').click()  # 2024/12/01
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('起始時間:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         close_btn = self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]')  # 關閉
#         self.driver.execute_script("$(arguments[0]).click()", close_btn)
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/span/button/i').click()  # 結束時間
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[6]/td[2]/a').click()  # 2024/12/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('結束時間:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         close_btn = self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]')  # 關閉
#         self.driver.execute_script("$(arguments[0]).click()", close_btn)
#         time.sleep(1)   

#         self.driver.find_element(By.ID, f'search_time_type').click()  # 搜尋類別
#         time.sleep(1) 

#         agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="search_time_type"]/option[1]')  # 派彩時間
#         agentTypeSelect.click()
#         print('搜尋類別:', agentTypeSelect.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'count').click()  # 顯示筆數
#         time.sleep(1)

#         show_count_num = self.driver.find_element(By.XPATH, f'//*[@id="count"]/option[3]')  # 100 
#         show_count_num.click() 
#         print('顯示筆數:', show_count_num.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'filter_key').click()  # 篩選
#         time.sleep(1)

#         filter_name = self.driver.find_element(By.XPATH, f'//*[@id="filter_key"]/option[@value="allWin"]')  # All 奖 
#         filter_name.click() 
#         print('篩選:', filter_name.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'jackpot_info_sh_btn').click()  
#         time.sleep(1)

#         page_end = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[4]')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#         time.sleep(2)

#         page_head = self.driver.find_element(By.XPATH, '//*[@id="export"]') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head) # 匯出
#         page_head.click()
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 中獎名單(抽獎)搜尋(結帳時間) **************************

#     def test_winner_list_search_func_identify_ct_zhCN(self):
#         '''中獎名單(抽獎)搜尋(結帳時間)功能驗證'''  

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div[1]/div/div/span/button/i').click()  # 起始時間
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[7]/a').click()  # 2024/12/01
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('起始時間:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         close_btn = self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]')  # 關閉
#         self.driver.execute_script("$(arguments[0]).click()", close_btn)
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/span/button/i').click()  # 結束時間
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[6]/td[2]/a').click()  # 2024/12/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('結束時間:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         close_btn = self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]')  # 關閉
#         self.driver.execute_script("$(arguments[0]).click()", close_btn)
#         time.sleep(1)  

#         self.driver.find_element(By.ID, f'search_time_type').click()  # 搜尋類別
#         time.sleep(1) 

#         agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="search_time_type"]/option[2]')  # 結帳時間
#         agentTypeSelect.click()
#         print('搜尋類別:', agentTypeSelect.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'count').click()  # 顯示筆數
#         time.sleep(1)

#         show_count_num = self.driver.find_element(By.XPATH, f'//*[@id="count"]/option[3]')  # 100 
#         show_count_num.click() 
#         print('顯示筆數:', show_count_num.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'filter_key').click()  # 篩選
#         time.sleep(1)

#         filter_name = self.driver.find_element(By.XPATH, f'//*[@id="filter_key"]/option[@value="1stPrize"]')  # 第一獎 
#         filter_name.click() 
#         print('篩選:', filter_name.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'jackpot_info_sh_btn').click()   
#         time.sleep(1)

#         page_end = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[4]')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#         time.sleep(2)

#         page_head = self.driver.find_element(By.XPATH, '//*[@id="export"]') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head) # 匯出
#         page_head.click()
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# # --------------------------- 中獎名單(抽獎)(EN) ---------------------------

#     def test_Event_Winner_List_Lottery_EN(self):
#         '''【活動】中獎名單(抽獎)語系切換(英)''' 

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[1]').click()
#         print("語系已切換:", 'English', '\n')
#         time.sleep(1)


# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_wl_EN(self):
#         '''搜尋列功能驗證(英)'''


# # ************************** Start Time **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div[1]/div/div/span/button/i').click()  # Start Time
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[1]/a').click()  # 2024/12/01
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('Start Time:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]').click()  # 關閉
#         time.sleep(1)

  
# # ************************** End Time **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/span/button/i').click()  # End Time
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[3]/a').click()  # 2024/12/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('End Time:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]').click()  # 關閉
#         time.sleep(1)   

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** Search Type **************************

#         for search_class in range(1, 3):
#             self.driver.find_element(By.ID, f'search_time_type').click()  # Search Type
#             time.sleep(1)
            
#             agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="search_time_type"]/option[{search_class}]')  # Payout Time / Checkout Time
#             agentTypeSelect.click()
#             print('Search Type:', agentTypeSelect.text, '\n')
#             time.sleep(1)
            
#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** Show Count **************************

#         show_count_list = [2, 3, 6]
#         for show_count in show_count_list:
            
#             self.driver.find_element(By.ID, 'count').click()  # Show Count
#             time.sleep(1)
            
#             show_count_num = self.driver.find_element(By.XPATH, f'//*[@id="count"]/option[{show_count}]')  # 50 / 100 / 1000
#             show_count_num.click() 
#             print('Show Count:', show_count_num.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** Agent Class **************************

#         agent_class_list = ["All", 'Platform', 'PS', 'Test']
#         for agent_class in agent_class_list:
            
#             self.driver.find_element(By.ID, 'agent_attr').click()  # Agent Class
#             time.sleep(1)
            
#             agent_class_name = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="{agent_class}"]')  # -- Select All -- / Platform / PS / Test
#             agent_class_name.click() 
#             print('Agent Class:', agent_class_name.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** Agent **************************

#         agent_list = ['PLAYSTAR', 'platform', 'QATEST', 'TEST', 'TEST-IDR', 'TEST-MYR']
#         for agent_ in agent_list:
#             self.driver.find_element(By.ID, 'agent').click()  # Agent
#             time.sleep(1)
            
#             agent_name = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="{agent_}"]')  # PLAYSTAR / Platform / QATEST / Test / Test-IDR / Test-MYR
#             agent_name.click() 
#             print('Agent:', agent_name.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** Filter **************************

#         filter_key_list = ['allWin', '2stPrize', '5stPrize', '8stPrize', 'GoodLuck']
#         for filter_ in filter_key_list:
#             self.driver.find_element(By.ID, 'filter_key').click()  # Filter
#             time.sleep(1)
            
#             filter_name = self.driver.find_element(By.XPATH, f'//*[@id="filter_key"]/option[@value="{filter_}"]')  # All / 2ndPrize / 5stPrize / 8thPrize / GoodLuck
#             filter_name.click() 
#             print('Filter:', filter_name.text, '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** Winner List(Lottery)搜尋(Payout Time) **************************

#     def test_winner_list_search_func_identify_pt_EN(self):
#         '''中獎名單(抽獎)搜尋(派彩時間)功能驗證(英)'''    

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div[1]/div/div/span/button/i').click()  # Start Time
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[1]/a').click()  # 2024/12/01
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('Start Time:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         close_btn = self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]')  # Close
#         self.driver.execute_script("$(arguments[0]).click()", close_btn)
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/span/button/i').click()  # End Time
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[3]/a').click()  # 2024/12/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('End Time:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         close_btn = self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]')  # Close
#         self.driver.execute_script("$(arguments[0]).click()", close_btn)
#         time.sleep(1)  

#         self.driver.find_element(By.ID, f'search_time_type').click()  # Search Type
#         time.sleep(1)

#         agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="search_time_type"]/option[1]')  # Payout Time
#         agentTypeSelect.click()
#         print('Search Type:', agentTypeSelect.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'count')  # Show Count
#         time.sleep(1)

#         show_count_num = self.driver.find_element(By.XPATH, f'//*[@id="count"]/option[3]')  # 100 
#         show_count_num.click() 
#         print('Show Count:', show_count_num.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'filter_key').click()  # Filter
#         time.sleep(1)

#         filter_name = self.driver.find_element(By.XPATH, f'//*[@id="filter_key"]/option[@value="allWin"]')  # All 奖 
#         filter_name.click() 
#         print('Filter:', filter_name.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'jackpot_info_sh_btn').click()  
#         time.sleep(1)

#         page_end = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[4]')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#         time.sleep(2)

#         page_head = self.driver.find_element(By.XPATH, '//*[@id="export"]') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head) # 匯出
#         page_head.click()
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** Winner List(Lottery)搜尋(Checkout Time) **************************

#     def test_winner_list_search_func_identify_ct_EN(self):
#         '''中獎名單(抽獎)搜尋(結帳時間)功能驗證(英)'''  

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div[1]/div/div/span/button/i').click()  # Start Time
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[1]/a').click()  # 2024/12/01
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('Start Time:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         close_btn = self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]')  # Close
#         self.driver.execute_script("$(arguments[0]).click()", close_btn)
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/span/button/i').click()  # End Time
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[3]/a').click()  # 2024/12/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('End Time:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         close_btn = self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]')  # Close
#         self.driver.execute_script("$(arguments[0]).click()", close_btn)
#         time.sleep(1)
        
#         self.driver.find_element(By.ID, f'search_time_type').click()  # Search Type
#         time.sleep(1) 

#         agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="search_time_type"]/option[2]')  # Checkout Time
#         agentTypeSelect.click()
#         print('Search Type:', agentTypeSelect.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'count').click()  # Show Count
#         time.sleep(1)

#         show_count_num = self.driver.find_element(By.XPATH, f'//*[@id="count"]/option[3]')  # 100 
#         show_count_num.click() 
#         print('Show Count:', show_count_num.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'filter_key').click()  # Filter
#         time.sleep(1)

#         filter_name = self.driver.find_element(By.XPATH, f'//*[@id="filter_key"]/option[@value="1stPrize"]')  # 1stPrize 
#         filter_name.click() 
#         print('Filter:', filter_name.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'jackpot_info_sh_btn').click()   
#         time.sleep(1)

#         page_end = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[4]')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#         time.sleep(2)

#         page_head = self.driver.find_element(By.XPATH, '//*[@id="export"]') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head) # 匯出
#         page_head.click()
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# # --------------------------- 中獎名單(抽獎)(Tai) ---------------------------

#     def test_Event_Winner_List_Lottery_Tai(self):
#         '''【活動】中獎名單(抽獎)語系切換(泰)''' 

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[4]').click()
#         print("語系已切換:", 'ไทย', '\n')
#         time.sleep(1)


# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_wl_Tai(self):
#         '''搜尋列功能驗證(泰)'''


# # ************************** เวลาเริ่มต้น **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div[1]/div/div/span/button/i').click()  # เวลาเริ่มต้น
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[1]/a').click()  # 2024/12/01
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('เวลาเริ่มต้น:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2] ').click()  # 關閉
#         time.sleep(1)

  
# # ************************** เวลาสิ้นสุด **************************

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/span/button/i').click()  # เวลาสิ้นสุด
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[3]/a').click()  # 2024/12/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('เวลาสิ้นสุด:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2] ').click()  # 關閉
#         time.sleep(1)  

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** หมวดหมู่การค้นหา **************************

#         for search_class in range(1, 3):
#             self.driver.find_element(By.ID, f'search_time_type').click()  # หมวดหมู่การค้นหา
#             time.sleep(1)
            
#             agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="search_time_type"]/option[{search_class}]')  # เวลาจ่ายเงิน / เวลาชำระเงิน
#             agentTypeSelect.click()
#             print('หมวดหมู่การค้นหา:', agentTypeSelect.text, '\n')
#             time.sleep(1)
            
#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** จำนวนครังที่แสดง **************************

#         show_count_list = [2, 3, 6]
#         for show_count in show_count_list:
            
#             self.driver.find_element(By.ID, 'count').click()  # จำนวนครังที่แสดง
#             time.sleep(1)
            
#             show_count_num = self.driver.find_element(By.XPATH, f'//*[@id="count"]/option[{show_count}]')  # 50 / 100 / 1000
#             show_count_num.click() 
#             print('จำนวนครังที่แสดง:', show_count_num.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** หมวดตัวแทน **************************

#         agent_class_list = ["All", 'Platform', 'PS', 'Test']
#         for agent_class in agent_class_list:
            
#             self.driver.find_element(By.ID, 'agent_attr').click()  # หมวดตัวแทน
#             time.sleep(1)
            
#             agent_class_name = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="{agent_class}"]')  # -- เลือกทั้งหมด -- / Platform / PS / Test
#             agent_class_name.click() 
#             print('หมวดตัวแทน:', agent_class_name.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** ตัวแทน **************************

#         agent_list = ['PLAYSTAR', 'platform', 'QATEST', 'TEST', 'TEST-IDR', 'TEST-MYR']
#         for agent_ in agent_list:
#             self.driver.find_element(By.ID, 'agent').click()  # Agent
#             time.sleep(1)
            
#             agent_name = self.driver.find_element(By.XPATH, f'//*[@id="agent"]/option[@value="{agent_}"]')  # PLAYSTAR / Platform / QATEST / Test / Test-IDR / Test-MYR
#             agent_name.click() 
#             print('ตัวแทน:', agent_name.get_attribute('value'), '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** กรอง **************************

#         filter_key_list = ['allWin', '2stPrize', '5stPrize', '8stPrize', 'GoodLuck']
#         for filter_ in filter_key_list:
#             self.driver.find_element(By.ID, 'filter_key').click()  # Filter
#             time.sleep(1)
            
#             filter_name = self.driver.find_element(By.XPATH, f'//*[@id="filter_key"]/option[@value="{filter_}"]')  # All / รางวัลที่ 2 / 5stPrize / 8thPrize / GoodLuck
#             filter_name.click() 
#             print('กรอง:', filter_name.text, '\n')
#             time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** รายชื่อผู้ชนะ (ลอตเตอรี่)搜尋(เวลาจ่ายเงิน) **************************

#     def test_winner_list_search_func_identify_pt_Tai(self):
#         '''中獎名單(抽獎)搜尋(派彩時間)功能驗證(泰)'''    

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div[1]/div/div/span/button/i').click()  # เวลาเริ่มต้น
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[1]/a').click()  # 2024/12/01
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('เวลาเริ่มต้น:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         close_btn = self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]')  # Close
#         self.driver.execute_script("$(arguments[0]).click()", close_btn)
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/span/button/i').click()  # เวลาสิ้นสุด
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[3]/a').click()  # 2024/12/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('เวลาสิ้นสุด:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         close_btn = self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]')  # Close
#         self.driver.execute_script("$(arguments[0]).click()", close_btn)
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'search_time_type').click()  # หมวดหมู่การค้นหา
#         time.sleep(1)

#         agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="search_time_type"]/option[1]')  # เวลาจ่ายเงิน
#         agentTypeSelect.click()
#         print('หมวดหมู่การค้นหา:', agentTypeSelect.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'count').click()  # จำนวนครังที่แสดง
#         time.sleep(1)

#         show_count_num = self.driver.find_element(By.XPATH, f'//*[@id="count"]/option[3]')  # 100 
#         show_count_num.click() 
#         print('จำนวนครังที่แสดง:', show_count_num.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'filter_key').click()  # กรอง
#         time.sleep(1)

#         filter_name = self.driver.find_element(By.XPATH, f'//*[@id="filter_key"]/option[2]')  # All 奖 
#         filter_name.click() 
#         print('กรอง:', filter_name.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'jackpot_info_sh_btn').click()   
#         time.sleep(1)

#         page_end = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[4]')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#         time.sleep(2)

#         page_head = self.driver.find_element(By.XPATH, '//*[@id="export"]') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head) # 匯出
#         page_head.click()
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** รายชื่อผู้ชนะ (ลอตเตอรี่)搜尋(เวลาชำระเงิน) **************************

#     def test_winner_list_search_func_identify_ct_Tai(self):
#         '''中獎名單(抽獎)搜尋(結帳時間)功能驗證(泰)'''  

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div[1]/div/div/span/button/i').click()  # เวลาเริ่มต้น
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[1]/a').click()  # 2024/12/01
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="start"]')
#         print('เวลาเริ่มต้น:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         close_btn = self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]') # Close
#         self.driver.execute_script("$(arguments[0]).click()", close_btn)
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/span/button/i').click()  # เวลาสิ้นสุด
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[3]/a').click()  # 2024/12/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="end"]')
#         print('เวลาสิ้นสุด:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         close_btn = self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]')  # Close
#         self.driver.execute_script("$(arguments[0]).click()", close_btn)
#         time.sleep(1)  

#         self.driver.find_element(By.ID, f'search_time_type').click()  # หมวดหมู่การค้นหา
#         time.sleep(1)

#         agentTypeSelect = self.driver.find_element(By.XPATH, f'//*[@id="search_time_type"]/option[2]')  # เวลาชำระเงิน
#         agentTypeSelect.click()
#         print('หมวดหมู่การค้นหา:', agentTypeSelect.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'count').click()  # จำนวนครังที่แสดง
#         time.sleep(1)

#         show_count_num = self.driver.find_element(By.XPATH, f'//*[@id="count"]/option[3]')  # 100 
#         show_count_num.click() 
#         print('จำนวนครังที่แสดง:', show_count_num.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'filter_key').click()  # กรอง
#         time.sleep(1)

#         filter_name = self.driver.find_element(By.XPATH, f'//*[@id="filter_key"]/option[3]')  # 1stPrize 
#         filter_name.click() 
#         print('กรอง:', filter_name.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'jackpot_info_sh_btn').click()
#         time.sleep(1)

#         page_end = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[4]')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#         time.sleep(2)

#         page_head = self.driver.find_element(By.XPATH, '//*[@id="export"]') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head) # 匯出
#         page_head.click()
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# # ============================= 後台功能巡測【活動】============================= 
# # *************************** 紅包活動設定 ***************************

#     def test_Event_Red_Envelope_Event_Setting_zhCN(self):
#         '''【活動】紅包活動設定功能頁切換'''     

#         self.driver.find_element(By.ID, 'Event2').click()
#         print("進入活動功能選單!", '\n')
#         time.sleep(1)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Scratch/Red_Envelope_Event_Setting')
#         print("切換紅包活動設定選單!", '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[3]').click()
#         print("語系已切換'简体中文'!", '\n')
#         time.sleep(1)
        

# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_ree_zhCN(self):
#         '''搜尋列功能驗證'''

        
# # ************************** 活動開始時間 **************************

#         self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[2]/div[2]/div[1]/div/div/span/button').click()  # 活動開始時間
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[7]/a').click()  # 2024/12/01
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="search_start"]')
#         print('活動開始時間:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         starttime_close = self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]')  # 關閉
#         self.driver.execute_script("$(arguments[0]).click()", starttime_close)
#         time.sleep(1)

  
# # ************************** 活動結束時間 **************************

#         self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/span/button').click()  # 活動結束時間
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[6]/td[2]/a')  # 2024/12/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="search_end"]')
#         print('活動結束時間:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         endtime_close = self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]')  # 關閉
#         self.driver.execute_script("$(arguments[0]).click()", endtime_close)
#         time.sleep(1)
        
#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 幣別 **************************

#         self.driver.find_element(By.ID, f'search_currency').click()  # 幣別
#         time.sleep(1)
        
#         currency_list = ['CNY', 'IDR', 'JPY', 'THB']
#         for currency_value in currency_list:
#             currency_slt = self.driver.find_element(By.XPATH, f'//*[@id="search_currency"]/option[@value="{currency_value}"]')  # CNY / IDR / JPY / THB
#             currency_slt.click()
#             print('幣別:', currency_slt.get_attribute('value'), '\n')
#             time.sleep(1)
            
#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 代理商 **************************

#     def test_agent_with_currency_setting_zhCN(self):
#         '''代理商與幣別設定功能驗證'''

#         self.driver.find_element(By.ID, f'search_currency').click()  # 幣別
#         time.sleep(1)
        
#         currency_slt = self.driver.find_element(By.XPATH, f'//*[@id="search_currency"]/option[@value="CNY"]')  # CNY
#         currency_slt.click()
#         print('幣別:', currency_slt.get_attribute('value'), '\n')
#         time.sleep(1) 

#         self.driver.find_element(By.ID, f'select_host').click()  # 代理商
#         time.sleep(1)

#         agent_list = ['65921-selectable', '-1632647290-selectable', '2571410-selectable', '79713760-selectable']
#         for agent_list_id in agent_list:
#             agent_slt = self.driver.find_element(By.XPATH, f'//*[@id="{agent_list_id}"]')  # -- 全選 -- / PLAYSTAR / Test / Test-2                                            
#             agent_slt.click()
#             print('(勾選)代理商:', agent_slt.text, '\n')
#             time.sleep(1)
            
#             self.driver.find_element(By.ID, 'sh_btn').click()  # 送出  
#             time.sleep(1)

#             page_end = self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[3]')  
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#             time.sleep(2)

#             page_head = self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[1]/h5')  
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 搜尋
#             page_head.click()
#             time.sleep(2)
            
#             agent_slt_cls = self.driver.find_element(By.XPATH, f'//*[@id="{agent_list_id}"]')  # (取消) -- 全選 -- / PLAYSTAR / Test / Test-2                                                            
#             agent_slt_cls.click()
#             print('(取消勾選)代理商:', agent_slt_cls.text, '\n')
#             time.sleep(1)
            
#         self.driver.refresh()
#         time.sleep(1)
        

# #  ************************** 參加活動遊戲 **************************

#     def test_participate_games_setting_zhCN(self):
#         '''參加活動遊戲設定功能驗證'''

#         game_id_list = ['1839777149-selectable', '1839777156-selectable', '1839777187-selectable']
#         for game_list_id in game_id_list:
#             self.driver.find_element(By.ID, f'select_game').click()  # 活動遊戲
#             time.sleep(1) 
            
#             game_slt = self.driver.find_element(By.XPATH, f'//*[@id="{game_list_id}"]')  # PSS-ON-00141 麻將胡了3 / PSS-ON-00148 爆爆糖果 / PSS-ON-00158 赛博魔方                                                        
#             game_slt.click()
#             print('(勾選)活動遊戲:', game_slt.text, '\n')
#             time.sleep(1)
            
#             self.driver.find_element(By.ID, 'sh_btn').click()  # 送出   
#             time.sleep(1)
            
#             self.driver.find_element(By.XPATH, f'//*[@id="search_event_result"]/tbody/tr[1]/td[5]/button').click()  # 查看    
#             time.sleep(1)
                
#             game_name = self.driver.find_element(By.XPATH, f'//*[@id="check_cost_game"]/div/div/div[2]')
#             game_name.click()
#             print(game_name.text, '\n')
#             time.sleep(1)
            
#             self.driver.find_element(By.XPATH, f'//*[@id="check_cost_game"]/div/div/div[3]/button').click()
#             time.sleep(1)
                
#             page_end = self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[3]')  
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#             time.sleep(2)
            
#             page_head = self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[1]/h5')  
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 頁首
#             time.sleep(1)
            
#             game_slt_cls = self.driver.find_element(By.XPATH, f'//*[@id="{game_list_id}"]')  # PSS-ON-00141 麻將胡了3 / PSS-ON-00148 爆爆糖果 / PSS-ON-00158 赛博魔方                                                         
#             game_slt_cls.click()
#             print('(取消勾選)活動遊戲:', game_slt_cls.text, '\n')
#             time.sleep(1)
                                                                                                                                                                            
#         self.driver.refresh()
#         time.sleep(1)                                                            
                                                            

# #  ************************** 紅包活動設定紀錄搜尋 **************************

#     def test_red_envelope_event_setting_identify_zhCN(self):
#         '''紅包活動設定功能驗證'''

#         self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[2]/div[2]/div[1]/div/div/span/button').click()  # 活動開始時間
#         time.sleep(1)

#         for _ in range(7):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 
        
#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[5]/a').click()  # 2024/11/01
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="search_start"]')
#         print('活動開始時間:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         starttime_close = self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]')  # 關閉
#         self.driver.execute_script("$(arguments[0]).click()", starttime_close)
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/span/button').click()  # 活動結束時間
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[6]/td[2]/a').click()  # 2024/12/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="search_end"]')
#         print('活動結束時間:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         endtime_close = self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]').click() # 關閉
#         self.driver.execute_script("$(arguments[0]).click()", endtime_close)
#         time.sleep(1) 

#         self.driver.find_element(By.ID, f'search_currency').click()  # 幣別
#         time.sleep(1) 

#         currency_slt = self.driver.find_element(By.XPATH, f'//*[@id="search_currency"]/option[@value="CNY"]')  # CNY
#         currency_slt.click()
#         print('幣別:', currency_slt.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'select_host').click()  # 代理商
#         time.sleep(1) 

#         agent_slt = self.driver.find_element(By.XPATH, f'//*[@id="2571410-selectable"]')  # Test                                                            
#         agent_slt.click()
#         print('(勾選)代理商:', agent_slt.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'select_game').click()  # 活動遊戲
#         time.sleep(1)

#         game_select = self.driver.find_element(By.XPATH, f'//*[@id="1839777187-selectable"]')  # PSS-ON-00158 赛博魔方                                                      
#         game_select.click()
#         print('(勾選)活動遊戲:', game_select.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'sh_btn').click()    
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="search_event_result"]/tbody/tr/td[5]/button').click()
#         time.sleep(1)

#         game_name = self.driver.find_element(By.XPATH, f'//*[@id="check_cost_game"]/div/div/div[2]')
#         game_name.click()
#         print(game_name.text, '\n')
#         time.sleep(1)
                
#         self.driver.find_element(By.XPATH, f'//*[@id="check_cost_game"]/div/div/div[3]/button').click()
#         time.sleep(1)
                
#         page_end = self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[3]')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#         time.sleep(2)

#         page_head = self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[1]/h5')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 頁首
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1) 


# # --------------------------- 紅包活動設定(EN) ---------------------------

#     def test_Event_Red_Envelope_Event_Setting_EN(self):
#         '''【活動】紅包活動設定功能頁切換(英)'''

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[1]').click()
#         print("語系已切換:", 'English', '\n')
#         time.sleep(1)


# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_ree_EN(self):
#         '''搜尋列功能驗證(英)'''        
        

# # ************************** Event Start Time **************************

#         self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[2]/div[2]/div[1]/div/div/span/button').click()  # Event Start Time
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[1]/a').click()  # 2024/12/01
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="search_start"]')
#         print('Event Start Time:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         starttime_close = self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]')  # Done
#         self.driver.execute_script("$(arguments[0]).click()", starttime_close)
#         time.sleep(1)

  
# # ************************** Event End Time **************************

#         self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/span/button').click()  # Event End Time
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[3]/a').click()  # 2024/12/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="search_end"]')
#         print('Event End Time:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         endtime_close = self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]')  # Done
#         self.driver.execute_script("$(arguments[0]).click()", endtime_close)
#         time.sleep(1)   

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** Currency **************************

#         self.driver.find_element(By.ID, f'search_currency').click()  # Currency
#         time.sleep(1)
        
#         currency_list = ['CNY', 'IDR', 'JPY', 'THB']
#         for currency_value in currency_list:
#             currency_slt = self.driver.find_element(By.XPATH, f'//*[@id="search_currency"]/option[@value="{currency_value}"]')  # CNY / IDR / JPY / THB
#             currency_slt.click()
#             print('Currency:', currency_slt.get_attribute('value'), '\n')
#             time.sleep(1)
            
#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** Agent **************************

#     def test_agent_with_currency_setting_EN(self):
#         '''代理商與幣別設定功能驗證(英)'''
    
#         self.driver.find_element(By.ID, f'search_currency').click()  # Currency
#         time.sleep(1)
        
#         currency_slt = self.driver.find_element(By.XPATH, f'//*[@id="search_currency"]/option[@value="CNY"]')  # CNY
#         currency_slt.click()
#         print('Currency:', currency_slt.get_attribute('value'), '\n')
#         time.sleep(1) 

#         self.driver.find_element(By.ID, f'select_host').click()  # Agent
#         time.sleep(1)

#         agent_list = ['65921-selectable', '-1632647290-selectable', '2571410-selectable', '79713760-selectable']
#         for agent_list_id in agent_list:
#             agent_slt = self.driver.find_element(By.XPATH, f'//*[@id="{agent_list_id}"]')  # -- Select All -- / PLAYSTAR / Test / Test-2                                             
#             agent_slt.click()
#             print('(勾選)Agent:', agent_slt.text, '\n')
#             time.sleep(1)
            
#             self.driver.find_element(By.ID, 'sh_btn').click()  # Submit  
#             time.sleep(1)

#             page_end = self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[3]')  
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#             time.sleep(2)

#             page_head = self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[1]/h5')  
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 搜尋
#             page_head.click()
#             time.sleep(2)
            
#             agent_slt_cls = self.driver.find_element(By.XPATH, f'//*[@id="{agent_list_id}"]')  # (取消) -- Select All -- / PLAYSTAR / Test / Test-2                                                             
#             agent_slt_cls.click()
#             print('(取消勾選)Agent:', agent_slt_cls.text, '\n')
#             time.sleep(1)
            
#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 參加活動遊戲 **************************

#     def test_participate_games_setting_EN(self):
#         '''參加活動遊戲設定功能驗證(英)'''
        
#         game_id_list = ['1839777149-selectable', '1839777156-selectable', '1839777187-selectable']
#         for game_list_id in game_id_list:
#             self.driver.find_element(By.ID, f'select_game').click()  # Participate Games
#             time.sleep(1) 
            
#             game_slt = self.driver.find_element(By.XPATH, f'//*[@id="{game_list_id}"]')  # PSS-ON-00141 MAHJONG WAYS 3 / PSS-ON-00148 SUGAR BOOM / PSS-ON-00158 CYBER CUBE                                                        
#             game_slt.click()
#             print('(勾選)Participate Games:', game_slt.text, '\n')
#             time.sleep(1)
            
#             self.driver.find_element(By.ID, 'sh_btn').click()  # 送出   
#             time.sleep(1)
            
#             self.driver.find_element(By.XPATH, f'//*[@id="search_event_result"]/tbody/tr[1]/td[5]/button').click()  # Check    
#             time.sleep(1)
                
#             game_name = self.driver.find_element(By.XPATH, f'//*[@id="check_cost_game"]/div/div/div[2]')
#             game_name.click()
#             print(game_name.text, '\n')
#             time.sleep(1)
            
#             self.driver.find_element(By.XPATH, f'//*[@id="check_cost_game"]/div/div/div[3]/button').click()
#             time.sleep(1)
                
#             page_end = self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[3]')  
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#             time.sleep(2)
            
#             page_head = self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[1]/h5')  
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 頁首
#             time.sleep(1)
            
#             game_slt_cls = self.driver.find_element(By.XPATH, f'//*[@id="{game_list_id}"]')  # PSS-ON-00141 MAHJONG WAYS 3 / PSS-ON-00148 SUGAR BOOM / PSS-ON-00158 CYBER CUBE                                                        
#             game_slt_cls.click()
#             print('(取消勾選)Participate Games:', game_slt_cls.text, '\n')
#             time.sleep(1)
                                                                                                                                                                            
#         self.driver.refresh()
#         time.sleep(1)        


# #  ************************** 紅包活動設定搜尋 **************************

#     def test_red_envelope_event_setting_identify_EN(self):
#         '''紅包活動設定功能驗證(英)'''
    
#         self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[2]/div[2]/div[1]/div/div/span/button').click()  # Event Start Time
#         time.sleep(1)

#         for _ in range(7):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
        
#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[6]/a').click()  # 2024/11/01
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="search_start"]')
#         print('Event Start Time:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         starttime_close = self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]')  # Done
#         self.driver.execute_script("$(arguments[0]).click()", starttime_close)
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/span/button').click()  # Event End Time
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[3]/a').click()  # 2024/12/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="search_end"]')
#         print('Event End Time:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         endtime_close = self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]').click() # Done
#         self.driver.execute_script("$(arguments[0]).click()", endtime_close)
#         time.sleep(1) 

#         self.driver.find_element(By.ID, f'search_currency').click()  # Currency
#         time.sleep(1) 

#         currency_slt = self.driver.find_element(By.XPATH, f'//*[@id="search_currency"]/option[@value="CNY"]')  # CNY
#         currency_slt.click()
#         print('Currency:', currency_slt.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'select_host').click()  # Agent
#         time.sleep(1) 

#         agent_slt = self.driver.find_element(By.XPATH, f'//*[@id="2571410-selectable"]')  # Test                                                            
#         agent_slt.click()
#         print('(勾選)Agent:', agent_slt.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'select_game').click()  # Participate Games
#         time.sleep(1)

#         game_select = self.driver.find_element(By.XPATH, f'//*[@id="1839777187-selectable"]')  # PSS-ON-00158 CYBER CUBE                                                      
#         game_select.click()
#         print('(勾選)Participate Games:', game_select.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'sh_btn').click()    
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="search_event_result"]/tbody/tr/td[5]/button').click()
#         time.sleep(1)

#         game_name = self.driver.find_element(By.XPATH, f'//*[@id="check_cost_game"]/div/div/div[2]')
#         game_name.click()
#         print(game_name.text, '\n')
#         time.sleep(1)
                
#         self.driver.find_element(By.XPATH, f'//*[@id="check_cost_game"]/div/div/div[3]/button').click()
#         time.sleep(1)
                
#         page_end = self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[3]')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#         time.sleep(2)

#         page_head = self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[1]/h5')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 頁首
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1) 


# # --------------------------- 紅包活動設定(Tai) ---------------------------

#     def test_Event_Red_Envelope_Event_Setting_Tai(self):
#         '''【活動】紅包活動設定功能頁切換(泰)'''

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[4]').click()
#         print("語系已切換:", 'ไทย', '\n')
#         time.sleep(1)


# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_ree_Tai(self):
#         '''搜尋列功能驗證(泰)'''            


# # ************************** 活動開始時間 **************************

#         self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[2]/div[2]/div[1]/div/div/span/button').click()  # 活動開始時間
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[1]/a').click()  # 2024/12/01
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="search_start"]')
#         print('เวลาเริ่มต้นกิจกรรม:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         starttime_close = self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]')  # Done
#         self.driver.execute_script("$(arguments[0]).click()", starttime_close)
#         time.sleep(1)

  
# # ************************** 活動結束時間 **************************

#         self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/span/button').click()  # 活動結束時間
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[3]/a').click()  # 2024/12/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="search_end"]')
#         print('เวลาสิ้นสุดกิจกรรม:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         endtime_close = self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]')  # Done
#         self.driver.execute_script("$(arguments[0]).click()", endtime_close)
#         time.sleep(1)   

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 幣別 **************************

#         self.driver.find_element(By.ID, f'search_currency').click()  # 幣別
#         time.sleep(1)
        
#         currency_list = ['CNY', 'IDR', 'JPY', 'THB']
#         for currency_value in currency_list:
#             currency_slt = self.driver.find_element(By.XPATH, f'//*[@id="search_currency"]/option[@value="{currency_value}"]')  # CNY / IDR / JPY / THB
#             currency_slt.click()
#             print('สกุลเงิน:', currency_slt.get_attribute('value'), '\n')
#             time.sleep(1)
            
#         self.driver.refresh()
#         time.sleep(1)
    

# #  ************************** 代理商 **************************

#     def test_agent_with_currency_setting_Tai(self):
#         '''代理商與幣別設定功能驗證(泰)'''    

#         self.driver.find_element(By.ID, f'search_currency').click()  # Currency
#         time.sleep(1)
        
#         currency_slt = self.driver.find_element(By.XPATH, f'//*[@id="search_currency"]/option[@value="CNY"]')  # CNY
#         currency_slt.click()
#         print('สกุลเงิน:', currency_slt.get_attribute('value'), '\n')
#         time.sleep(1) 

#         self.driver.find_element(By.ID, f'select_host').click()  # Agent
#         time.sleep(1)

#         agent_list = ['65921-selectable', '-1632647290-selectable', '2571410-selectable', '79713760-selectable']
#         for agent_list_id in agent_list:
#             agent_slt = self.driver.find_element(By.XPATH, f'//*[@id="{agent_list_id}"]')  # -- เลือกทั้งหมด -- / PLAYSTAR / Test / Test-2                                             
#             agent_slt.click()
#             print('(勾選)ตัวแทน:', agent_slt.text, '\n')
#             time.sleep(1)
            
#             self.driver.find_element(By.ID, 'sh_btn').click()  # Submit  
#             time.sleep(1)

#             page_end = self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[3]')  
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#             time.sleep(2)

#             page_head = self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[1]/h5')  
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 搜尋
#             page_head.click()
#             time.sleep(2)
            
#             agent_slt_cls = self.driver.find_element(By.XPATH, f'//*[@id="{agent_list_id}"]')  # (取消) -- เลือกทั้งหมด -- / PLAYSTAR / Test / Test-2                                                             
#             agent_slt_cls.click()
#             print('(取消勾選)ตัวแทน:', agent_slt_cls.text, '\n')
#             time.sleep(1)
            
#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 參加活動遊戲 **************************

#     def test_participate_games_setting_Tai(self):
#         '''參加活動遊戲設定功能驗證(泰)'''

#         game_id_list = ['1839777149-selectable', '1839777156-selectable', '1839777187-selectable']
#         for game_list_id in game_id_list:
#             self.driver.find_element(By.ID, f'select_game').click()  # Participate Games
#             time.sleep(1) 
            
#             game_slt = self.driver.find_element(By.XPATH, f'//*[@id="{game_list_id}"]')  # PSS-ON-00141 เส้นทางมาจอง 3 / PSS-ON-00148 SUGAR BOOM / PSS-ON-00158 ไซเบอร์คิวบ์                                                        
#             game_slt.click()
#             print('(勾選)เข้าร่วมเกม:', game_slt.text, '\n')
#             time.sleep(1)
            
#             self.driver.find_element(By.ID, 'sh_btn').click()  # 送出   
#             time.sleep(1)
            
#             self.driver.find_element(By.XPATH, f'//*[@id="search_event_result"]/tbody/tr[1]/td[5]/button').click()  # Check    
#             time.sleep(1)
                
#             game_name = self.driver.find_element(By.XPATH, f'//*[@id="check_cost_game"]/div/div/div[2]')
#             game_name.click()
#             print(game_name.text, '\n')
#             time.sleep(1)
            
#             self.driver.find_element(By.XPATH, f'//*[@id="check_cost_game"]/div/div/div[3]/button').click()
#             time.sleep(1)
                
#             page_end = self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[3]')  
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#             time.sleep(2)
            
#             page_head = self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[1]/h5')  
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 頁首
#             time.sleep(1)
            
#             game_slt_cls = self.driver.find_element(By.XPATH, f'//*[@id="{game_list_id}"]')  # PSS-ON-00141 เส้นทางมาจอง 3 / PSS-ON-00148 SUGAR BOOM / PSS-ON-00158 ไซเบอร์คิวบ์                                                        
#             game_slt_cls.click()
#             print('(取消勾選)เข้าร่วมเกม:', game_slt_cls.text, '\n')
#             time.sleep(1)
                                                                                                                                                                            
#         self.driver.refresh()
#         time.sleep(1)                                                         
                       

# #  ************************** 紅包活動設定搜尋 **************************

#     def test_red_envelope_event_setting_identify_Tai(self):
#         '''紅包活動設定功能驗證(泰)'''

#         self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[2]/div[2]/div[1]/div/div/span/button').click()  # Event Start Time
#         time.sleep(1)

#         for _ in range(7):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 
        
#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[6]/a').click()  # 2024/11/01
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="search_start"]')
#         print('เวลาเริ่มต้นกิจกรรม:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         starttime_close = self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]')  # Done
#         self.driver.execute_script("$(arguments[0]).click()", starttime_close)
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/span/button').click()  # Event End Time
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[3]/a').click()  # 2024/12/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="search_end"]')
#         print('เวลาสิ้นสุดกิจกรรม:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         endtime_close = self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[3]/button[2]').click() # Done
#         self.driver.execute_script("$(arguments[0]).click()", endtime_close)
#         time.sleep(1) 

#         self.driver.find_element(By.ID, f'search_currency').click()  # Currency
#         time.sleep(1) 

#         currency_slt = self.driver.find_element(By.XPATH, f'//*[@id="search_currency"]/option[@value="CNY"]')  # CNY
#         currency_slt.click()
#         print('สกุลเงิน:', currency_slt.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'select_host').click()  # Agent
#         time.sleep(1) 

#         agent_slt = self.driver.find_element(By.XPATH, f'//*[@id="2571410-selectable"]')  # Test                                                            
#         agent_slt.click()
#         print('(勾選)ตัวแทน:', agent_slt.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'select_game').click()  # Participate Games
#         time.sleep(1)

#         game_select = self.driver.find_element(By.XPATH, f'//*[@id="1839777187-selectable"]')  # PSS-ON-00158 CYBER CUBE                                                      
#         game_select.click()
#         print('(勾選)เข้าร่วมเกม:', game_select.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'sh_btn').click()    
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="search_event_result"]/tbody/tr/td[5]/button').click()
#         time.sleep(1)

#         game_name = self.driver.find_element(By.XPATH, f'//*[@id="check_cost_game"]/div/div/div[2]')
#         game_name.click()
#         print(game_name.text, '\n')
#         time.sleep(1)
                
#         self.driver.find_element(By.XPATH, f'//*[@id="check_cost_game"]/div/div/div[3]/button').click()
#         time.sleep(1)
                
#         page_end = self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[3]')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#         time.sleep(2)

#         page_head = self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[1]/h5')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 頁首
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1) 
        

# # ============================= 後台功能巡測【活動】============================= 
# # *************************** 紅包活動成本計算 ***************************

#     def test_Event_Calculate_Cost_of_Red_Envelope_Event_zhCN(self):
#         '''【活動】紅包活動成本計算功能頁切換'''     

#         self.driver.find_element(By.ID, 'Event2').click()
#         print("進入活動功能選單!", '\n')
#         time.sleep(1)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Scratch/Red_Envelope_Event_Cost')
#         print("切換紅包活動成本計算選單!", '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[3]').click()
#         print("語系已切換'简体中文'!", '\n')
#         time.sleep(1)
        
        
# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_ccree_zhCN(self):
#         '''搜尋列功能驗證'''


# #  ************************** 幣別 **************************
        
#         self.driver.find_element(By.ID, f'search_currency').click()  # 幣別
#         time.sleep(1)
            
#         currency_list = ['MYR', 'PHP', 'USD']
#         for currency_slt in currency_list:
#             currencySelect = self.driver.find_element(By.XPATH, f'//*[@id="search_currency"]/option[@value="{currency_slt}"]')  # MYR / PHP / USD
#             currencySelect.click()
#             print('幣別:', currencySelect.get_attribute('value'), '\n')
#             time.sleep(1)
            
#         self.driver.refresh()
#         time.sleep(1)


# # ************************** 活動預計開始時間 **************************

#         self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[2]/div[2]/div[1]/div/div/span/button').click()  # 活動預計開始時間
#         time.sleep(1)

#         for _ in range(7):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1) 

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[5]/a').click()  # 2024/11/01
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="search_start"]')
#         print('活動預計開始時間:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

  
# # ************************** 活動預計結束時間 **************************

#         self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/span/button').click()  # 活動預計結束時間
#         time.sleep(1)

#         for _ in range(2):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[2]/span').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[5]/a').click()  # 2025/01/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="search_end"]')
#         print('活動預計結束時間:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 代理商 **************************

#         self.driver.find_element(By.ID, f'search_currency').click()  # 幣別
#         time.sleep(1) 

#         currencySelect = self.driver.find_element(By.XPATH, f'//*[@id="search_currency"]/option[@value="CNY"]')  # CNY
#         currencySelect.click()
#         print('幣別:', currencySelect.text, '\n')
#         time.sleep(1)
        
#         self.driver.find_element(By.ID, f'select_host').click()  # 代理商
#         time.sleep(1)        

#         agent_list = ['65921-selectable', '1874684019-selectable', '79713760-selectable']
#         for agent_list_id in agent_list:
#             agent_slt = self.driver.find_element(By.XPATH, f'//*[@id="{agent_list_id}"]')  # -- 全選 -- / Platform / Test-2                                                            
#             agent_slt.click()
#             print('(勾選)代理商:', agent_slt.text, '\n')
#             time.sleep(1)
            
#             agent_slt_cls = self.driver.find_element(By.XPATH, f'//*[@id="{agent_list_id}"]')  # (取消) -- 全選 -- / Platform / Test-2                                                              
#             agent_slt_cls.click()
#             print('(取消勾選)代理商:', agent_slt_cls.text, '\n')
#             time.sleep(1)
                      
#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 參加活動遊戲 **************************

#         self.driver.find_element(By.ID, f'select_game').click()  # 參加活動遊戲
#         time.sleep(1)
        
#         game_list = ['65921-selectable', '1839777181-selectable', '1844393676-selectable']
#         for game_list_id in game_list:
#             game_slt = self.driver.find_element(By.XPATH, f'//*[@id="{game_list_id}"]')  # -- 全選 -- / PSS-ON-00152 財神爸爸 / PSS-ON-50008 火焰金字塔 BLACK                                                        
#             game_slt.click()
#             print('(勾選)參加活動遊戲:', game_slt.text, '\n')
#             time.sleep(1)
            
#             game_slt_cls = self.driver.find_element(By.XPATH, f'//*[@id="{game_list_id}"]')  # -- 全選 -- / PSS-ON-00152 財神爸爸 / PSS-ON-50008 火焰金字塔 BLACK                                                          
#             game_slt_cls.click()
#             print('(取消勾選)參加活動遊戲:', game_slt_cls.text, '\n')
#             time.sleep(1)
                                                                                                                                                                                
#         self.driver.refresh()
#         time.sleep(1)                                                            
                                                                    

# #  ************************** 紅包活動成本計算搜尋 **************************

#     def test_calculate_cost_red_envelope_event_zhCN(self):
#         '''紅包活動成本計算功能驗證'''

#         self.driver.find_element(By.ID, f'search_currency').click()  # 幣別
#         time.sleep(1)

#         currency_slt = self.driver.find_element(By.XPATH, f'//*[@id="search_currency"]/option[@value="CNY"]')  # CNY
#         currency_slt.click()
#         print('幣別:', currency_slt.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[2]/div[2]/div[1]/div/div/span/button').click()  # 活動預計開始時間
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)
      
#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[7]/a').click()  # 2024/12/01
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="search_start"]')
#         print('活動預計開始時間:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/span/button').click()  # 活動預計結束時間
#         time.sleep(1)

#         for _ in range(1):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[2]').click()
#             time.sleep(1)
            
#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[5]/a').click()  # 2025/01/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="search_end"]')
#         print('活動預計結束時間:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'select_host').click()  # 代理商
#         time.sleep(1)

#         agent_slt = self.driver.find_element(By.XPATH, f'//*[@id="-1632647290-selectable"]')  # PLAYSTAR                                                            
#         agent_slt.click()
#         print('(勾選)代理商:', agent_slt.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'select_game').click()  # 參加活動遊戲
#         time.sleep(1)

#         game_slt = self.driver.find_element(By.XPATH, f'//*[@id="65921-selectable"]')  # -- 全選 --                                                      
#         game_slt.click()
#         print('(勾選)參加活動遊戲:', game_slt.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'sh_btn').click()
#         time.sleep(1)
                
#         page_end = self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[3]')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#         time.sleep(1)

#         page_head = self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[1]/h5')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 頁首
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1) 


# # --------------------------- 紅包活動成本計算(EN) ---------------------------

#     def test_Event_Calculate_Cost_of_Red_Envelope_Event_EN(self): 
#         '''【活動】紅包活動成本計算功能頁切換(英)'''

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[1]').click()
#         print("語系已切換:", 'English', '\n')
#         time.sleep(1)


# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_ccree_EN(self):
#         '''搜尋列功能驗證(英)'''
        

# #  ************************** Currency **************************

#         self.driver.find_element(By.ID, f'search_currency').click()  # 幣別
#         time.sleep(1)
            
#         currency_list = ['MYR', 'PHP', 'USD']
#         for currency_slt in currency_list:
#             currencySelect = self.driver.find_element(By.XPATH, f'//*[@id="search_currency"]/option[@value="{currency_slt}"]')  # MYR / PHP / USD
#             currencySelect.click()
#             print('Currency:', currencySelect.get_attribute('value'), '\n')
#             time.sleep(1)
            
#         self.driver.refresh()
#         time.sleep(1)


# # ************************** 活動預計開始時間 **************************

#         self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[2]/div[2]/div[1]/div/div/span/button').click()  # 活動預計開始時間
#         time.sleep(1)

#         for _ in range(7):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[6]/a').click()  # 2024/11/01
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="search_start"]')
#         print('Expected Start Time:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

 
# # ************************** 活動預計結束時間 **************************

#         self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/span/button').click()  # 活動預計結束時間
#         time.sleep(1)

#         for _ in range(2):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[2]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[6]/a').click()  # 2025/01/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="search_end"]')
#         print('Expected End Time:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** Agent **************************

#         self.driver.find_element(By.ID, f'search_currency').click()  # 幣別
#         time.sleep(1) 

#         currencySelect = self.driver.find_element(By.XPATH, f'//*[@id="search_currency"]/option[@value="CNY"]')  # CNY
#         currencySelect.click()
#         print('Currency:', currencySelect.text, '\n')
#         time.sleep(1)
        
#         self.driver.find_element(By.ID, f'select_host').click()  # 代理商
#         time.sleep(1)        

#         agent_list = ['65921-selectable', '1874684019-selectable', '79713760-selectable']
#         for agent_list_id in agent_list:
#             agent_slt = self.driver.find_element(By.XPATH, f'//*[@id="{agent_list_id}"]')  # -- Select All -- / Platform / Test-2                                                            
#             agent_slt.click()
#             print('(勾選)Agent:', agent_slt.text, '\n')
#             time.sleep(1)
            
#             agent_slt_cls = self.driver.find_element(By.XPATH, f'//*[@id="{agent_list_id}"]')  # (取消) -- Select All -- / Platform / Test-2                                                              
#             agent_slt_cls.click()
#             print('(取消勾選)Agent:', agent_slt_cls.text, '\n')
#             time.sleep(1)
                      
#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 參加活動遊戲 **************************

#         self.driver.find_element(By.ID, f'select_game').click()  # 參加活動遊戲
#         time.sleep(1)
        
#         game_list = ['65921-selectable', '1839777181-selectable', '1844393676-selectable']
#         for game_list_id in game_list:
#             game_slt = self.driver.find_element(By.XPATH, f'//*[@id="{game_list_id}"]')  # -- Select All -- / PSS-ON-00152 CAISHEN DADDY / PSS-ON-50008 PYRAMID OF FLAMES BLACK                                                           
#             game_slt.click()
#             print('(勾選)Participate Games:', game_slt.text, '\n')
#             time.sleep(1)
            
#             game_slt_cls = self.driver.find_element(By.XPATH, f'//*[@id="{game_list_id}"]')  # -- Select All -- / PSS-ON-00152 CAISHEN DADDY / PSS-ON-50008 PYRAMID OF FLAMES BLACK                                                           
#             game_slt_cls.click()
#             print('(取消勾選)Participate Games:', game_slt_cls.text, '\n')
#             time.sleep(1)
                                                                                                                                                                                
#         self.driver.refresh()
#         time.sleep(1)       
                                 

# #  ************************** 紅包活動成本計算搜尋 **************************

#     def test_calculate_cost_red_envelope_event_EN(self):
#         '''紅包活動成本計算功能驗證(英)'''

#         self.driver.find_element(By.ID, f'search_currency').click()  # 幣別
#         time.sleep(1)

#         currency_slt = self.driver.find_element(By.XPATH, f'//*[@id="search_currency"]/option[@value="CNY"]')  # CNY
#         currency_slt.click()
#         print('Currency:', currency_slt.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[2]/div[2]/div[1]/div/div/span/button').click()  # 活動預計開始時間
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[1]/a').click()  # 2024/12/01
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="search_start"]')
#         print('Expected Start Time:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/span/button').click()  # 活動預計結束時間
#         time.sleep(1)

#         for _ in range(1):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[2]').click()
#             time.sleep(1)
            
#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[6]/a').click()  # 2025/01/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="search_end"]')
#         print('Expected End Time:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'select_host').click()  # 代理商
#         time.sleep(1)

#         agent_slt = self.driver.find_element(By.XPATH, f'//*[@id="-1632647290-selectable"]')  # PLAYSTAR                                                            
#         agent_slt.click()
#         print('(勾選)Agent:', agent_slt.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'select_game').click()  # 參加活動遊戲
#         time.sleep(1)

#         game_slt = self.driver.find_element(By.XPATH, f'//*[@id="65921-selectable"]')  # -- Select All --                                                     
#         game_slt.click()
#         print('(勾選)Participate Games:', game_slt.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'sh_btn').click()
#         time.sleep(1)
                
#         page_end = self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[3]')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#         time.sleep(1)

#         page_head = self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[1]/h5')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 頁首
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1) 


# # --------------------------- 紅包活動成本計算(Tai) ---------------------------

#     def test_Event_Calculate_Cost_of_Red_Envelope_Event_Tai(self):
#         '''【活動】紅包活動成本計算功能頁切換(泰)''' 
        
#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[4]').click()
#         print("語系已切換:", 'ไทย', '\n')
#         time.sleep(1)
        

# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_ccree_Tai(self):
#         '''搜尋列功能驗證(泰)'''


# #  ************************** 幣別 **************************

#         self.driver.find_element(By.ID, f'search_currency').click()  # 幣別
#         time.sleep(1)
            
#         currency_list = ['MYR', 'PHP', 'USD']
#         for currency_slt in currency_list:
#             currencySelect = self.driver.find_element(By.XPATH, f'//*[@id="search_currency"]/option[@value="{currency_slt}"]')  # MYR / PHP / USD
#             currencySelect.click()
#             print('สกุลเงิน:', currencySelect.get_attribute('value'), '\n')
#             time.sleep(1)
            
#         self.driver.refresh()
#         time.sleep(1)


# # ************************** 活動預計開始時間 **************************

#         self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[2]/div[2]/div[1]/div/div/span/button').click()  # 活動預計開始時間
#         time.sleep(1)

#         for _ in range(7):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click() 
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[6]/a').click()  # 2024/11/01
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="search_start"]')
#         print('เวลาคาดเริ่มต้น:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

 
# # ************************** 活動預計結束時間 **************************

#         self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/span/button').click()  # 活動預計結束時間
#         time.sleep(1)

#         for _ in range(2):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[2]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[6]/a').click()  # 2025/01/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="search_end"]')
#         print('เวลาคาดสิ้นสุด:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 代理商 **************************

#         self.driver.find_element(By.ID, f'search_currency').click()  # 幣別
#         time.sleep(1) 

#         currencySelect = self.driver.find_element(By.XPATH, f'//*[@id="search_currency"]/option[@value="CNY"]')  # CNY
#         currencySelect.click()
#         print('สกุลเงิน:', currencySelect.text, '\n')
#         time.sleep(1)
        
#         self.driver.find_element(By.ID, f'select_host').click()  # 代理商
#         time.sleep(1)        

#         agent_list = ['65921-selectable', '1874684019-selectable', '79713760-selectable']
#         for agent_list_id in agent_list:
#             agent_slt = self.driver.find_element(By.XPATH, f'//*[@id="{agent_list_id}"]')  # -- เลือกทั้งหมด -- / Platform / Test-2                                                            
#             agent_slt.click()
#             print('(勾選)ตัวแทน:', agent_slt.text, '\n')
#             time.sleep(1)
            
#             agent_slt_cls = self.driver.find_element(By.XPATH, f'//*[@id="{agent_list_id}"]')  # (取消) -- เลือกทั้งหมด -- / Platform / Test-2                                                              
#             agent_slt_cls.click()
#             print('(取消勾選)ตัวแทน:', agent_slt_cls.text, '\n')
#             time.sleep(1)
                      
#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 參加活動遊戲 **************************

#         self.driver.find_element(By.ID, f'select_game').click()  # 參加活動遊戲
#         time.sleep(1)
        
#         game_list = ['65921-selectable', '1839777181-selectable', '1844393676-selectable']
#         for game_list_id in game_list:
#             game_slt = self.driver.find_element(By.XPATH, f'//*[@id="{game_list_id}"]')  # -- เลือกทั้งหมด -- / PSS-ON-00152 แด๊ดดี้ เทพเจ้าแห่งโชคลาภ / PSS-ON-50008 ปิรามิดในกองไฟ BLACK                                                         
#             game_slt.click()
#             print('(勾選)เข้าร่วมเกม:', game_slt.text, '\n')
#             time.sleep(1)
            
#             game_slt_cls = self.driver.find_element(By.XPATH, f'//*[@id="{game_list_id}"]')  # -- เลือกทั้งหมด -- / PSS-ON-00152 แด๊ดดี้ เทพเจ้าแห่งโชคลาภ / PSS-ON-50008 ปิรามิดในกองไฟ BLACK                                                            
#             game_slt_cls.click()
#             print('(取消勾選)เข้าร่วมเกม:', game_slt_cls.text, '\n')
#             time.sleep(1)
                                                                                                                                                                                
#         self.driver.refresh()
#         time.sleep(1)                                                          
                                                            

# #  ************************** 紅包活動成本計算搜尋 **************************

#     def test_calculate_cost_red_envelope_event_Tai(self):
#         '''紅包活動成本計算功能驗證(泰)'''

#         self.driver.find_element(By.ID, f'search_currency').click()  # 幣別
#         time.sleep(1)

#         currency_slt = self.driver.find_element(By.XPATH, f'//*[@id="search_currency"]/option[@value="CNY"]')  # CNY
#         currency_slt.click()
#         print('สกุลเงิน:', currency_slt.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[2]/div[2]/div[1]/div/div/span/button').click()  # 活動預計開始時間
#         time.sleep(1)

#         for _ in range(6):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div[1]/a[1]').click()
#             time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[1]/td[1]/a').click()  # 2024/12/01
#         time.sleep(1)

#         start_time = self.driver.find_element(By.XPATH, f'//*[@id="search_start"]')
#         print('เวลาคาดเริ่มต้น:', start_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/span/button').click()  # 活動預計結束時間
#         time.sleep(1)

#         for _ in range(1):
#             self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/div/a[2]').click()
#             time.sleep(1)
            
#         self.driver.find_element(By.XPATH, f'//*[@id="ui-datepicker-div"]/table/tbody/tr[5]/td[6]/a').click()  # 2025/01/31
#         time.sleep(1)

#         end_time = self.driver.find_element(By.XPATH, f'//*[@id="search_end"]')
#         print('เวลาคาดสิ้นสุด:', end_time.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'select_host').click()  # 代理商
#         time.sleep(1)

#         agent_slt = self.driver.find_element(By.XPATH, f'//*[@id="-1632647290-selectable"]')  # PLAYSTAR                                                            
#         agent_slt.click()
#         print('(勾選)ตัวแทน:', agent_slt.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'select_game').click()  # 參加活動遊戲
#         time.sleep(1)

#         game_slt = self.driver.find_element(By.XPATH, f'//*[@id="65921-selectable"]')  # -- 全選 --                                                      
#         game_slt.click()
#         print('(勾選)เข้าร่วมเกม:', game_slt.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, 'sh_btn').click()
#         time.sleep(1)
                
#         page_end = self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[3]')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)  # 頁尾
#         time.sleep(1)

#         page_head = self.driver.find_element(By.XPATH, '//*[@id="search_area"]/div[1]/div/div/div[1]/h5')  
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 頁首
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1) 


# # ============================= 後台功能巡測【配置】============================= 
# # *************************** 代理商管理 ***************************

#     def test_Configuration_Agent_Management_zhCN(self):
#         '''【配置】代理商管理功能頁切換'''     

#         self.driver.find_element(By.ID, 'Configuration').click()
#         print("進入配置功能選單!", '\n')
#         time.sleep(1)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Configuration/agent_managemen')
#         print("切換代理商管理選單!", '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[3]').click()
#         print("語系已切換'简体中文'!", '\n')
#         time.sleep(1)
        

# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_am_zhCN(self):
#         '''搜尋列功能驗證'''


# #  ************************** 代理商 **************************

#         self.driver.find_element(By.ID, f'agent_type').click()  # 代理商    
#         time.sleep(1) 
        
#         agent_list = ['default', 'PS', 'Test']
#         for agent_id in agent_list:
#             agent_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_type"]/option[@value="{agent_id}"]')  # 預設 / PS / test                                                            
#             agent_select.click()
#             print('代理商:', agent_select.get_attribute('value'), '\n')
#             time.sleep(1)
            
#             try:
#                 alertt = self.driver.switch_to.alert
#                 print(alertt.text, '\n')
#                 time.sleep(3)
#                 alertt.accept()
                
#             except NoAlertPresentException:
#                 pass

#         self.driver.refresh()
#         time.sleep(1)
        

# # ********************** 狀態 **************************

#         self.driver.find_element(By.ID, f'status_type').click()  # 狀態
#         time.sleep(1)
        
#         status_value = ['All', "0", "1", "2", "3"]
#         for cond_list in status_value:
#             cond_select_1 = self.driver.find_element(By.XPATH, f'//*[@id="status_type"]/option[@value="{cond_list}"]')  #  -- 全選 -- / 停用 / 啟用 (已營運) / 啟用 (未營運) / 維護                                                            
#             cond_select_1.click()
#             print('狀態:', cond_select_1.text, '\n')
#             time.sleep(1)
            
#             try:
#                 alertt = self.driver.switch_to.alert
#                 print(alertt.text, '\n')
#                 time.sleep(3)
#                 alertt.accept()
                
#             except NoAlertPresentException:
#                 pass

#         self.driver.refresh()
#         time.sleep(1)


# # ********************** 批次執行 **************************

#     def test_agent_management_batch_execution_zhCN(self):
#         '''代理商管理批次執行功能驗證'''
        
#         batch_exe = self.driver.find_element(By.XPATH, f'/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[5]/div/button')  # 批次執行
#         self.driver.execute_script("$(arguments[0]).click()", batch_exe)  
#         time.sleep(1)
        
#         for game_gs in range(1, 3):    
#             gs_icon_ch = self.driver.find_element(By.XPATH, f'//*[@id="myModal_host_batch"]/div/div/div[2]/div[1]/div/div/label[{game_gs}]')  # 老虎機 GS 切換 / 捕魚機 GS 切换
#             self.driver.execute_script("$(arguments[0]).click()", gs_icon_ch)
#             print('已切換:', gs_icon_ch.text, '\n')  
#             time.sleep(1)
            
#             self.driver.find_element(By.ID, f'batch_confirm_submit').click()  # 送出
#             time.sleep(1)

#             notice_info = self.driver.find_element(By.XPATH, f'/html/body/div[2]') 
#             print(notice_info.text, '\n')
#             time.sleep(2)
            
#         slot_gs = self.driver.find_element(By.XPATH, f'//*[@id="myModal_host_batch"]/div/div/div[2]/div[1]/div/div/label[1]/input')
#         slot_gs.click()
#         print(slot_gs.get_attribute('innerText'), '\n')

#         self.driver.find_element(By.XPATH, f'//*[@id="option"]/div[2]/div[1]/label/input').click()  # 123
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'batch_confirm_submit').click()  # 送出
#         time.sleep(1)

#         notice_info = self.driver.find_element(By.XPATH, f'/html/body/div[2]')  
#         print(notice_info.text, '\n')
#         time.sleep(2)

#         self.driver.refresh()
#         time.sleep(1)


# # ********************** 匯出 **************************

#     def test_agent_management_export_zhCN(self):
#         '''匯出功能驗證'''

#         self.driver.find_element(By.ID, f'export').click()  # 匯出
#         time.sleep(15)
            
#         self.driver.refresh()
#         time.sleep(1)
         

# # ********************** 代理商搜尋與重啟 **************************

#     def test_agent_search_with_reset_zhCN(self):
#         '''代理商搜尋與重啟功能驗證'''

#         search_box = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0_filter"]/label/input')  # 篩選
#         search_box.click()
#         search_box.send_keys('_ivantest')

#         self.driver.find_element(By.NAME, f'disabled').click()  # 切換
#         time.sleep(1)

#         for cond_change in range(1, 5):
#             self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[5]/select/option[{cond_change}]').click()  # 停用 / 啟用 (已營運) / 啟用 (未營運) / 維護
#             time.sleep(2)

#             notice_info = self.driver.find_element(By.XPATH, f'/html/body/div[2]')
#             print((notice_info.text), '\n')
#             time.sleep(2)
        
#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[5]/select').click()  # 切換
#         time.sleep(1)
        
#         self.driver.find_element(By.XPATH, '//*[@id="DataTables_Table_0"]/tbody/tr/td[5]/select/option[1]').click()  # 停用 (代理商ID = _ivantest)      
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[6]/button[5]').click()  # 重啟
#         time.sleep(1)
        
#         note_info = self.driver.find_element(By.XPATH, f'//*[@id="myModal_3"]/div')  # 重啟確認
#         print(note_info.text, '\n')

#         self.driver.find_element(By.ID, 'reset_confirm_close').click()  # 關閉視窗
#         time.sleep(2)

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[6]/button[5]').click()  # 重啟
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'reset_confirm_submit').click()  # 送出
#         time.sleep(2)

#         note_info_1 = self.driver.find_element(By.XPATH, f'/html/body/div[2]')  # 重啟成功
#         print(note_info_1.text, '\n')
#         time.sleep(1)
        
#         self.driver.refresh() 
#         time.sleep(1)


# # --------------------------- 代理商管理(EN) ---------------------------

#     def test_Configuration_Agent_Management_EN(self):
#         '''【配置】代理商管理功能頁切換(英)'''  
        
#         self.driver.find_element(By.ID, 'Configuration').click()
#         print("進入配置功能選單!", '\n')
#         time.sleep(1)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Configuration/agent_managemen')
#         print("切換代理商管理選單!", '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[1]').click()
#         print("語系已切換:", 'English', '\n')
#         time.sleep(1)


# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_am_EN(self):
#         '''搜尋列功能驗證(英)'''


# #  ************************** Agent **************************

#         self.driver.find_element(By.ID, f'agent_type').click()  # 代理商    
#         time.sleep(1) 
        
#         agent_list = ['default', 'PS', 'Test']
#         for agent_id in agent_list:
#             agent_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_type"]/option[@value="{agent_id}"]')  # Default / PS / test                                                            
#             agent_select.click()
#             print('Agent:', agent_select.get_attribute('value'), '\n')
#             time.sleep(1)
            
#             try:
#                 alertt = self.driver.switch_to.alert
#                 print(alertt.text, '\n')
#                 time.sleep(3)
#                 alertt.accept()
                
#             except NoAlertPresentException:
#                 pass

#         self.driver.refresh()
#         time.sleep(1)


# # ********************** Status **************************

#         self.driver.find_element(By.ID, f'status_type').click()  # 狀態
#         time.sleep(1)
        
#         status_value = ['All', "0", "1", "2", "3"]
#         for cond_list in status_value:
#             cond_select_1 = self.driver.find_element(By.XPATH, f'//*[@id="status_type"]/option[@value="{cond_list}"]')  #   -- Select All -- / Disabled / Enabled (Operating) / Enabled (Non-Operating) / Maintenance                                                            
#             cond_select_1.click()
#             print('Status:', cond_select_1.text, '\n')
#             time.sleep(1)
            
#             try:
#                 alertt = self.driver.switch_to.alert
#                 print(alertt.text, '\n')
#                 time.sleep(3)
#                 alertt.accept()
                
#             except NoAlertPresentException:
#                 pass

#         self.driver.refresh()
#         time.sleep(1)


# # ********************** Batch Execution  **************************

#     def test_agent_management_batch_execution_EN(self):
#         '''代理商管理批次執行功能驗證(英)'''

#         batch_exe = self.driver.find_element(By.XPATH, f'/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[5]/div/button')  # 批次執行
#         self.driver.execute_script("$(arguments[0]).click()", batch_exe)  
#         time.sleep(1)
        
#         for game_gs in range(1, 3):    
#             gs_icon_ch = self.driver.find_element(By.XPATH, f'//*[@id="myModal_host_batch"]/div/div/div[2]/div[1]/div/div/label[{game_gs}]')  # 老虎機 GS 切換 / 捕魚機 GS 切换
#             self.driver.execute_script("$(arguments[0]).click()", gs_icon_ch)
#             print('已切換:', gs_icon_ch.text, '\n')  
#             time.sleep(1)
            
#             self.driver.find_element(By.ID, f'batch_confirm_submit').click()  # 送出
#             time.sleep(1)

#             notice_info = self.driver.find_element(By.XPATH, f'/html/body/div[2]') 
#             print(notice_info.text, '\n')
#             time.sleep(2)
            
#         slot_gs = self.driver.find_element(By.XPATH, f'//*[@id="myModal_host_batch"]/div/div/div[2]/div[1]/div/div/label[1]/input')
#         slot_gs.click()
#         print(slot_gs.get_attribute('innerText'), '\n')

#         self.driver.find_element(By.XPATH, f'//*[@id="option"]/div[2]/div[1]/label/input').click()  # 123
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'batch_confirm_submit').click()  # 送出
#         time.sleep(1)

#         notice_info = self.driver.find_element(By.XPATH, f'/html/body/div[2]')  
#         print(notice_info.text, '\n')
#         time.sleep(2)

#         self.driver.refresh()
#         time.sleep(1)


# # ********************** Export **************************

#     def test_agent_management_export_EN(self):
#         '''匯出功能驗證(英)'''

#         self.driver.find_element(By.ID, f'export').click()  # 匯出
#         time.sleep(15)
            
#         self.driver.refresh()
#         time.sleep(1)


# # ********************** Agent搜尋與重啟 **************************

#     def test_agent_search_with_reset_EN(self):
#         '''代理商搜尋與重啟功能驗證(英)'''

#         search_box = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0_filter"]/label/input')  # 篩選
#         search_box.click()
#         search_box.send_keys('_ivantest')

#         self.driver.find_element(By.NAME, f'disabled').click()  # 切換
#         time.sleep(1)

#         for cond_change in reversed(range(1, 5)):
#             self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[5]/select/option[{cond_change}]').click()  # 停用 / 啟用 (已營運) / 啟用 (未營運) / 維護
#             time.sleep(2)

#             notice_info = self.driver.find_element(By.XPATH, f'/html/body/div[2]')
#             print((notice_info.text), '\n')
#             time.sleep(2)
                
#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[5]/select').click()  # 切換
#         time.sleep(1)
        
#         self.driver.find_element(By.XPATH, '//*[@id="DataTables_Table_0"]/tbody/tr/td[5]/select/option[2]').click()  # 啟用 (已營運) (代理商ID = _ivantest)      
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[6]/button[5]').click()  # 重啟
#         time.sleep(1)
        
#         note_info = self.driver.find_element(By.XPATH, f'//*[@id="myModal_3"]/div')  # 重啟確認
#         print(note_info.text, '\n')

#         self.driver.find_element(By.ID, f'reset_confirm_close').click()  # 關閉視窗
#         time.sleep(2)

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[6]/button[5]').click()  # 重啟
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'reset_confirm_submit').click()  # 送出
#         time.sleep(2)

#         note_info_1 = self.driver.find_element(By.XPATH, f'/html/body/div[2]')  # 重啟成功
#         print(note_info_1.text, '\n')
#         time.sleep(1)
        
#         self.driver.refresh() 
#         time.sleep(1)


# # --------------------------- 代理商管理(Tai) ---------------------------

#     def test_Configuration_Agent_Management_Tai(self):
#         '''【配置】代理商管理功能頁切換(泰)'''  
        
#         self.driver.find_element(By.ID, 'Configuration').click()
#         print("進入配置功能選單!", '\n')
#         time.sleep(1)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Configuration/agent_managemen')
#         print("切換代理商管理選單!", '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[4]').click()
#         print("語系已切換:", 'ไทย', '\n')
#         time.sleep(1)


# # ************************** 搜尋功能 **************************

#     def test_search_func_identify_am_Tai(self):
#         '''搜尋列功能驗證(泰)'''


# #  ************************** 代理商 **************************

#         self.driver.find_element(By.ID, f'agent_type').click()  # 代理商    
#         time.sleep(1) 
        
#         agent_list = ['default', 'PS', 'Test']
#         for agent_id in agent_list:
#             agent_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_type"]/option[@value="{agent_id}"]')  # Default / PS / test                                                            
#             agent_select.click()
#             print('ตัวแทน:', agent_select.get_attribute('value'), '\n')
#             time.sleep(1)
            
#             try:
#                 alertt = self.driver.switch_to.alert
#                 print(alertt.text, '\n')
#                 time.sleep(3)
#                 alertt.accept()
                
#             except NoAlertPresentException:
#                 pass

#         self.driver.refresh()
#         time.sleep(1)


# # ********************** 狀態 **************************

#         self.driver.find_element(By.ID, f'status_type').click()  # 狀態
#         time.sleep(1)
        
#         status_value = ['All', "0", "1", "2", "3"]
#         for cond_list in status_value:
#             cond_select_1 = self.driver.find_element(By.XPATH, f'//*[@id="status_type"]/option[@value="{cond_list}"]')  #   -- Select All -- / Disabled / Enabled (Operating) / Enabled (Non-Operating) / Maintenance                                                            
#             cond_select_1.click()
#             print('สถานะ:', cond_select_1.text, '\n')
#             time.sleep(1)
            
#             try:
#                 alertt = self.driver.switch_to.alert
#                 print(alertt.text, '\n')
#                 time.sleep(3)
#                 alertt.accept()
                
#             except NoAlertPresentException:
#                 pass

#         self.driver.refresh()
#         time.sleep(1)


# # ********************** การดำเนินการแบทช์  **************************

#     def test_agent_management_batch_execution_Tai(self):
#         '''代理商管理批次執行功能驗證(泰)'''

#         batch_exe = self.driver.find_element(By.XPATH, f'/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[5]/div/button')  # 批次執行
#         self.driver.execute_script("$(arguments[0]).click()", batch_exe)  
#         time.sleep(1)
        
#         for game_gs in range(1, 3):    
#             gs_icon_ch = self.driver.find_element(By.XPATH, f'//*[@id="myModal_host_batch"]/div/div/div[2]/div[1]/div/div/label[{game_gs}]')  # 老虎機 GS 切換 / 捕魚機 GS 切换
#             self.driver.execute_script("$(arguments[0]).click()", gs_icon_ch)
#             print('已切換:', gs_icon_ch.text, '\n')  
#             time.sleep(1)
            
#             self.driver.find_element(By.ID, f'batch_confirm_submit').click()  # 送出
#             time.sleep(1)

#             notice_info = self.driver.find_element(By.XPATH, f'/html/body/div[2]') 
#             print(notice_info.text, '\n')
#             time.sleep(2)
            
#         slot_gs = self.driver.find_element(By.XPATH, f'//*[@id="myModal_host_batch"]/div/div/div[2]/div[1]/div/div/label[1]/input')
#         slot_gs.click()
#         print(slot_gs.get_attribute('innerText'), '\n')

#         self.driver.find_element(By.XPATH, f'//*[@id="option"]/div[2]/div[1]/label/input').click()  # 123
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'batch_confirm_submit').click()  # 送出
#         time.sleep(1)

#         notice_info = self.driver.find_element(By.XPATH, f'/html/body/div[2]')  
#         print(notice_info.text, '\n')
#         time.sleep(2)

#         self.driver.refresh()
#         time.sleep(1)
        

# # ********************** 匯出 **************************

#     def test_agent_management_export_Tai(self):
#         '''匯出功能驗證(泰)'''

#         self.driver.find_element(By.ID, f'export').click()  # 匯出
#         time.sleep(15)
            
#         self.driver.refresh()
#         time.sleep(1)


# # ********************** 代理商搜尋與重啟 **************************

#     def test_agent_search_with_reset_Tai(self):
#         '''代理商搜尋與重啟功能驗證(泰)'''

#         search_box = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0_filter"]/label/input')  # 篩選
#         search_box.click()
#         search_box.send_keys('_ivantest')

#         self.driver.find_element(By.NAME, f'disabled').click()  # 切換
#         time.sleep(1)

#         for cond_change in reversed(range(1, 5)):
#             self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[5]/select/option[{cond_change}]').click()  # 停用 / 啟用 (已營運) / 啟用 (未營運) / 維護
#             time.sleep(2)

#             notice_info = self.driver.find_element(By.XPATH, f'/html/body/div[2]')
#             print((notice_info.text), '\n')
#             time.sleep(2)
                
#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[5]/select').click()  # 切換
#         time.sleep(1)
        
#         self.driver.find_element(By.XPATH, '//*[@id="DataTables_Table_0"]/tbody/tr/td[5]/select/option[1]').click()  # 停用 (代理商ID = _ivantest)      
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[6]/button[5]').click()  # 重啟
#         time.sleep(1)
        
#         note_info = self.driver.find_element(By.XPATH, f'//*[@id="myModal_3"]/div')  # 重啟確認
#         print(note_info.text, '\n')

#         self.driver.find_element(By.ID, f'reset_confirm_close').click()  # 關閉視窗
#         time.sleep(2)

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[6]/button[5]').click()  # 重啟
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'reset_confirm_submit').click()  # 送出
#         time.sleep(2)

#         note_info_1 = self.driver.find_element(By.XPATH, f'/html/body/div[2]')  # 重啟成功
#         print(note_info_1.text, '\n')
#         time.sleep(1)
        
#         self.driver.refresh() 
#         time.sleep(1)


# # ============================= 後台功能巡測【配置】============================= 
# # *************************** 遊戲配置 ***************************

#     def test_Configuration_Game_Configuration_zhCN(self):
#         '''【配置】遊戲配置功能頁切換'''     

        
#         self.driver.find_element(By.ID, 'Configuration').click()
#         print("進入配置功能選單!", '\n')
#         time.sleep(2)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Configuration/game_config')
#         print("切換遊戲配置選單!", '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[3]').click()
#         print("語系已切換'简体中文'!", '\n')
#         time.sleep(1)


# #  ************************** 代理商類別 **************************

#     def test_agent_class_identify_gc_zhCN(self):
#         '''代理商類別功能驗證'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_list = ['All', 'default', 'Test']
#         for agent_class_name in agent_class_list:
#             agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="{agent_class_name}"]')  # -- 全選 -- / 預設 / Test                                                            
#             agent_class_select.click()
#             print('代理商類別:', agent_class_select.get_attribute('value'), '\n')
#             time.sleep(1)
            
#             try:
#                 alertt = self.driver.switch_to.alert
#                 print(alertt.text, '\n')
#                 time.sleep(3)
#                 alertt.accept()
                
#             except NoAlertPresentException:
#                 pass
#             self.driver.refresh()
#             time.sleep(1)
            
#             page_end = self.driver.find_element(By.XPATH, f'/html/body/div[3]/div/div[4]/div[2]/div[2]')  # 頁尾
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)    
#             time.sleep(2)
            
#             page_head = self.driver.find_element(By.CLASS_NAME, f'text-semibold') 
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 遊戲配置
#             time.sleep(1)
            
#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 代理商 **************************

#     def test_agent_identify_gc_zhCN(self):
#         '''代理商功能驗證'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="All"]')  # -- 全選 --                                                             
#         agent_class_select.click()
#         print('代理商類別:', agent_class_select.get_attribute('value'), '\n')

#         self.driver.find_element(By.ID, f'slot_agent').click()  # (老虎機)代理商
#         time.sleep(1)

#         agent_list = ['MATH', 'TEST2']
#         for agent_num in agent_list:
#             agent_select = self.driver.find_element(By.XPATH, f'//*[@id="slot_agent"]/option[@value="{agent_num}"]')  # MATH / Test-2
#             agent_select.click() 
#             print('代理商:', agent_select.get_attribute('value'), '\n')
#             time.sleep(1)

#             page_end = self.driver.find_element(By.XPATH, f'/html/body/div[3]/div/div[4]/div[2]/div[2]')  # 頁尾
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)    
#             time.sleep(2)

#             page_head = self.driver.find_element(By.CLASS_NAME, f'text-semibold')
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 遊戲配置
#             time.sleep(1)
            
#         self.driver.refresh()
#         time.sleep(1)
        

# # ********************** (老虎機)遊戲狀態確認 **************************

#     def test_slot_game_cond_identify_gc_zhCN(self):
#         '''(老虎機)遊戲狀態確認'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  #  Test                                                            
#         agent_class_select.click()
#         print('代理商類別:', agent_class_select.get_attribute('value'), '\n')

#         self.driver.find_element(By.ID, f'slot_agent').click()  # (老虎機)代理商
#         time.sleep(1) 

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="slot_agent"]/option[@value="TEST5"]')  # Test-5
#         agent_select.click() 
#         print('代理商:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'status_type').click()  # 狀態
#         time.sleep(1)

#         for condition in reversed(range(1, 4)):
#             cond_select = self.driver.find_element(By.XPATH, f'//*[@id="status_type"]/option[{condition}]')  #  -- 全選 -- / 停用 / 啟用
#             cond_select.click() 
#             print('狀態:', cond_select.text, '\n')
#             time.sleep(1)
            
#             page_end = self.driver.find_element(By.XPATH, f'/html/body/div[3]/div/div[4]/div[2]/div[2]')  # 頁尾
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)    
#             time.sleep(2)
            
#             page_head = self.driver.find_element(By.CLASS_NAME, f'text-semibold') 
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 遊戲配置
#             time.sleep(1)
        
#         self.driver.refresh()   
#         time.sleep(1)
         

# # ********************** 複製遊戲 **************************

#     def test_slot_game_copy_gc_zhCN(self):
#         '''(老虎機)遊戲複製'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1) 

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="All"]')  # -- 全選 --                                                             
#         agent_class_select.click()
#         print('代理商類別:', agent_class_select.get_attribute('value'), '\n')

#         self.driver.find_element(By.ID, f'slot_agent').click()  # (老虎機)代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="slot_agent"]/option[@value="TEST15"]')  # Test-15
#         agent_select.click() 
#         print('代理商:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'batch_btn').click()  # 批次執行
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'reset_btn').click()  # 更新後重啟
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'currency_search').click()  # 帳戶幣別
#         time.sleep(1)

#         currency_select = self.driver.find_element(By.XPATH, f'//*[@id="currency_search"]/option[@value="CNY"]')  # CNY
#         currency_select.click()
#         print('帳戶幣別:', currency_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         agent_list = ['-1632647290-selectable', '2359048-selectable'] 
#         for agent_id in agent_list:
#             agent_select = self.driver.find_element(By.XPATH, f'//*[@id="{agent_id}"]')  # PLAYSTAR(CNY) / MATH(CNY)
#             self.driver.execute_script("arguments[0].scrollIntoView();", agent_select)
#             agent_select.click()
#             print('欲異動的代理商:', agent_select.get_attribute('textContent'), '\n')
#             time.sleep(1)     
        
#         game_select = self.driver.find_element(By.XPATH, f'//*[@id="-1500625730-selectable"]')  # PSS-ON-00160 熊熊战争
#         game_select.click()
#         print('欲複製的遊戲:', game_select.get_attribute('textContent'), '\n')
#         time.sleep(1)  
            
#         self.driver.find_element(By.ID, f'batch_confirm_submit').click()  # 送出
#         time.sleep(3)

#         note_info = self.driver.find_element(By.XPATH, f'/html/body/div[1]')  # 通知訊息
#         print(note_info.text, '\n')
#         time.sleep(2)

#         self.driver.refresh()
#         time.sleep(1)   


# # ********************** 變更遊戲狀態(停用) **************************

#     def test_slot_game_cond_dis_gc_zhCN(self):
#         '''(老虎機)變更遊戲狀態(停用)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test                                                           
#         agent_class_select.click()
#         print('代理商類別:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'slot_agent').click()  # (老虎機)代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="slot_agent"]/option[@value="TEST15"]')  # Test-15
#         agent_select.click() 
#         print('代理商:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[12]/select').click()  # 狀態
#         time.sleep(1)

#         for op_code_0 in range(1, 3):
#             op_cond_0 = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[12]/select/option[{op_code_0}]')  # 停用/啟用
#             op_num_0 = op_cond_0.get_attribute('value')
#             if op_num_0 == '1':
#                 print('(目前)狀態:', op_cond_0.text, '\n')    

#         op_status_stop = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[12]/select/option[1]')  # 停用
#         op_status_stop.click() 
#         print('(變更後)狀態:', op_status_stop.text, '\n')
#         time.sleep(1)

#         note_info = self.driver.find_element(By.XPATH, f'/html/body/div[1]')  # 通知訊息
#         print(note_info.text, '\n')
#         time.sleep(2)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Configuration/agent_managemen')  # 代理商管理
#         time.sleep(1)
            
#         filter_box = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0_filter"]/label/input')  # 篩選
#         filter_box.click()
#         filter_box.send_keys('test15')
#         time.sleep(1) 

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[6]/button[5]').click()  # 重啟
#         time.sleep(2)

#         self.driver.find_element(By.XPATH, f'//*[@id="reset_confirm_submit"]').click()  # 送出
#         time.sleep(2)

#         note_info_00 = self.driver.find_element(By.XPATH, f'/html/body/div[2]')  # 通知訊息 
#         print(note_info_00.text, '\n')
#         time.sleep(2)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Configuration/game_config')  # 遊戲配置
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test                                                           
#         agent_class_select.click()
#         print('代理商類別:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'slot_agent').click()  # (老虎機)代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="slot_agent"]/option[@value="TEST15"]')  # Test-15
#         agent_select.click() 
#         print('代理商:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         filter_box_1 = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0_filter"]/label/input')  # 篩選
#         filter_box_1.click()
#         filter_box_1.send_keys('PSS-ON-00156') 
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[12]/select').click()  # 狀態
#         time.sleep(1)

#         for op_code_1 in range(1, 3):
#             op_cond_1 = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[12]/select/option[{op_code_1}]')  # 停用/啟用
#             op_num_1 = op_cond_1.get_attribute('value')
#             if op_num_1 == '0':
#                 print('(目前)狀態:', op_cond_1.text, '\n')    
            
#         time.sleep(3)    
#         self.driver.refresh()
#         time.sleep(1)   


# # ********************** 變更遊戲狀態(啟用) **************************

#     def test_slot_game_cond_en_gc_zhCN(self):
#         '''(老虎機)變更遊戲狀態(啟用)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1) 

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test                                                             
#         agent_class_select.click()
#         print('代理商類別:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)
        
#         self.driver.find_element(By.ID, f'slot_agent').click()  # (老虎機)代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="slot_agent"]/option[@value="TEST15"]')  # Test-15
#         agent_select.click() 
#         print('代理商:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[12]/select').click()  # 狀態
#         time.sleep(1)

#         for op_code_0 in range(1, 3):
#             op_cond_0 = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[12]/select/option[{op_code_0}]')  # 停用/啟用
#             op_num_0 = op_cond_0.get_attribute('value')
#             if op_num_0 == '0':
#                 print('(目前)狀態:', op_cond_0.text, '\n')
#             time.sleep(1)

#         op_status_stop = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[12]/select/option[2]')  # 啟用
#         op_status_stop.click() 
#         print('(變更後)狀態:', op_status_stop.text, '\n')
#         time.sleep(1)

#         note_info = self.driver.find_element(By.XPATH, f'/html/body/div[1]')  # 通知訊息
#         print(note_info.text, '\n')
#         time.sleep(2)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Configuration/agent_managemen')  # 代理商管理
#         time.sleep(1)

#         filter_box = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0_filter"]/label/input')  # 篩選
#         filter_box.click()
#         filter_box.send_keys('test15')
#         time.sleep(1) 

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[6]/button[5]').click()  # 重啟
#         time.sleep(2)

#         self.driver.find_element(By.XPATH, f'//*[@id="reset_confirm_submit"]').click()  # 送出
#         time.sleep(3)

#         note_info_01 = self.driver.find_element(By.XPATH, f'/html/body/div[2]')  # 通知訊息
#         print(note_info_01.text, '\n')
#         time.sleep(2)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Configuration/game_config')  # 遊戲配置
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="All"]')  # -- 全選 --                                                             
#         agent_class_select.click()
#         print('代理商類別:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         filter_box_1 = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0_filter"]/label/input')  # 篩選
#         filter_box_1.click()
#         filter_box_1.send_keys('PSS-ON-00156') 
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[12]/select').click()  # 狀態
#         time.sleep(1)

#         for op_code in range(1, 3):
#             op_cond_01 = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[12]/select/option[{op_code}]')  # 停用/啟用
#             op_num_2 = op_cond_01.get_attribute('value')
#             if op_num_2 == '1':
#                 print('(目前)狀態:', op_cond_01.text, '\n')   
            
#         time.sleep(3)    
#         self.driver.refresh()
#         time.sleep(1)   


# # ********************** 試玩遊戲(老虎機) **************************

#     def slot_game_demo_gc_zhCN(self):
#         '''(老虎機)遊戲試玩'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="All"]')  # -- 全選 --                                                             
#         agent_class_select.click()
#         print('代理商類別:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'slot_agent').click()  # 代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="slot_agent"]/option[@value="TEST6"]')  # Test-6
#         agent_select.click() 
#         print('代理商:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         game_num = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[1]')  # 遊戲編號
#         game_num.click() 
#         print('遊戲編號:', game_num.text, '\n')
#         time.sleep(1)

#         game_name = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[3]')  # 遊戲名稱
#         game_name.click() 
#         print('遊戲名稱:', game_name.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[18]/a').click()  # 試玩
#         time.sleep(10)

#         self.driver.switch_to.window(self.driver.window_handles[1])  # 關閉標籤頁
#         self.driver.close()

#         self.driver.switch_to.window(self.driver.window_handles[0])  # 返回主頁
#         time.sleep(2)
        
#         self.driver.refresh()
#         time.sleep(1)


# # ********************** 刪除遊戲(老虎機) **************************

#     def test_slot_game_del_gc_zhCN(self):
#         '''(老虎機)遊戲刪除'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="All"]')  # -- 全選 --                                                             
#         agent_class_select.click()
#         print('代理商類別:', agent_class_select.get_attribute('value'), '\n')

#         self.driver.find_element(By.ID, f'slot_agent').click()  # 代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="slot_agent"]/option[@value="TEST5"]')  # Test-5
#         agent_select.click() 
#         print('代理商:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         game_num = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[1]')  # 遊戲編號
#         game_num.click() 
#         print('遊戲編號:', game_num.text, '\n')
#         time.sleep(1)

#         game_name = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[3]')  # 遊戲名稱
#         game_name.click() 
#         print('遊戲名稱:', game_name.text, '\n')
#         time.sleep(1)
        
#         self.driver.find_element(By.XPATH, '//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[18]/button[3]').click()  # 刪除
#         time.sleep(1) 

#         notice_info = self.driver.find_element(By.XPATH, f'//*[@id="myModal_3"]/div/div/div[2]')  # 提示訊息 (確認刪除 您即將刪除 TEST5 (PSS-ON-00xxx))
#         print(notice_info.text, '\n')
#         time.sleep(2) 

#         self.driver.find_element(By.XPATH, '//*[@id="slot_confirm_submit"]').click()  # 送出
#         time.sleep(1) 

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Configuration/agent_managemen')  # 代理商管理
#         time.sleep(1)

#         filter_box = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0_filter"]/label/input')  # 篩選
#         filter_box.click()
#         filter_box.send_keys('test5')
#         time.sleep(1) 

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[6]/button[5]').click()  # 重啟
#         time.sleep(2)

#         note_info = self.driver.find_element(By.XPATH, f'//*[@id="myModal_3"]/div/div/div[2]')  # 通知訊息 (重啟確認 您目前要重啟的代理是 Test-5 [SID : 1]，確定要送出指令嗎？)
#         print(note_info.text, '\n')
#         time.sleep(3)

#         self.driver.find_element(By.XPATH, f'//*[@id="reset_confirm_submit"]').click()  # 送出 
#         time.sleep(3)

#         note_info_1 = self.driver.find_element(By.XPATH, f'/html/body/div[2]')  # 通知訊息 (重啟成功)
#         print(note_info_1.text, '\n')
#         time.sleep(2) 

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Configuration/game_config')  # 遊戲配置
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1) 

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="All"]')  # -- 全選 --                                                             
#         agent_class_select.click()
#         print('代理商類別:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'slot_agent').click()  # 代理商
#         time.sleep(1) 

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="slot_agent"]/option[@value="TEST5"]')  # Test-5
#         agent_select.click() 
#         print('代理商:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)   


# # ********************** 切換設定(棋牌設定) **************************

#     def test_switch_card_game_interface_gc_zhCN(self):
#         '''切換(棋牌)遊戲功能頁'''

#         self.driver.find_element(By.XPATH, '/html/body/div[3]/div/div[4]/div[2]/div[1]/div/div/div/ul/li[3]/a').click()  # 棋牌設定
#         time.sleep(2)
#         print('切換"棋牌設定"功能頁!')
        

# #  ************************** 代理商類別 **************************

#     def test_agent_class_identify_card_gc_zhCN(self):
#         '''代理商類別功能驗證(棋牌)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)
        
#         agent_class_list = ['All', 'PS', 'Test']
#         for agent_class_name in agent_class_list:
#             agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="{agent_class_name}"]')  # -- 全選 -- / PS / Test                                                            
#             agent_class_select.click()
#             print('代理商類別:', agent_class_select.get_attribute('value'), '\n')
            
#             try:
#                 alertt = self.driver.switch_to.alert
#                 print(alertt.text, '\n')
#                 time.sleep(3)
#                 alertt.accept()
                
#             except NoAlertPresentException:
#                 pass
#             self.driver.refresh()
#             time.sleep(1)
            
#             page_end = self.driver.find_element(By.XPATH, f'/html/body/div[3]/div/div[4]/div[2]/div[2]')  # 頁尾
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)    
#             time.sleep(2)
            
#             page_head = self.driver.find_element(By.CLASS_NAME, f'text-semibold')
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 遊戲配置
#             time.sleep(1)
            
#         self.driver.refresh() 
#         time.sleep(1)
 

# #  ************************** 代理商 **************************

#     def test_agent_identify_card_gc_zhCN(self):
#         '''代理商功能驗證(棋牌)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="All"]')  # -- 全選 --                                                             
#         agent_class_select.click()
#         print('代理商類別:', agent_class_select.get_attribute('value'), '\n')

#         self.driver.find_element(By.ID, f'card_game_agent').click()  # (棋牌)代理商
#         time.sleep(1)

#         agent_list = ['MATH', 'TEST2']
#         for agent_num in agent_list:
#             agent_select = self.driver.find_element(By.XPATH, f'//*[@id="card_game_agent"]/option[@value="{agent_num}"]')  # MATH / Test-2
#             agent_select.click() 
#             print('代理商:', agent_select.get_attribute('value'), '\n')
#             time.sleep(1)

#             page_end = self.driver.find_element(By.XPATH, f'/html/body/div[3]/div/div[4]/div[2]/div[2]')  # 頁尾
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)    
#             time.sleep(2)

#             page_head = self.driver.find_element(By.CLASS_NAME, f'text-semibold') 
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 遊戲配置
#             time.sleep(1)
            
#         self.driver.refresh()
#         time.sleep(1)


# # ********************** (棋牌)遊戲狀態確認 **************************

#     def test_card_game_cond_identify_gc_zhCN(self):
#         '''(棋牌)遊戲狀態確認'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  #  Test                                                            
#         agent_class_select.click()
#         print('代理商類別:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'card_game_agent').click()  # (棋牌)代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="card_game_agent"]/option[@value="TEST3"]')  # Test-3
#         agent_select.click() 
#         print('代理商:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'status_type').click()  # 狀態
#         time.sleep(1)

#         for condition in reversed(range(1, 4)):
#             cond_select = self.driver.find_element(By.XPATH, f'//*[@id="status_type"]/option[{condition}]')  #  -- 全選 -- / 停用 / 啟用
#             cond_select.click() 
#             print('狀態:', cond_select.text, '\n')
#             time.sleep(1)
            
#             page_end = self.driver.find_element(By.XPATH, f'/html/body/div[3]/div/div[4]/div[2]/div[2]')  # 頁尾
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)    
#             time.sleep(2)
            
#             page_head = self.driver.find_element(By.CLASS_NAME, f'text-semibold')
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 遊戲配置
#             time.sleep(1)
        
#         self.driver.refresh()
#         time.sleep(1)   


# # ********************** (棋牌)複製遊戲 **************************

#     def test_card_game_copy_gc_zhCN(self):
#         '''(棋牌)遊戲複製'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="All"]')  # -- 全選 --                                                             
#         agent_class_select.click()
#         print('代理商類別:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'card_game_agent').click()  # (棋牌)代理商
#         time.sleep(1) 

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="card_game_agent"]/option[@value="TEST6"]')  # Test-6
#         agent_select.click() 
#         print('代理商:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'batch_btn').click()  # 批次執行
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'reset_btn').click()  # 更新後重啟
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'currency_search').click()  # 帳戶幣別
#         time.sleep(1)

#         currency_select = self.driver.find_element(By.XPATH, f'//*[@id="currency_search"]/option[@value="CNY"]')  # CNY
#         currency_select.click()
#         print('帳戶幣別:', currency_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         agent_list = ['-1632647290-selectable', '2359048-selectable'] 
#         for agent_id in agent_list:
#             agent_select = self.driver.find_element(By.XPATH, f'//*[@id="{agent_id}"]')  # PLAYSTAR(CNY) / MATH(CNY)
#             self.driver.execute_script("arguments[0].scrollIntoView();", agent_select)
#             agent_select.click()
#             print('欲異動的代理商:', agent_select.get_attribute('textContent'), '\n')
#             time.sleep(1)     
            
#         game_select = self.driver.find_element(By.XPATH, f'//*[@id="727920600-selectable"]')  # RPC-ON-00001 博八博九(0)
#         game_select.click()
#         print('欲複製的遊戲:', game_select.get_attribute('textContent'), '\n')
#         time.sleep(1)  
            
#         self.driver.find_element(By.ID, f'batch_confirm_submit').click()  # 送出
#         time.sleep(3)

#         note_info = self.driver.find_element(By.XPATH, f'/html/body/div[1]')  # 通知訊息
#         print(note_info.text, '\n')
#         time.sleep(2)

#         self.driver.refresh()
#         time.sleep(1)   


# # ********************** 變更遊戲狀態(停用) **************************

#     def test_card_game_cond_dis_gc_zhCN(self):
#         '''(棋牌)變更遊戲狀態(停用)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test                                                             
#         agent_class_select.click()
#         print('代理商類別:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'card_game_agent').click()  # (棋牌)代理商
#         time.sleep(1)
        
#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="card_game_agent"]/option[@value="TEST3"]')  # Test-3
#         agent_select.click() 
#         print('代理商:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)
        
#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[5]/select').click()  # 狀態
#         time.sleep(1)
    
#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[5]/select/option[1]').click()  # 停用
#         time.sleep(1)
        
#         op_cond_0_value = Select(self.driver.find_element(By.XPATH, '//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[5]/select'))
#         op_cond_0_text = op_cond_0_value.first_selected_option
#         print('(目前)狀態:', op_cond_0_text.text, '\n')

#         note_info = self.driver.find_element(By.XPATH, f'/html/body/div[1]')  # 通知訊息
#         print(note_info.text, '\n')
#         time.sleep(2)   
            
#         self.driver.refresh()
#         time.sleep(1)  


# # ********************** 變更遊戲狀態(啟用) **************************

#     def test_card_game_cond_en_gc_zhCN(self):
#         '''(棋牌)變更遊戲狀態(啟用)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test                                                             
#         agent_class_select.click()
#         print('代理商類別:', agent_class_select.get_attribute('value'), '\n')

#         self.driver.find_element(By.ID, f'card_game_agent').click()  # (棋牌)代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="card_game_agent"]/option[@value="TEST3"]')  # Test-3
#         agent_select.click() 
#         print('代理商:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[5]/select').click()  # 狀態
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[5]/select/option[2]').click()  # 啟用
#         time.sleep(1)
             
#         op_cond_1_value = Select(self.driver.find_element(By.XPATH, '//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[5]/select'))
#         op_cond_1_value = op_cond_1_value.first_selected_option
#         print('(目前)狀態:', op_cond_1_value.text, '\n')

#         note_info = self.driver.find_element(By.XPATH, f'/html/body/div[1]')  # 通知訊息 
#         print(note_info.text, '\n')
#         time.sleep(2)   
            
#         self.driver.refresh()
#         time.sleep(1)   


# # ********************** 試玩遊戲(棋牌) **************************

#     def test_card_game_demo_gc_zhCN(self):
#         '''(棋牌)遊戲試玩'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test                                                             
#         agent_class_select.click()
#         print('代理商類別:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'card_game_agent').click()  # 代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="card_game_agent"]/option[@value="TEST2"]')  # Test-2
#         agent_select.click() 
#         print('代理商:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         game_num =self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[1]')  # 遊戲編號
#         game_num.click() 
#         print('遊戲編號:', game_num.text, '\n')
#         time.sleep(1)

#         game_name = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[2]')  # 遊戲名稱
#         game_name.click() 
#         print('遊戲名稱:', game_name.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[6]/a').click()  # 試玩
#         time.sleep(10) 

#         self.driver.switch_to.window(self.driver.window_handles[1])  # 關閉標籤頁
#         self.driver.close()

#         self.driver.switch_to.window(self.driver.window_handles[0])  # 返回主頁
#         self.driver.refresh()
#         time.sleep(1)          


# # ********************** 刪除遊戲(棋牌) **************************

#     def test_card_game_del_gc_zhCN(self):
#         '''(棋牌)遊戲刪除'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)   

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test                                                             
#         agent_class_select.click()
#         print('代理商類別:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'card_game_agent').click()  # 代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="card_game_agent"]/option[@value="TEST3"]')  # Test-3
#         agent_select.click() 
#         print('代理商:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         game_num = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[1]')  # 遊戲編號 
#         game_num.click() 
#         print('遊戲編號:', game_num.text, '\n')
#         time.sleep(1)

#         game_name = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[2]')  # 遊戲名稱 
#         game_name.click() 
#         print('遊戲名稱:', game_name.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[6]/button[2]').click()  # 刪除
#         time.sleep(1) 

#         notice_info = self.driver.find_element(By.XPATH, f'//*[@id="confirm_modal"]/div/div/div[2]')  # 提示訊息 
#         print(notice_info.text, '\n')
#         time.sleep(2) 

#         self.driver.find_element(By.XPATH, '//*[@id="confirm_close"]').click()  # 關閉
#         time.sleep(1) 
        
#         self.driver.find_element(By.XPATH, '//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[6]/button[2]').click()  # 刪除
#         time.sleep(1) 

#         notice_info = self.driver.find_element(By.XPATH, f'//*[@id="confirm_modal"]/div/div/div[2]')  # 提示訊息 
#         print(notice_info.text, '\n')
#         time.sleep(2) 

#         self.driver.find_element(By.XPATH, '//*[@id="confirm_submit"]').click()  # 送出
#         time.sleep(1) 

#         note_info_1 = self.driver.find_element(By.XPATH, f'/html/body/div[1]')  # 通知訊息 (移除成功)
#         print(note_info_1.text, '\n')
#         time.sleep(5) 

#         self.driver.refresh()
#         time.sleep(1)   


# # --------------------------- 遊戲配置(EN) ---------------------------

#     def test_Configuration_Game_Configuration_EN(self):
#         '''【配置】遊戲配置功能頁切換(英)'''  
        
#         self.driver.find_element(By.ID, 'Configuration').click()
#         print("進入配置功能選單!", '\n')
#         time.sleep(2)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Configuration/game_config')
#         print("切換遊戲配置選單!", '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[1]').click()
#         print("語系已切換:", 'English', '\n')
#         time.sleep(1)


# # ********************** 切換設定(老虎機設定) **************************

#     def test_switch_slot_game_interface_gc_EN(self):
#         '''切換(老虎機)遊戲功能頁(英)'''

#         self.driver.find_element(By.XPATH, '/html/body/div[3]/div/div[4]/div[2]/div[1]/div/div/div/ul/li[1]/a').click()  # 老虎機設定
#         time.sleep(2)
#         print('切換"老虎機設定"功能頁!')
        

# #  ************************** Agent Class **************************

#     def test_agent_class_identify_gc_EN(self):
#         '''代理商類別功能驗證(英)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_list = ['All', 'default', 'Test']
#         for agent_class_name in agent_class_list:
#             agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="{agent_class_name}"]')  # -- Select All -- / Default / Test                                                             
#             agent_class_select.click()
#             print('Agent Class:', agent_class_select.get_attribute('value'), '\n')
#             time.sleep(1)
            
#             try:
#                 alertt = self.driver.switch_to.alert
#                 print(alertt.text, '\n')
#                 time.sleep(3)
#                 alertt.accept()
                
#             except NoAlertPresentException:
#                 pass
#             self.driver.refresh()
#             time.sleep(1)
            
#             page_end = self.driver.find_element(By.XPATH, f'/html/body/div[3]/div/div[4]/div[2]/div[2]')  # 頁尾
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)    
#             time.sleep(2)
            
#             page_head = self.driver.find_element(By.CLASS_NAME, f'text-semibold') 
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 遊戲配置
#             time.sleep(1)
            
#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** Agent **************************

#     def test_agent_identify_gc_EN(self):
#         '''代理商功能驗證(英)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="All"]')  # -- Select All --                                                             
#         agent_class_select.click()
#         print('Agent Class:', agent_class_select.get_attribute('value'), '\n')

#         self.driver.find_element(By.ID, f'slot_agent').click()  # (老虎機)代理商
#         time.sleep(1)

#         agent_list = ['MATH', 'TEST2']
#         for agent_num in agent_list:
#             agent_select = self.driver.find_element(By.XPATH, f'//*[@id="slot_agent"]/option[@value="{agent_num}"]')  # MATH / Test-2
#             agent_select.click() 
#             print('Agent:', agent_select.get_attribute('value'), '\n')
#             time.sleep(1)

#             page_end = self.driver.find_element(By.XPATH, f'/html/body/div[3]/div/div[4]/div[2]/div[2]')  # 頁尾
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)    
#             time.sleep(2)

#             page_head = self.driver.find_element(By.CLASS_NAME, f'text-semibold')
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 遊戲配置
#             time.sleep(1)
            
#         self.driver.refresh()
#         time.sleep(1)


# # ********************** (Slot)遊戲狀態確認 **************************

#     def test_slot_game_cond_identify_gc_EN(self):
#         '''(老虎機)遊戲狀態確認(英)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  #  Test                                                            
#         agent_class_select.click()
#         print('Agent Class:', agent_class_select.get_attribute('value'), '\n')

#         self.driver.find_element(By.ID, f'slot_agent').click()  # (老虎機)代理商
#         time.sleep(1) 

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="slot_agent"]/option[@value="TEST5"]')  # Test-5
#         agent_select.click() 
#         print('Agent:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'status_type').click()  # 狀態
#         time.sleep(1)

#         for condition in reversed(range(1, 4)):
#             cond_select = self.driver.find_element(By.XPATH, f'//*[@id="status_type"]/option[{condition}]')  #  -- 全選 -- / 停用 / 啟用
#             cond_select.click() 
#             print('Status:', cond_select.text, '\n')
#             time.sleep(1)
            
#             page_end = self.driver.find_element(By.XPATH, f'/html/body/div[3]/div/div[4]/div[2]/div[2]')  # 頁尾
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)    
#             time.sleep(2)
            
#             page_head = self.driver.find_element(By.CLASS_NAME, f'text-semibold') 
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 遊戲配置
#             time.sleep(1)
        
#         self.driver.refresh()   
#         time.sleep(1)


# # ********************** 複製遊戲(SLOT) **************************

#     def test_slot_game_copy_gc_EN(self):
#         '''(老虎機)遊戲複製(英)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1) 

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="All"]')  # -- 全選 --                                                             
#         agent_class_select.click()
#         print('Agent Class:', agent_class_select.get_attribute('value'), '\n')

#         self.driver.find_element(By.ID, f'slot_agent').click()  # (老虎機)代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="slot_agent"]/option[@value="TEST15"]')  # Test-15
#         agent_select.click() 
#         print('Agent:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'batch_btn').click()  # 批次執行
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'reset_btn').click()  # 更新後重啟
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'currency_search').click()  # 帳戶幣別
#         time.sleep(1)

#         currency_select = self.driver.find_element(By.XPATH, f'//*[@id="currency_search"]/option[@value="CNY"]')  # CNY
#         currency_select.click()
#         print('Account currency:', currency_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         agent_list = ['-1632647290-selectable', '2359048-selectable'] 
#         for agent_id in agent_list:
#             agent_select = self.driver.find_element(By.XPATH, f'//*[@id="{agent_id}"]')  # PLAYSTAR(CNY) / MATH(CNY)
#             self.driver.execute_script("arguments[0].scrollIntoView();", agent_select)
#             agent_select.click()
#             print('Updated agent:', agent_select.get_attribute('textContent'), '\n')
#             time.sleep(1)     
        
#         game_select = self.driver.find_element(By.XPATH, f'//*[@id="-1500625730-selectable"]')  # PSS-ON-00160 熊熊战争
#         game_select.click()
#         print('Copyied game:', game_select.get_attribute('textContent'), '\n')
#         time.sleep(1)  
            
#         self.driver.find_element(By.ID, f'batch_confirm_submit').click()  # 送出
#         time.sleep(3)

#         note_info = self.driver.find_element(By.XPATH, f'/html/body/div[1]')  # 通知訊息
#         print(note_info.text, '\n')
#         time.sleep(2)

#         self.driver.refresh()
#         time.sleep(1)


# # ********************** 變更遊戲狀態(Disabled) **************************

#     def test_slot_game_cond_dis_gc_EN(self):
#         '''(老虎機)變更遊戲狀態(停用)(英)'''    

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test                                                           
#         agent_class_select.click()
#         print('Agent Class:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'slot_agent').click()  # (老虎機)代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="slot_agent"]/option[@value="TEST15"]')  # Test-15
#         agent_select.click() 
#         print('Agent:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[12]/select').click()  # 狀態
#         time.sleep(1)

#         for op_code_0 in range(1, 3):
#             op_cond_0 = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[12]/select/option[{op_code_0}]')  # 停用/啟用
#             op_num_0 = op_cond_0.get_attribute('value')
#             if op_num_0 == '1':
#                 print('(Current)Status:', op_cond_0.text, '\n')    

#         op_status_stop = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[12]/select/option[1]')  # 停用
#         op_status_stop.click() 
#         print('(Updated)Status:', op_status_stop.text, '\n')
#         time.sleep(1)

#         note_info = self.driver.find_element(By.XPATH, f'/html/body/div[1]')  # 通知訊息
#         print(note_info.text, '\n')
#         time.sleep(2)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Configuration/agent_managemen')  # 代理商管理
#         time.sleep(1)

#         filter_box = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0_filter"]/label/input')  # 篩選
#         filter_box.click()
#         filter_box.send_keys('test15')
#         time.sleep(1) 

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[6]/button[5]').click()  # 重啟
#         time.sleep(2)

#         self.driver.find_element(By.XPATH, f'//*[@id="reset_confirm_submit"]').click()  # 送出
#         time.sleep(2)

#         note_info_00 = self.driver.find_element(By.XPATH, f'/html/body/div[2]')  # 通知訊息 
#         print(note_info_00.text, '\n')
#         time.sleep(2)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Configuration/game_config')  # 遊戲配置
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test                                                           
#         agent_class_select.click()
#         print('Agent Class:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'slot_agent').click()  # (老虎機)代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="slot_agent"]/option[@value="TEST15"]')  # Test-15
#         agent_select.click() 
#         print('Agent:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         filter_box_1 = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0_filter"]/label/input')  # 篩選
#         filter_box_1.click()
#         filter_box_1.send_keys('PSS-ON-00156') 
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[12]/select').click()  # 狀態
#         time.sleep(1)

#         for op_code_1 in range(1, 3):
#             op_cond_1 = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[12]/select/option[{op_code_1}]')  # 停用/啟用
#             op_num_1 = op_cond_1.get_attribute('value')
#             if op_num_1 == '0':
#                 print('(Current)Status:', op_cond_1.text, '\n')    
            
#         time.sleep(3)    
#         self.driver.refresh()
#         time.sleep(1)   


# # ********************** 變更遊戲狀態(Enabled) **************************

#     def test_slot_game_cond_en_gc_EN(self):
#         '''(老虎機)變更遊戲狀態(啟用)(英)'''   

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1) 

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test                                                             
#         agent_class_select.click()
#         print('Agent Class:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)
        
#         self.driver.find_element(By.ID, f'slot_agent').click()  # (老虎機)代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="slot_agent"]/option[@value="TEST15"]')  # Test-15
#         agent_select.click() 
#         print('Agent:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[12]/select').click()  # 狀態
#         time.sleep(1)

#         for op_code_0 in range(1, 3):
#             op_cond_0 = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[12]/select/option[{op_code_0}]')  # 停用/啟用
#             op_num_0 = op_cond_0.get_attribute('value')
#             if op_num_0 == '0':
#                 print('(Current)Status:', op_cond_0.text, '\n')
#             time.sleep(1)

#         op_status_stop = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[12]/select/option[2]')  # 啟用
#         op_status_stop.click() 
#         print('(Updated)Status:', op_status_stop.text, '\n')
#         time.sleep(1)

#         note_info = self.driver.find_element(By.XPATH, f'/html/body/div[1]')  # 通知訊息
#         print(note_info.text, '\n')
#         time.sleep(2)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Configuration/agent_managemen')  # 代理商管理
#         time.sleep(1)

#         filter_box = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0_filter"]/label/input')  # 篩選
#         filter_box.click()
#         filter_box.send_keys('test15')
#         time.sleep(1) 

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[6]/button[5]').click()  # 重啟
#         time.sleep(2)

#         self.driver.find_element(By.XPATH, f'//*[@id="reset_confirm_submit"]').click()  # 送出
#         time.sleep(3)

#         note_info_01 = self.driver.find_element(By.XPATH, f'/html/body/div[2]')  # 通知訊息
#         print(note_info_01.text, '\n')
#         time.sleep(2)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Configuration/game_config')  # 遊戲配置
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="All"]')  # -- 全選 --                                                             
#         agent_class_select.click()
#         print('Agent Class:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         filter_box_1 = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0_filter"]/label/input')  # 篩選
#         filter_box_1.click()
#         filter_box_1.send_keys('PSS-ON-00156') 
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[12]/select').click()  # 狀態
#         time.sleep(1)

#         for op_code in range(1, 3):
#             op_cond_01 = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[12]/select/option[{op_code}]')  # 停用/啟用
#             op_num_2 = op_cond_01.get_attribute('value')
#             if op_num_2 == '1':
#                 print('(Current)Status:', op_cond_01.text, '\n')   
            
#         time.sleep(3)    
#         self.driver.refresh()
#         time.sleep(1)   


# # ********************** 試玩遊戲(老虎機) **************************

#     def test_slot_game_demo_gc_EN(self):
#         '''(老虎機)遊戲試玩(英)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="All"]')  # -- 全選 --                                                             
#         agent_class_select.click()
#         print('Agent Class:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'slot_agent').click()  # 代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="slot_agent"]/option[@value="TEST6"]')  # Test-6
#         agent_select.click() 
#         print('Agent:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         game_num = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[1]')  # 遊戲編號
#         game_num.click() 
#         print('Game ID:', game_num.text, '\n')
#         time.sleep(1)

#         game_name = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[3]')  # 遊戲名稱
#         game_name.click() 
#         print('Game Name:', game_name.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[18]/a').click()  # 試玩
#         time.sleep(10)

#         self.driver.switch_to.window(self.driver.window_handles[1])  # 關閉標籤頁
#         self.driver.close()

#         self.driver.switch_to.window(self.driver.window_handles[0])  # 返回主頁
#         time.sleep(2)
        
#         self.driver.refresh()
#         time.sleep(1)


# # ********************** 刪除遊戲(Slot game) **************************
    
#     def test_slot_game_del_gc_EN(self):
#         '''(老虎機)遊戲刪除(英)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="All"]')  # -- 全選 --                                                             
#         agent_class_select.click()
#         print('代理商類別:', agent_class_select.get_attribute('value'), '\n')

#         self.driver.find_element(By.ID, f'slot_agent').click()  # 代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="slot_agent"]/option[@value="TEST5"]')  # Test-5
#         agent_select.click() 
#         print('代理商:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         game_num = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[1]')  # 遊戲編號
#         game_num.click() 
#         print('遊戲編號:', game_num.text, '\n')
#         time.sleep(1)

#         game_name = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[3]')  # 遊戲名稱
#         game_name.click() 
#         print('遊戲名稱:', game_name.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[18]/button[3]').click()  # 刪除
#         time.sleep(1) 

#         notice_info = self.driver.find_element(By.XPATH, f'//*[@id="myModal_3"]/div/div/div[2]')  # 提示訊息 (確認刪除 您即將刪除 TEST5 (PSS-ON-00xxx))
#         print(notice_info.text, '\n')
#         time.sleep(2) 

#         self.driver.find_element(By.XPATH, '//*[@id="slot_confirm_submit"]').click()  # 送出
#         time.sleep(1) 

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Configuration/agent_managemen')  # 代理商管理
#         time.sleep(1)

#         filter_box = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0_filter"]/label/input')  # 篩選
#         filter_box.click()
#         filter_box.send_keys('test5')
#         time.sleep(1) 

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[6]/button[5]').click()  # 重啟
#         time.sleep(2)

#         note_info = self.driver.find_element(By.XPATH, f'//*[@id="myModal_3"]/div/div/div[2]')  # 通知訊息 (重啟確認 您目前要重啟的代理是 Test-5 [SID : 1]，確定要送出指令嗎？)
#         print(note_info.text, '\n')
#         time.sleep(3)

#         self.driver.find_element(By.XPATH, f'//*[@id="reset_confirm_submit"]').click()  # 送出 
#         time.sleep(3)

#         note_info_1 = self.driver.find_element(By.XPATH, f'/html/body/div[2]')  # 通知訊息 (重啟成功)
#         print(note_info_1.text, '\n')
#         time.sleep(2) 

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Configuration/game_config')  # 遊戲配置
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1) 

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="All"]')  # -- 全選 --                                                             
#         agent_class_select.click()
#         print('代理商類別:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'slot_agent').click()  # 代理商
#         time.sleep(1) 

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="slot_agent"]/option[@value="TEST5"]')  # Test-5
#         agent_select.click() 
#         print('代理商:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)   


# # ********************** 切換設定(Chess Setting) **************************

#     def test_switch_card_game_interface_gc_EN(self):
#         '''切換(棋牌)遊戲功能頁(英)'''
    
#         self.driver.find_element(By.XPATH, '/html/body/div[3]/div/div[4]/div[2]/div[1]/div/div/div/ul/li[3]/a').click()  # 棋牌設定
#         time.sleep(2)
#         print('切換"棋牌設定"功能頁!') 
        
 
# #  ************************** Agent Class **************************

#     def test_agent_class_identify_card_gc_EN(self):
#         '''代理商類別功能驗證(棋牌)(英)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)
        
#         agent_class_list = ['All', 'PS', 'Test']
#         for agent_class_name in agent_class_list:
#             agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="{agent_class_name}"]')  # -- Select All -- / PS / Test                                                            
#             agent_class_select.click()
#             print('Agent Class:', agent_class_select.get_attribute('value'), '\n')
            
#             try:
#                 alertt = self.driver.switch_to.alert
#                 print(alertt.text, '\n')
#                 time.sleep(3)
#                 alertt.accept()
                
#             except NoAlertPresentException:
#                 pass
#             self.driver.refresh()
#             time.sleep(1)
            
#             page_end = self.driver.find_element(By.XPATH, f'/html/body/div[3]/div/div[4]/div[2]/div[2]')  # 頁尾
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)    
#             time.sleep(2)
            
#             page_head = self.driver.find_element(By.CLASS_NAME, f'text-semibold')
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 遊戲配置
#             time.sleep(1)
            
#         self.driver.refresh() 
#         time.sleep(1)
 

# #  ************************** Agent **************************

#     def test_agent_identify_card_gc_EN(self):
#         '''代理商功能驗證(棋牌)(英)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="All"]')  # -- 全選 --                                                             
#         agent_class_select.click()
#         print('Agent Class:', agent_class_select.get_attribute('value'), '\n')

#         self.driver.find_element(By.ID, f'card_game_agent').click()  # (棋牌)代理商
#         time.sleep(1)

#         agent_list = ['MATH', 'TEST2']
#         for agent_num in agent_list:
#             agent_select = self.driver.find_element(By.XPATH, f'//*[@id="card_game_agent"]/option[@value="{agent_num}"]')  # MATH / Test-2
#             agent_select.click() 
#             print('Agent:', agent_select.get_attribute('value'), '\n')
#             time.sleep(1)

#             page_end = self.driver.find_element(By.XPATH, f'/html/body/div[3]/div/div[4]/div[2]/div[2]')  # 頁尾
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)    
#             time.sleep(2)

#             page_head = self.driver.find_element(By.CLASS_NAME, f'text-semibold') 
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 遊戲配置
#             time.sleep(1)
            
#         self.driver.refresh()
#         time.sleep(1)


# # ********************** (Chess)遊戲狀態確認 **************************

#     def test_card_game_cond_identify_gc_EN(self):
#         '''(棋牌)遊戲狀態確認(英)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  #  Test                                                            
#         agent_class_select.click()
#         print('Agent Class:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'card_game_agent').click()  # (棋牌)代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="card_game_agent"]/option[@value="TEST3"]')  # Test-3
#         agent_select.click() 
#         print('Agent:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'status_type').click()  # 狀態
#         time.sleep(1)

#         for condition in reversed(range(1, 4)):
#             cond_select = self.driver.find_element(By.XPATH, f'//*[@id="status_type"]/option[{condition}]')  #  -- 全選 -- / 停用 / 啟用
#             cond_select.click() 
#             print('Status:', cond_select.text, '\n')
#             time.sleep(1)
            
#             page_end = self.driver.find_element(By.XPATH, f'/html/body/div[3]/div/div[4]/div[2]/div[2]')  # 頁尾
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)    
#             time.sleep(2)
            
#             page_head = self.driver.find_element(By.CLASS_NAME, f'text-semibold')
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 遊戲配置
#             time.sleep(1)
        
#         self.driver.refresh()
#         time.sleep(1)   


# # ********************** 複製遊戲 **************************

#     def test_card_game_copy_gc_EN(self):
#         '''(棋牌)遊戲複製(英)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="All"]')  # -- 全選 --                                                             
#         agent_class_select.click()
#         print('Agent Class:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'card_game_agent').click()  # (棋牌)代理商
#         time.sleep(1) 

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="card_game_agent"]/option[@value="TEST6"]')  # Test-6
#         agent_select.click() 
#         print('Agent:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'batch_btn').click()  # 批次執行
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'reset_btn').click()  # 更新後重啟
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'currency_search').click()  # 帳戶幣別
#         time.sleep(1)

#         currency_select = self.driver.find_element(By.XPATH, f'//*[@id="currency_search"]/option[@value="CNY"]')  # CNY
#         currency_select.click()
#         print('Account currency:', currency_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         agent_list = ['-1632647290-selectable', '2359048-selectable'] 
#         for agent_id in agent_list:
#             agent_select = self.driver.find_element(By.XPATH, f'//*[@id="{agent_id}"]')  # PLAYSTAR(CNY) / MATH(CNY)
#             self.driver.execute_script("arguments[0].scrollIntoView();", agent_select)
#             agent_select.click()
#             print('Updated agent:', agent_select.get_attribute('textContent'), '\n')
#             time.sleep(1)     
            
#         game_select = self.driver.find_element(By.XPATH, f'//*[@id="727920600-selectable"]')  # RPC-ON-00001 博八博九(0)
#         game_select.click()
#         print('Copied game:', game_select.get_attribute('textContent'), '\n')
#         time.sleep(1)  
            
#         self.driver.find_element(By.ID, f'batch_confirm_submit').click()  # 送出
#         time.sleep(3)

#         note_info = self.driver.find_element(By.XPATH, f'/html/body/div[1]')  # 通知訊息
#         print(note_info.text, '\n')
#         time.sleep(2)

#         self.driver.refresh()
#         time.sleep(1)   


# # ********************** 變更遊戲狀態(Disabled) **************************

#     def test_card_game_cond_dis_gc_EN(self):
#         '''(棋牌)變更遊戲狀態(停用)(英)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test                                                             
#         agent_class_select.click()
#         print('Agent Class:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'card_game_agent').click()  # (棋牌)代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="card_game_agent"]/option[@value="TEST3"]')  # Test-3
#         agent_select.click() 
#         print('Agent:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[5]/select').click()  # 狀態
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[5]/select/option[1]').click()  # 停用
#         time.sleep(1)
        
#         op_cond_0_value = Select(self.driver.find_element(By.XPATH, '//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[5]/select'))
#         op_cond_0_text = op_cond_0_value.first_selected_option
#         print('(Current)Status:', op_cond_0_text.text, '\n')

#         note_info = self.driver.find_element(By.XPATH, f'/html/body/div[1]')  # 通知訊息
#         print(note_info.text, '\n')
#         time.sleep(2)   
            
#         self.driver.refresh()
#         time.sleep(1)    


# # ********************** 變更遊戲狀態(Enabled) **************************

#     def test_card_game_cond_en_gc_EN(self):
#         '''(棋牌)變更遊戲狀態(啟用)(英)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test                                                             
#         agent_class_select.click()
#         print('Agent Class:', agent_class_select.get_attribute('value'), '\n')

#         self.driver.find_element(By.ID, f'card_game_agent').click()  # (棋牌)代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="card_game_agent"]/option[@value="TEST3"]')  # Test-3
#         agent_select.click() 
#         print('Agent:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[5]/select').click()  # 狀態
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[5]/select/option[2]').click()  # 啟用
#         time.sleep(1)
             
#         op_cond_1_value = Select(self.driver.find_element(By.XPATH, '//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[5]/select'))
#         op_cond_1_value = op_cond_1_value.first_selected_option
#         print('(Current)Status:', op_cond_1_value.text, '\n')

#         note_info = self.driver.find_element(By.XPATH, f'/html/body/div[1]')  # 通知訊息 
#         print(note_info.text, '\n')
#         time.sleep(2)   
            
#         self.driver.refresh()
#         time.sleep(1)   


# # ********************** 試玩遊戲(Chess) **************************

#     def card_game_demo_gc_EN(self):
#         '''(棋牌)遊戲試玩(英)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test                                                             
#         agent_class_select.click()
#         print('Agent Class:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'card_game_agent').click()  # 代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="card_game_agent"]/option[@value="TEST2"]')  # Test-2
#         agent_select.click() 
#         print('Agent:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         game_num =self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[1]')  # 遊戲編號
#         game_num.click() 
#         print('Game ID:', game_num.text, '\n')
#         time.sleep(1)

#         game_name = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[2]')  # 遊戲名稱
#         game_name.click() 
#         print('Game Name:', game_name.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[6]/a').click()  # 試玩
#         time.sleep(10) 

#         self.driver.switch_to.window(self.driver.window_handles[1])  # 關閉標籤頁
#         self.driver.close()

#         self.driver.switch_to.window(self.driver.window_handles[0])  # 返回主頁
#         self.driver.refresh()
#         time.sleep(1)    


# # ********************** 刪除遊戲(Chess) **************************

#     def test_card_game_del_gc_EN(self):
#         '''(棋牌)遊戲刪除(英)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)   

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test                                                             
#         agent_class_select.click()
#         print('Agent Class:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'card_game_agent').click()  # 代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="card_game_agent"]/option[@value="TEST3"]')  # Test-3
#         agent_select.click() 
#         print('Agent:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         game_num = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[1]')  # 遊戲編號 
#         game_num.click() 
#         print('Game ID:', game_num.text, '\n')
#         time.sleep(1)

#         game_name = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[2]')  # 遊戲名稱 
#         game_name.click() 
#         print('Game Name:', game_name.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[6]/button[2]').click()  # 刪除
#         time.sleep(1) 

#         notice_info = self.driver.find_element(By.XPATH, f'//*[@id="confirm_modal"]/div/div/div[2]')  # 提示訊息 
#         print(notice_info.text, '\n')
#         time.sleep(2) 

#         self.driver.find_element(By.XPATH, '//*[@id="confirm_close"]').click()  # 關閉
#         time.sleep(1) 

#         self.driver.find_element(By.XPATH, '//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[6]/button[2]').click()  # 刪除
#         time.sleep(1) 

#         notice_info = self.driver.find_element(By.XPATH, f'//*[@id="confirm_modal"]/div/div/div[2]')  # 提示訊息 
#         print(notice_info.text, '\n')
#         time.sleep(2) 

#         self.driver.find_element(By.XPATH, '//*[@id="confirm_submit"]').click()  # 送出
#         time.sleep(1) 

#         note_info_1 = self.driver.find_element(By.XPATH, f'/html/body/div[1]')  # 通知訊息 (移除成功)
#         print(note_info_1.text, '\n')
#         time.sleep(5) 

#         self.driver.refresh()
#         time.sleep(1)   


# # --------------------------- 遊戲配置(Tai) ---------------------------

#     def test_Configuration_Game_Configuration_Tai(self):
#         '''【配置】遊戲配置功能頁切換(泰)'''  
        
#         self.driver.find_element(By.ID, 'Configuration').click()
#         print("進入配置功能選單!", '\n')
#         time.sleep(2)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Configuration/game_config')
#         print("切換遊戲配置選單!", '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[4]').click()
#         print("語系已切換:", 'ไทย', '\n')
#         time.sleep(1)
    

# # ********************** 切換設定(老虎機設定) **************************

#     def test_switch_slot_game_interface_gc_Tai(self):
#         '''切換(老虎機)遊戲功能頁(泰)'''

#         self.driver.find_element(By.XPATH, '/html/body/div[3]/div/div[4]/div[2]/div[1]/div/div/div/ul/li[1]/a').click()  # 老虎機設定
#         time.sleep(2)
        

# #  ************************** Agent Class **************************

#     def test_agent_class_identify_gc_Tai(self):
#         '''代理商類別功能驗證(泰)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_list = ['All', 'default', 'Test']
#         for agent_class_name in agent_class_list:
#             agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="{agent_class_name}"]')  # -- เลือกทั้งหมด -- / Default / Test                                                             
#             agent_class_select.click()
#             print('หมวดตัวแทน:', agent_class_select.get_attribute('value'), '\n')
#             time.sleep(1)
            
#             try:
#                 alertt = self.driver.switch_to.alert
#                 print(alertt.text, '\n')
#                 time.sleep(3)
#                 alertt.accept()
                
#             except NoAlertPresentException:
#                 pass
#             self.driver.refresh()
#             time.sleep(1)
            
#             page_end = self.driver.find_element(By.XPATH, f'/html/body/div[3]/div/div[4]/div[2]/div[2]')  # 頁尾
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)    
#             time.sleep(2)
            
#             page_head = self.driver.find_element(By.CLASS_NAME, f'text-semibold') 
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 遊戲配置
#             time.sleep(1)
            
#         self.driver.refresh()
#         time.sleep(1)


# #  ************************** 代理商 **************************

#     def test_agent_identify_gc_Tai(self):
#         '''代理商功能驗證(泰)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="All"]')  # -- Select All --                                                             
#         agent_class_select.click()
#         print('หมวดตัวแทน:', agent_class_select.get_attribute('value'), '\n')

#         self.driver.find_element(By.ID, f'slot_agent').click()  # (老虎機)代理商
#         time.sleep(1)

#         agent_list = ['MATH', 'TEST2']
#         for agent_num in agent_list:
#             agent_select = self.driver.find_element(By.XPATH, f'//*[@id="slot_agent"]/option[@value="{agent_num}"]')  # MATH / Test-2
#             agent_select.click() 
#             print('ตัวแทน:', agent_select.get_attribute('value'), '\n')
#             time.sleep(1)

#             page_end = self.driver.find_element(By.XPATH, f'/html/body/div[3]/div/div[4]/div[2]/div[2]')  # 頁尾
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)    
#             time.sleep(2)

#             page_head = self.driver.find_element(By.CLASS_NAME, f'text-semibold')
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 遊戲配置
#             time.sleep(1)
            
#         self.driver.refresh()
#         time.sleep(1)
                                     

# # ********************** (Slot)遊戲狀態確認 **************************

#     def test_slot_game_cond_identify_gc_Tai(self):
#         '''(老虎機)遊戲狀態確認(泰)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  #  Test                                                            
#         agent_class_select.click()
#         print('หมวดตัวแทน:', agent_class_select.get_attribute('value'), '\n')

#         self.driver.find_element(By.ID, f'slot_agent').click()  # (老虎機)代理商
#         time.sleep(1) 

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="slot_agent"]/option[@value="TEST5"]')  # Test-5
#         agent_select.click() 
#         print('ตัวแทน:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'status_type').click()  # 狀態
#         time.sleep(1)

#         for condition in reversed(range(1, 4)):
#             cond_select = self.driver.find_element(By.XPATH, f'//*[@id="status_type"]/option[{condition}]')  #  -- เลือกทั้งหมด -- / 停用 / 啟用
#             cond_select.click() 
#             print('สถานะ:', cond_select.text, '\n')
#             time.sleep(1)
            
#             page_end = self.driver.find_element(By.XPATH, f'/html/body/div[3]/div/div[4]/div[2]/div[2]')  # 頁尾
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)    
#             time.sleep(2)
            
#             page_head = self.driver.find_element(By.CLASS_NAME, f'text-semibold') 
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 遊戲配置
#             time.sleep(1)
        
#         self.driver.refresh()   
#         time.sleep(1)


# # ********************** 複製遊戲 **************************

#     def test_slot_game_copy_gc_Tai(self):
#         '''(老虎機)遊戲複製(泰)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1) 

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="All"]')  # -- 全選 --                                                             
#         agent_class_select.click()
#         print('หมวดตัวแทน:', agent_class_select.get_attribute('value'), '\n')

#         self.driver.find_element(By.ID, f'slot_agent').click()  # (老虎機)代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="slot_agent"]/option[@value="TEST15"]')  # Test-15
#         agent_select.click() 
#         print('ตัวแทน:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'batch_btn').click()  # 批次執行
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'reset_btn').click()  # 更新後重啟
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'currency_search').click()  # 帳戶幣別
#         time.sleep(1)

#         currency_select = self.driver.find_element(By.XPATH, f'//*[@id="currency_search"]/option[@value="CNY"]')  # CNY
#         currency_select.click()
#         print('สกุลเงินในบัญชี:', currency_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         agent_list = ['-1632647290-selectable', '2359048-selectable'] 
#         for agent_id in agent_list:
#             agent_select = self.driver.find_element(By.XPATH, f'//*[@id="{agent_id}"]')  # PLAYSTAR(CNY) / MATH(CNY)
#             self.driver.execute_script("arguments[0].scrollIntoView();", agent_select)
#             agent_select.click()
#             print('Updated agent:', agent_select.get_attribute('textContent'), '\n')
#             time.sleep(1)     
        
#         game_select = self.driver.find_element(By.XPATH, f'//*[@id="-1500625730-selectable"]')  # PSS-ON-00160 熊熊战争
#         game_select.click()
#         print('Copyied game:', game_select.get_attribute('textContent'), '\n')
#         time.sleep(1)  
            
#         self.driver.find_element(By.ID, f'batch_confirm_submit').click()  # 送出
#         time.sleep(3)

#         note_info = self.driver.find_element(By.XPATH, f'/html/body/div[1]')  # 通知訊息
#         print(note_info.text, '\n')
#         time.sleep(2)

#         self.driver.refresh()
#         time.sleep(1)


# # ********************** 變更遊戲狀態(ปิดใช้งาน) **************************

#     def test_slot_game_cond_dis_gc_Tai(self):
#         '''(老虎機)變更遊戲狀態(停用)(泰)'''    

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test                                                           
#         agent_class_select.click()
#         print('หมวดตัวแทน:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'slot_agent').click()  # (老虎機)代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="slot_agent"]/option[@value="TEST15"]')  # Test-15
#         agent_select.click() 
#         print('ตัวแทน:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[12]/select').click()  # 狀態
#         time.sleep(1)

#         for op_code_0 in range(1, 3):
#             op_cond_0 = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[12]/select/option[{op_code_0}]')  # 停用/啟用
#             op_num_0 = op_cond_0.get_attribute('value')
#             if op_num_0 == '1':
#                 print('(Current)สถานะ:', op_cond_0.text, '\n')    

#         op_status_stop = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[12]/select/option[1]')  # 停用
#         op_status_stop.click() 
#         print('(Updated)สถานะ:', op_status_stop.text, '\n')
#         time.sleep(1)

#         note_info = self.driver.find_element(By.XPATH, f'/html/body/div[1]')  # 通知訊息
#         print(note_info.text, '\n')
#         time.sleep(2)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Configuration/agent_managemen')  # 代理商管理
#         time.sleep(1)

#         filter_box = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0_filter"]/label/input')  # 篩選
#         filter_box.click()
#         filter_box.send_keys('test15')
#         time.sleep(1) 

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[6]/button[5]').click()  # 重啟
#         time.sleep(2)

#         self.driver.find_element(By.XPATH, f'//*[@id="reset_confirm_submit"]').click()  # 送出
#         time.sleep(2)

#         note_info_00 = self.driver.find_element(By.XPATH, f'/html/body/div[2]')  # 通知訊息 
#         print(note_info_00.text, '\n')
#         time.sleep(2)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Configuration/game_config')  # 遊戲配置
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test                                                           
#         agent_class_select.click()
#         print('หมวดตัวแทน:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'slot_agent').click()  # (老虎機)代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="slot_agent"]/option[@value="TEST15"]')  # Test-15
#         agent_select.click() 
#         print('ตัวแทน:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         filter_box_1 = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0_filter"]/label/input')  # 篩選
#         filter_box_1.click()
#         filter_box_1.send_keys('PSS-ON-00156') 
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[12]/select').click()  # 狀態
#         time.sleep(1)

#         for op_code_1 in range(1, 3):
#             op_cond_1 = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[12]/select/option[{op_code_1}]')  # 停用/啟用
#             op_num_1 = op_cond_1.get_attribute('value')
#             if op_num_1 == '0':
#                 print('(Current)Status:', op_cond_1.text, '\n')    
            
#         time.sleep(3)    
#         self.driver.refresh()
#         time.sleep(1)   


# # ********************** 變更遊戲狀態(เปิดใช้งาน) **************************

#     def test_slot_game_cond_en_gc_Tai(self):
#         '''(老虎機)變更遊戲狀態(啟用)(泰)'''  

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1) 

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test                                                             
#         agent_class_select.click()
#         print('หมวดตัวแทน:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)
        
#         self.driver.find_element(By.ID, f'slot_agent').click()  # (老虎機)代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="slot_agent"]/option[@value="TEST15"]')  # Test-15
#         agent_select.click() 
#         print('ตัวแทน:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[12]/select').click()  # 狀態
#         time.sleep(1)

#         for op_code_0 in range(1, 3):
#             op_cond_0 = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[12]/select/option[{op_code_0}]')  # 停用/啟用
#             op_num_0 = op_cond_0.get_attribute('value')
#             if op_num_0 == '0':
#                 print('(Current)สถานะ:', op_cond_0.text, '\n')
#             time.sleep(1)

#         op_status_stop = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[12]/select/option[2]')  # 啟用
#         op_status_stop.click() 
#         print('(Updated)สถานะ:', op_status_stop.text, '\n')
#         time.sleep(1)

#         note_info = self.driver.find_element(By.XPATH, f'/html/body/div[1]')  # 通知訊息
#         print(note_info.text, '\n')
#         time.sleep(2)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Configuration/agent_managemen')  # 代理商管理
#         time.sleep(1)

#         filter_box = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0_filter"]/label/input')  # 篩選
#         filter_box.click()
#         filter_box.send_keys('test15')
#         time.sleep(1) 

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[6]/button[5]').click()  # 重啟
#         time.sleep(2)

#         self.driver.find_element(By.XPATH, f'//*[@id="reset_confirm_submit"]').click()  # 送出
#         time.sleep(3)

#         note_info_01 = self.driver.find_element(By.XPATH, f'/html/body/div[2]')  # 通知訊息
#         print(note_info_01.text, '\n')
#         time.sleep(2)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Configuration/game_config')  # 遊戲配置
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="All"]')  # -- 全選 --                                                             
#         agent_class_select.click()
#         print('หมวดตัวแทน:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         filter_box_1 = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0_filter"]/label/input')  # 篩選
#         filter_box_1.click()
#         filter_box_1.send_keys('PSS-ON-00156') 
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[12]/select').click()  # 狀態
#         time.sleep(1)

#         for op_code in range(1, 3):
#             op_cond_01 = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr/td[12]/select/option[{op_code}]')  # 停用/啟用
#             op_num_2 = op_cond_01.get_attribute('value')
#             if op_num_2 == '1':
#                 print('(Current)สถานะ:', op_cond_01.text, '\n')   
            
#         time.sleep(3)    
#         self.driver.refresh()
#         time.sleep(1)    


# # ********************** 試玩遊戲(老虎機) **************************

#     def test_slot_game_demo_gc_Tai(self):
#         '''(老虎機)遊戲試玩(泰)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="All"]')  # -- 全選 --                                                             
#         agent_class_select.click()
#         print('หมวดตัวแทน:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'slot_agent').click()  # 代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="slot_agent"]/option[@value="TEST6"]')  # Test-6
#         agent_select.click() 
#         print('ตัวแทน:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         game_num = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[1]')  # 遊戲編號
#         game_num.click() 
#         print('หมายเลข:', game_num.text, '\n')
#         time.sleep(1)

#         game_name = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[3]')  # 遊戲名稱
#         game_name.click() 
#         print('ชื่อเกม:', game_name.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[18]/a').click()  # 試玩
#         time.sleep(10)

#         self.driver.switch_to.window(self.driver.window_handles[1])  # 關閉標籤頁
#         self.driver.close()

#         self.driver.switch_to.window(self.driver.window_handles[0])  # 返回主頁
#         time.sleep(2)
        
#         self.driver.refresh()
#         time.sleep(1)


# # ********************** 刪除遊戲(Slot game) **************************

#     def test_slot_game_del_gc_Tai(self):
#         '''(老虎機)遊戲刪除(泰)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="All"]')  # -- 全選 --                                                             
#         agent_class_select.click()
#         print('หมวดตัวแทน:', agent_class_select.get_attribute('value'), '\n')

#         self.driver.find_element(By.ID, f'slot_agent').click()  # 代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="slot_agent"]/option[@value="TEST5"]')  # Test-5
#         agent_select.click() 
#         print('ตัวแทน:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         game_num = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[1]')  # 遊戲編號
#         game_num.click() 
#         print('หมายเลข:', game_num.text, '\n')
#         time.sleep(1)

#         game_name = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[3]')  # 遊戲名稱
#         game_name.click() 
#         print('ชื่อเกม:', game_name.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[18]/button[3]').click()  # 刪除
#         time.sleep(1) 

#         notice_info = self.driver.find_element(By.XPATH, f'//*[@id="myModal_3"]/div/div/div[2]')  # 提示訊息 (確認刪除 您即將刪除 TEST5 (PSS-ON-00xxx))
#         print(notice_info.text, '\n')
#         time.sleep(2) 

#         self.driver.find_element(By.XPATH, '//*[@id="slot_confirm_submit"]').click()  # 送出
#         time.sleep(1) 

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Configuration/agent_managemen')  # 代理商管理
#         time.sleep(1)

#         filter_box = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0_filter"]/label/input')  # 篩選
#         filter_box.click()
#         filter_box.send_keys('test5')
#         time.sleep(1) 

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[6]/button[5]').click()  # 重啟
#         time.sleep(2)

#         note_info = self.driver.find_element(By.XPATH, f'//*[@id="myModal_3"]/div/div/div[2]')  # 通知訊息 (重啟確認 您目前要重啟的代理是 Test-5 [SID : 1]，確定要送出指令嗎？)
#         print(note_info.text, '\n')
#         time.sleep(3)

#         self.driver.find_element(By.XPATH, f'//*[@id="reset_confirm_submit"]').click()  # 送出 
#         time.sleep(3)

#         note_info_1 = self.driver.find_element(By.XPATH, f'/html/body/div[2]')  # 通知訊息 (重啟成功)
#         print(note_info_1.text, '\n')
#         time.sleep(2) 

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Configuration/game_config')  # 遊戲配置
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1) 

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="All"]')  # -- 全選 --                                                             
#         agent_class_select.click()
#         print('หมวดตัวแทน:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'slot_agent').click()  # 代理商
#         time.sleep(1) 

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="slot_agent"]/option[@value="TEST5"]')  # Test-5
#         agent_select.click() 
#         print('ตัวแทน:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)   


# # ********************** 切換設定(กำหนดค่าเกม) **************************

#     def test_switch_card_game_interface_gc_Tai(self):
#         '''切換(棋牌)遊戲功能頁(泰)'''

#         self.driver.find_element(By.XPATH, '/html/body/div[3]/div/div[4]/div[2]/div[1]/div/div/div/ul/li[3]/a').click()  # 棋牌設定
#         time.sleep(2) 
#         print('切換"棋牌設定"功能頁!') 
        
 
# #  ************************** 代理商類別 **************************

#     def test_agent_class_identify_card_gc_Tai(self):
#         '''代理商類別功能驗證(棋牌)(泰)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)
        
#         agent_class_list = ['All', 'PS', 'Test']
#         for agent_class_name in agent_class_list:
#             agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="{agent_class_name}"]')  # -- เลือกทั้งหมด -- / PS / Test                                                            
#             agent_class_select.click()
#             print('หมวดตัวแทน:', agent_class_select.get_attribute('value'), '\n')
            
#             try:
#                 alertt = self.driver.switch_to.alert
#                 print(alertt.text, '\n')
#                 time.sleep(3)
#                 alertt.accept()
                
#             except NoAlertPresentException:
#                 pass
#             self.driver.refresh()
#             time.sleep(1)
            
#             page_end = self.driver.find_element(By.XPATH, f'/html/body/div[3]/div/div[4]/div[2]/div[2]')  # 頁尾
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)    
#             time.sleep(2)
            
#             page_head = self.driver.find_element(By.CLASS_NAME, f'text-semibold')
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 遊戲配置
#             time.sleep(1)
            
#         self.driver.refresh() 
#         time.sleep(1)
 

# #  ************************** 代理商 **************************

#     def test_agent_identify_card_gc_Tai(self):
#         '''代理商功能驗證(棋牌)(泰)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="All"]')  # -- 全選 --                                                             
#         agent_class_select.click()
#         print('หมวดตัวแทน:', agent_class_select.get_attribute('value'), '\n')

#         self.driver.find_element(By.ID, f'card_game_agent').click()  # (棋牌)代理商
#         time.sleep(1)

#         agent_list = ['MATH', 'TEST2']
#         for agent_num in agent_list:
#             agent_select = self.driver.find_element(By.XPATH, f'//*[@id="card_game_agent"]/option[@value="{agent_num}"]')  # MATH / Test-2
#             agent_select.click() 
#             print('ตัวแทน:', agent_select.get_attribute('value'), '\n')
#             time.sleep(1)

#             page_end = self.driver.find_element(By.XPATH, f'/html/body/div[3]/div/div[4]/div[2]/div[2]')  # 頁尾
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)    
#             time.sleep(2)

#             page_head = self.driver.find_element(By.CLASS_NAME, f'text-semibold') 
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 遊戲配置
#             time.sleep(1)
            
#         self.driver.refresh()
#         time.sleep(1)


# # ********************** (Chess)遊戲狀態確認 **************************
    
#     def test_card_game_cond_identify_gc_Tai(self):
#         '''(棋牌)遊戲狀態確認(泰)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  #  Test                                                            
#         agent_class_select.click()
#         print('หมวดตัวแทน:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'card_game_agent').click()  # (棋牌)代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="card_game_agent"]/option[@value="TEST3"]')  # Test-3
#         agent_select.click() 
#         print('ตัวแทน:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'status_type').click()  # 狀態
#         time.sleep(1)

#         for condition in reversed(range(1, 4)):
#             cond_select = self.driver.find_element(By.XPATH, f'//*[@id="status_type"]/option[{condition}]')  #  -- เลือกทั้งหมด -- / 停用 / 啟用
#             cond_select.click() 
#             print('สถานะ:', cond_select.text, '\n')
#             time.sleep(1)
            
#             page_end = self.driver.find_element(By.XPATH, f'/html/body/div[3]/div/div[4]/div[2]/div[2]')  # 頁尾
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_end)    
#             time.sleep(2)
            
#             page_head = self.driver.find_element(By.CLASS_NAME, f'text-semibold')
#             self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 遊戲配置
#             time.sleep(1)
        
#         self.driver.refresh()
#         time.sleep(1)   


# # ********************** 複製遊戲 **************************

#     def test_card_game_copy_gc_Tai(self):
#         '''(棋牌)遊戲複製(泰)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="All"]')  # -- 全選 --                                                             
#         agent_class_select.click()
#         print('หมวดตัวแทน:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'card_game_agent').click()  # (棋牌)代理商
#         time.sleep(1) 

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="card_game_agent"]/option[@value="TEST6"]')  # Test-6
#         agent_select.click() 
#         print('ตัวแทน:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'batch_btn').click()  # 批次執行
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'reset_btn').click()  # 更新後重啟
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'currency_search').click()  # 帳戶幣別
#         time.sleep(1)

#         currency_select = self.driver.find_element(By.XPATH, f'//*[@id="currency_search"]/option[@value="CNY"]')  # CNY
#         currency_select.click()
#         print('สกุลเงินในบัญชี:', currency_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         agent_list = ['-1632647290-selectable', '2359048-selectable'] 
#         for agent_id in agent_list:
#             agent_select = self.driver.find_element(By.XPATH, f'//*[@id="{agent_id}"]')  # PLAYSTAR(CNY) / MATH(CNY)
#             self.driver.execute_script("arguments[0].scrollIntoView();", agent_select)
#             agent_select.click()
#             print('Updated agent:', agent_select.get_attribute('textContent'), '\n')
#             time.sleep(1)     
            
#         game_select = self.driver.find_element(By.XPATH, f'//*[@id="727920600-selectable"]')  # RPC-ON-00001 博八博九(0)
#         game_select.click()
#         print('Copied game:', game_select.get_attribute('textContent'), '\n')
#         time.sleep(1)  
            
#         self.driver.find_element(By.ID, f'batch_confirm_submit').click()  # 送出
#         time.sleep(3)

#         note_info = self.driver.find_element(By.XPATH, f'/html/body/div[1]')  # 通知訊息
#         print(note_info.text, '\n')
#         time.sleep(2)

#         self.driver.refresh()
#         time.sleep(1)   


# # ********************** 變更遊戲狀態(ปิดใช้งาน) **************************

#     def test_card_game_cond_dis_gc_Tai(self):
#         '''(棋牌)變更遊戲狀態(停用)(泰)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test                                                             
#         agent_class_select.click()
#         print('หมวดตัวแทน:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'card_game_agent').click()  # (棋牌)代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="card_game_agent"]/option[@value="TEST3"]')  # Test-3
#         agent_select.click() 
#         print('ตัวแทน:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[5]/select').click()  # 狀態
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[5]/select/option[1]').click()  # 停用
#         time.sleep(1)
        
#         op_cond_0_value = Select(self.driver.find_element(By.XPATH, '//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[5]/select'))
#         op_cond_0_text = op_cond_0_value.first_selected_option
#         print('(Current)สถานะ:', op_cond_0_text.text, '\n')

#         note_info = self.driver.find_element(By.XPATH, f'/html/body/div[1]')  # 通知訊息
#         print(note_info.text, '\n')
#         time.sleep(2)   
            
#         self.driver.refresh()
#         time.sleep(1)    


# # ********************** 變更遊戲狀態(หมวดตัวแทน) **************************

#     def test_card_game_cond_en_gc_Tai(self):
#         '''(棋牌)變更遊戲狀態(啟用)(泰)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test                                                             
#         agent_class_select.click()
#         print('หมวดตัวแทน:', agent_class_select.get_attribute('value'), '\n')

#         self.driver.find_element(By.ID, f'card_game_agent').click()  # (棋牌)代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="card_game_agent"]/option[@value="TEST3"]')  # Test-3
#         agent_select.click() 
#         print('ตัวแทน:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[5]/select').click()  # 狀態
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[5]/select/option[2]').click()  # 啟用
#         time.sleep(1)
             
#         op_cond_1_value = Select(self.driver.find_element(By.XPATH, '//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[5]/select'))
#         op_cond_1_value = op_cond_1_value.first_selected_option
#         print('(Current)สถานะ:', op_cond_1_value.text, '\n')

#         note_info = self.driver.find_element(By.XPATH, f'/html/body/div[1]')  # 通知訊息 
#         print(note_info.text, '\n')
#         time.sleep(2)   
            
#         self.driver.refresh()
#         time.sleep(1)   


# # ********************** 試玩遊戲(Chess) **************************

#     def test_card_game_demo_gc_Tai(self):
#         '''(棋牌)遊戲試玩(泰)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test                                                             
#         agent_class_select.click()
#         print('หมวดตัวแทน:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'card_game_agent').click()  # 代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="card_game_agent"]/option[@value="TEST2"]')  # Test-2
#         agent_select.click() 
#         print('ตัวแทน:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         game_num =self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[1]')  # 遊戲編號
#         game_num.click() 
#         print('Game ID:', game_num.text, '\n')
#         time.sleep(1)

#         game_name = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[2]')  # 遊戲名稱
#         game_name.click() 
#         print('Game Name:', game_name.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[6]/a').click()  # 試玩
#         time.sleep(10) 

#         self.driver.switch_to.window(self.driver.window_handles[1])  # 關閉標籤頁
#         self.driver.close()

#         self.driver.switch_to.window(self.driver.window_handles[0])  # 返回主頁
#         self.driver.refresh()
#         time.sleep(1)    


# # ********************** 刪除遊戲(Chess) **************************

#     def test_card_game_del_gc_Tai(self):
#         '''(棋牌)遊戲刪除(泰)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1)   

#         agent_class_select = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test                                                             
#         agent_class_select.click()
#         print('หมวดตัวแทน:', agent_class_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'card_game_agent').click()  # 代理商
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="card_game_agent"]/option[@value="TEST3"]')  # Test-3
#         agent_select.click() 
#         print('ตัวแทน:', agent_select.get_attribute('value'), '\n')
#         time.sleep(1)

#         game_num = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[1]')  # 遊戲編號 
#         game_num.click() 
#         print('Game ID:', game_num.text, '\n')
#         time.sleep(1)

#         game_name = self.driver.find_element(By.XPATH, f'//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[2]')  # 遊戲名稱 
#         game_name.click() 
#         print('Game Name:', game_name.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[6]/button[2]').click()  # 刪除
#         time.sleep(1) 

#         notice_info = self.driver.find_element(By.XPATH, f'//*[@id="confirm_modal"]/div/div/div[2]')  # 提示訊息 
#         print(notice_info.text, '\n')
#         time.sleep(2) 

#         self.driver.find_element(By.XPATH, '//*[@id="confirm_close"]').click()  # 關閉
#         time.sleep(1) 

#         self.driver.find_element(By.XPATH, '//*[@id="DataTables_Table_0"]/tbody/tr[1]/td[6]/button[2]').click()  # 刪除
#         time.sleep(1) 

#         notice_info = self.driver.find_element(By.XPATH, f'//*[@id="confirm_modal"]/div/div/div[2]')  # 提示訊息 
#         print(notice_info.text, '\n')
#         time.sleep(2) 

#         self.driver.find_element(By.XPATH, '//*[@id="confirm_submit"]').click()  # 送出
#         time.sleep(1) 

#         note_info_1 = self.driver.find_element(By.XPATH, f'/html/body/div[1]')  # 通知訊息 (移除成功)
#         print(note_info_1.text, '\n')
#         time.sleep(5) 

#         self.driver.refresh()
#         time.sleep(1)   


# # ============================= 後台功能巡測【配置】============================= 
# # *************************** RTP配置 ***************************

#     def test_Configuration_RTP_Configuration_zhCN(self):
#         '''【配置】RTP配置功能頁切換'''     
        
#         self.driver.find_element(By.ID, 'Configuration').click()
#         print("進入配置功能選單!", '\n')
#         time.sleep(2)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Configuration/rtp_config')
#         print("切換RTP配置選單!", '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[3]').click()
#         print("語系已切換'简体中文'!", '\n')
#         time.sleep(1)


# #  ************************** 搜尋類別 **************************

#     def test_search_class_identify_rtp_zhCN(self):
#         '''搜尋類別功能驗證'''

#         self.driver.find_element(By.ID, f'search_class').click()  # 搜尋類別
#         time.sleep(1)

#         for Search_class_list in reversed(range(1, 3)):    
#             Search_class_select = self.driver.find_element(By.XPATH, f'//*[@id="search_class"]/option[{Search_class_list}]')  # 勾選 / 文字                                                            
#             Search_class_select.click()
#             print('搜尋類別:', Search_class_select.text, '\n')
#             time.sleep(1)
            
#         self.driver.refresh()
#         time.sleep(1)


# # ********************** 代理商類別與代理商選取 **************************

#     def test_agent_class_and_agent_identify_zhCN(self):
#         '''代理商類別與代理商選取功能驗證'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # 代理商類別
#         time.sleep(1) 

#         agent_class_click = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test                                                            
#         agent_class_click.click()
#         print('代理商類別:', agent_class_click.text, '\n')
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="79713763-selectable"]')  # Test-5
#         agent_select.click() 
#         print('(勾選)代理商:', agent_select.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'sh_btn').click()  # 送出
#         time.sleep(1)

#         page_end = self.driver.find_element(By.XPATH, f'/html/body/div[4]/div/div[4]/div[2]/div[2]')  # 頁尾
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)
#         time.sleep(1)

#         page_head = self.driver.find_element(By.XPATH, f'/html/body/div[4]/div/div[4]/div[2]/div[1]/div[1]/div/div[2]/h5') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 頁首
#         time.sleep(1)

#         agent_select_1 = self.driver.find_element(By.XPATH, f'//*[@id="79713763-selectable"]')  # Test-5
#         agent_select_1.click() 
#         print('(取消勾選)代理商:', agent_select_1.text, '\n')
#         time.sleep(2)

#         agent_select_2 = self.driver.find_element(By.XPATH, f'//*[@id="79713760-selectable"]')  # Test-2
#         agent_select_2.click() 
#         print('(勾選)代理商:', agent_select_2.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'sh_btn').click()  # 送出 
#         time.sleep(1)

#         page_end = self.driver.find_element(By.XPATH, f'/html/body/div[4]/div/div[4]/div[2]/div[2]')  # 頁尾
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)
#         time.sleep(1)

#         page_head = self.driver.find_element(By.XPATH, f'/html/body/div[4]/div/div[4]/div[2]/div[1]/div[1]/div/div[2]/h5') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 頁首
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# # ********************** 代理商類別選取與代理商輸入 **************************

#     def test_agent_class_and_agent_input_identify_zhCN(self):
#         '''代理商類別選取與代理商輸入功能驗證'''

#         self.driver.find_element(By.ID, f'search_class').click()  # 搜尋類別
#         time.sleep(1)

#         Search_class_select = self.driver.find_element(By.XPATH, f'//*[@id="search_class"]/option[@value="1"]')  # 文字                                                            
#         Search_class_select.click()
#         print('搜尋類別:', Search_class_select.text, '\n')
#         time.sleep(1)

#         agent_input_box = self.driver.find_element(By.ID, f'batch_host_id')  # 代理商輸入框
#         agent_input_box.click() 
#         agent_input_box.send_keys('TEST')
#         time.sleep(1)
#         agent_input_box.send_keys(Keys.ENTER)
#         agent_input_box.send_keys('TEST3')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'sh_btn').click()  # 送出
#         time.sleep(1)

#         page_end = self.driver.find_element(By.XPATH, f'/html/body/div[4]/div/div[4]/div[2]/div[2]')  # 頁尾
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)
#         time.sleep(1)

#         page_head = self.driver.find_element(By.XPATH, f'/html/body/div[4]/div/div[4]/div[2]/div[1]/div[1]/div/div[2]/h5')
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 頁首
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)
        

# # ********************** 不輸入代理商(提示訊息驗證) **************************

#     def test_non_agent_data_and_notice_info_identify_zhCN(self):
#         '''不輸入代理商(提示訊息驗證)'''

#         self.driver.find_element(By.ID, f'search_class').click()  # 搜尋類別
#         time.sleep(1)

#         Search_class_select = self.driver.find_element(By.XPATH, f'//*[@id="search_class"]/option[@value="1"]')  # 文字                                                            
#         Search_class_select.click()
#         print('搜尋類別:', Search_class_select.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'sh_btn').click()  # 送出  
#         time.sleep(1)

#         alert_info = self.driver.switch_to.alert.text
#         print(str(alert_info), '\n')
#         time.sleep(2)
#         self.driver.switch_to.alert.accept()
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# # ********************** 代理商類別選取與代理商輸入(錯誤訊息驗證) **************************

#     def test_agent_class_and_agent_input_notice_info_identify_zhCN(self):  
#         '''代理商類別選取與代理商輸入(錯誤訊息驗證)'''

#         self.driver.find_element(By.ID, f'search_class').click()  # 搜尋類別
#         time.sleep(1)

#         Search_class_select = self.driver.find_element(By.XPATH, f'//*[@id="search_class"]/option[@value="1"]')  # 文字                                                            
#         Search_class_select.click()
#         print('搜尋類別:', Search_class_select.text, '\n')
#         time.sleep(1)

#         agent_input_box = self.driver.find_element(By.ID, f'batch_host_id')  # 代理商輸入框
#         agent_input_box.click() 
#         agent_input_box.send_keys('TEST-2')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'sh_btn').click()  # 送出
#         time.sleep(1)

#         alert_info = self.driver.switch_to.alert.text
#         print(str(alert_info), '\n')
#         time.sleep(2)
#         self.driver.switch_to.alert.accept()
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# # --------------------------- RTP配置(EN) ---------------------------

#     def test_Configuration_RTP_Configuration_EN(self):
#         '''【配置】RTP配置功能頁切換(英)'''  
        
#         self.driver.find_element(By.ID, 'Configuration').click()
#         print("進入配置功能選單!", '\n')
#         time.sleep(2)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Configuration/rtp_config')
#         print("切換RTP配置選單!", '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[1]').click()
#         print("語系已切換:", 'English', '\n')
#         time.sleep(1)    


# #  ************************** Search class **************************

#     def test_search_class_identify_rtp_EN(self):
#         '''搜尋類別功能驗證(英)'''

#         self.driver.find_element(By.ID, f'search_class').click()  # Search class
#         time.sleep(1)

#         for Search_class_list in reversed(range(1, 3)):    
#             Search_class_select = self.driver.find_element(By.XPATH, f'//*[@id="search_class"]/option[{Search_class_list}]')  # 勾選 / 文字                                                            
#             Search_class_select.click()
#             print('Search class:', Search_class_select.text, '\n')
#             time.sleep(1)
            
#         self.driver.refresh()
#         time.sleep(1)


# # ********************** Agent Class與Agent選取 **************************

#     def test_agent_class_and_agent_identify_EN(self):
#         '''代理商類別與代理商選取功能驗證(英)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # Agent Class
#         time.sleep(1) 

#         agent_class_click = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test                                                            
#         agent_class_click.click()
#         print('Agent Class:', agent_class_click.text, '\n')
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="79713763-selectable"]')  # Test-5
#         agent_select.click() 
#         print('(勾選)Agent:', agent_select.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'sh_btn').click()  # 送出
#         time.sleep(1)

#         page_end = self.driver.find_element(By.XPATH, f'/html/body/div[4]/div/div[4]/div[2]/div[2]')  # 頁尾
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)
#         time.sleep(1)

#         page_head = self.driver.find_element(By.XPATH, f'/html/body/div[4]/div/div[4]/div[2]/div[1]/div[1]/div/div[2]/h5') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 頁首
#         time.sleep(1)

#         agent_select_1 = self.driver.find_element(By.XPATH, f'//*[@id="79713763-selectable"]')  # Test-5
#         agent_select_1.click() 
#         print('(取消勾選)Agent:', agent_select_1.text, '\n')
#         time.sleep(2)

#         agent_select_2 = self.driver.find_element(By.XPATH, f'//*[@id="79713760-selectable"]')  # Test-2
#         agent_select_2.click() 
#         print('(勾選)Agent:', agent_select_2.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'sh_btn').click()  # 送出 
#         time.sleep(1)

#         page_end = self.driver.find_element(By.XPATH, f'/html/body/div[4]/div/div[4]/div[2]/div[2]')  # 頁尾
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)
#         time.sleep(1)

#         page_head = self.driver.find_element(By.XPATH, f'/html/body/div[4]/div/div[4]/div[2]/div[1]/div[1]/div/div[2]/h5') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 頁首
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# # ********************** Agent Class選取與Agent輸入 **************************

#     def test_agent_class_and_agent_input_identify_EN(self):
#         '''代理商類別選取與代理商輸入功能驗證(英)'''

#         self.driver.find_element(By.ID, f'search_class').click()  # Search class
#         time.sleep(1)

#         Search_class_select = self.driver.find_element(By.XPATH, f'//*[@id="search_class"]/option[@value="1"]')  # 文字                                                            
#         Search_class_select.click()
#         print('Search class:', Search_class_select.text, '\n')
#         time.sleep(1)

#         agent_input_box = self.driver.find_element(By.ID, f'batch_host_id')  # 代理商輸入框
#         agent_input_box.click() 
#         agent_input_box.send_keys('TEST')
#         time.sleep(1)
#         agent_input_box.send_keys(Keys.ENTER)
#         agent_input_box.send_keys('TEST3')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'sh_btn').click()  # 送出
#         time.sleep(1)

#         page_end = self.driver.find_element(By.XPATH, f'/html/body/div[4]/div/div[4]/div[2]/div[2]')  # 頁尾
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)
#         time.sleep(1)

#         page_head = self.driver.find_element(By.XPATH, f'/html/body/div[4]/div/div[4]/div[2]/div[1]/div[1]/div/div[2]/h5')
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 頁首
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# # ********************** 不輸入Agent(提示訊息驗證) **************************

#     def test_non_agent_data_and_notice_info_identify_EN(self):
#         '''不輸入代理商(提示訊息驗證)(英)'''

#         self.driver.find_element(By.ID, f'search_class').click()  # Search class
#         time.sleep(1)

#         Search_class_select = self.driver.find_element(By.XPATH, f'//*[@id="search_class"]/option[@value="1"]')  # 文字                                                            
#         Search_class_select.click()
#         print('Search class:', Search_class_select.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'sh_btn').click()  # 送出  
#         time.sleep(1)

#         alert_info = self.driver.switch_to.alert.text
#         print(str(alert_info), '\n')
#         time.sleep(2)
#         self.driver.switch_to.alert.accept()
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# # ********************** Agent Class選取與Agent輸入(錯誤訊息驗證) **************************

#     def test_agent_class_and_agent_input_notice_info_identify_EN(self):  
#         '''代理商類別選取與代理商輸入(錯誤訊息驗證)(英)'''

#         self.driver.find_element(By.ID, f'search_class').click()  # Search class
#         time.sleep(1)

#         Search_class_select = self.driver.find_element(By.XPATH, f'//*[@id="search_class"]/option[@value="1"]')  # 文字                                                            
#         Search_class_select.click()
#         print('Search class:', Search_class_select.text, '\n')
#         time.sleep(1)

#         agent_input_box = self.driver.find_element(By.ID, f'batch_host_id')  # 代理商輸入框
#         agent_input_box.click() 
#         agent_input_box.send_keys('TEST-2')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'sh_btn').click()  # 送出
#         time.sleep(1)

#         alert_info = self.driver.switch_to.alert.text
#         print(str(alert_info), '\n')
#         time.sleep(2)
#         self.driver.switch_to.alert.accept()
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# # --------------------------- RTP配置(Tai) ---------------------------

#     def test_Configuration_RTP_Configuration_Tai(self):
#         '''【配置】RTP配置功能頁切換(泰)'''  
        
#         self.driver.find_element(By.ID, 'Configuration').click()
#         print("進入配置功能選單!", '\n')
#         time.sleep(2)

#         self.driver.get('https://dev-admin-br-02.iplaystar.net/Configuration/rtp_config')
#         print("切換RTP配置選單!", '\n')
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="navbar-mobile"]/ul[2]/li/a').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]').click()
#         time.sleep(1)

#         self.driver.find_element(By.XPATH, '//*[@id="lang"]/option[4]').click()
#         print("語系已切換:", 'ไทย', '\n')
#         time.sleep(1)    
    

# #  ************************** Search class **************************

#     def test_search_class_identify_rtp_Tai(self):
#         '''搜尋類別功能驗證(泰)'''

#         self.driver.find_element(By.ID, f'search_class').click()  # Search class
#         time.sleep(1)

#         for Search_class_list in reversed(range(1, 3)):    
#             Search_class_select = self.driver.find_element(By.XPATH, f'//*[@id="search_class"]/option[{Search_class_list}]')  # 勾選 / 文字                                                            
#             Search_class_select.click()
#             print('หมวดหมู่การค้นหา:', Search_class_select.text, '\n')
#             time.sleep(1)
            
#         self.driver.refresh()
#         time.sleep(1)


# # ********************** Agent Class與Agent選取 **************************

#     def test_agent_class_and_agent_identify_Tai(self):
#         '''代理商類別與代理商選取功能驗證(泰)'''

#         self.driver.find_element(By.ID, f'agent_attr').click()  # Agent Class
#         time.sleep(1) 

#         agent_class_click = self.driver.find_element(By.XPATH, f'//*[@id="agent_attr"]/option[@value="Test"]')  # Test                                                            
#         agent_class_click.click()
#         print('หมวดตัวแทน:', agent_class_click.text, '\n')
#         time.sleep(1)

#         agent_select = self.driver.find_element(By.XPATH, f'//*[@id="79713763-selectable"]')  # Test-5
#         agent_select.click() 
#         print('(勾選)ตัวแทน:', agent_select.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'sh_btn').click()  # 送出
#         time.sleep(1)

#         page_end = self.driver.find_element(By.XPATH, f'/html/body/div[4]/div/div[4]/div[2]/div[2]')  # 頁尾
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)
#         time.sleep(1)

#         page_head = self.driver.find_element(By.XPATH, f'/html/body/div[4]/div/div[4]/div[2]/div[1]/div[1]/div/div[2]/h5') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 頁首
#         time.sleep(1)

#         agent_select_1 = self.driver.find_element(By.XPATH, f'//*[@id="79713763-selectable"]')  # Test-5
#         agent_select_1.click() 
#         print('(取消勾選)ตัวแทน:', agent_select_1.text, '\n')
#         time.sleep(2)

#         agent_select_2 = self.driver.find_element(By.XPATH, f'//*[@id="79713760-selectable"]')  # Test-2
#         agent_select_2.click() 
#         print('(勾選)ตัวแทน:', agent_select_2.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'sh_btn').click()  # 送出 
#         time.sleep(1)

#         page_end = self.driver.find_element(By.XPATH, f'/html/body/div[4]/div/div[4]/div[2]/div[2]')  # 頁尾
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)
#         time.sleep(1)

#         page_head = self.driver.find_element(By.XPATH, f'/html/body/div[4]/div/div[4]/div[2]/div[1]/div[1]/div/div[2]/h5') 
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 頁首
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)
        

# # ********************** Agent Class選取與Agent輸入 **************************

#     def test_agent_class_and_agent_input_identify_Tai(self):
#         '''代理商類別選取與代理商輸入功能驗證(泰)'''

#         self.driver.find_element(By.ID, f'search_class').click()  # Search class
#         time.sleep(1)

#         Search_class_select = self.driver.find_element(By.XPATH, f'//*[@id="search_class"]/option[@value="1"]')  # 文字                                                            
#         Search_class_select.click()
#         print('หมวดหมู่การค้นหา:', Search_class_select.text, '\n')
#         time.sleep(1)

#         agent_input_box = self.driver.find_element(By.ID, f'batch_host_id')  # 代理商輸入框
#         agent_input_box.click() 
#         agent_input_box.send_keys('TEST')
#         time.sleep(1)
#         agent_input_box.send_keys(Keys.ENTER)
#         agent_input_box.send_keys('TEST3')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'sh_btn').click()  # 送出
#         time.sleep(1)

#         page_end = self.driver.find_element(By.XPATH, f'/html/body/div[4]/div/div[4]/div[2]/div[2]')  # 頁尾
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_end)
#         time.sleep(1)

#         page_head = self.driver.find_element(By.XPATH, f'/html/body/div[4]/div/div[4]/div[2]/div[1]/div[1]/div/div[2]/h5')
#         self.driver.execute_script("arguments[0].scrollIntoView();", page_head)  # 頁首
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# # ********************** 不輸入Agent(提示訊息驗證) **************************

#     def test_non_agent_data_and_notice_info_identify_Tai(self):
#         '''不輸入代理商(提示訊息驗證)(泰)'''

#         self.driver.find_element(By.ID, f'search_class').click()  # หมวดหมู่การค้นหา
#         time.sleep(1)

#         Search_class_select = self.driver.find_element(By.XPATH, f'//*[@id="search_class"]/option[@value="1"]')  # 文字                                                            
#         Search_class_select.click()
#         print('หมวดหมู่การค้นหา:', Search_class_select.text, '\n')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'sh_btn').click()  # 送出  
#         time.sleep(1)

#         alert_info = self.driver.switch_to.alert.text
#         print(str(alert_info), '\n')
#         time.sleep(2)
#         self.driver.switch_to.alert.accept()
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


# # ********************** Agent Class選取與Agent輸入(錯誤訊息驗證) **************************

#     def test_agent_class_and_agent_input_notice_info_identify_Tai(self):  
#         '''代理商類別選取與代理商輸入(錯誤訊息驗證)(泰)'''

#         self.driver.find_element(By.ID, f'search_class').click()  # หมวดหมู่การค้นหา
#         time.sleep(1)

#         Search_class_select = self.driver.find_element(By.XPATH, f'//*[@id="search_class"]/option[@value="1"]')  # 文字                                                            
#         Search_class_select.click()
#         print('หมวดหมู่การค้นหา:', Search_class_select.text, '\n')
#         time.sleep(1)

#         agent_input_box = self.driver.find_element(By.ID, f'batch_host_id')  # 代理商輸入框
#         agent_input_box.click() 
#         agent_input_box.send_keys('TEST-2')
#         time.sleep(1)

#         self.driver.find_element(By.ID, f'sh_btn').click()  # 送出
#         time.sleep(1)

#         alert_info = self.driver.switch_to.alert.text
#         print(str(alert_info), '\n')
#         time.sleep(2)
#         self.driver.switch_to.alert.accept()
#         time.sleep(1)

#         self.driver.refresh()
#         time.sleep(1)


    def test_PSweb_driver_closed(self):
        '''關閉PS後台自動化測試瀏覽器'''
        
        
        time.sleep(1)
        end_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
        print('測試結束時間: ', end_time, '\n')        
        
        self.driver.quit()
        print('測試結束! PS後台瀏覽器已關閉。')
        time.sleep(1)
        
    

if __name__ == '__main__':
    
    # report_path = r'D:\AutomotiveTest_QA\Back_platform Test\v 3.0.3.1756 250514\Report'
    
    # pytest.main(['-s', '-v', '--html = ./Pytest/Test_report.html'])
    
    # pytest.main(['-s', '-v', '(Pytest)Integrated_Test-20250606_test.py'])
    # pytest.main()
    
    pytest.main(['-s', '-q', '(Pytest)Integrated_Test-20250606_test.py', '--clean-alluredir', '--alluredir=Test_data'])
    os.system(r'allure generate -c -o allure_report')
    
    # pytest.main([ 
    #         "--reportPath = ./AutoControl_test/(x). Back_platform Test/v 3.0.3.1756 250514",     # 報告生成路徑 Tips：当前文件的上一级目录同级目录reports文件夹下
    #         "--report = Test_report(br-02)全功能巡測測試報告.html",   # 報告名稱
    #         '--title = PS後台(br-02)全功能巡測',    # 報告標題
    #         '--tester = Ivan_Li',   # 測試人員
    #         '--desc = 報告描述訊息',   # 報告項目描述
    #         '--template = 2',    # 報告模板(1 or 2)
    #         '-W', "ignore:Module already imported:pytest.PytestWarning"
    #         ])
    
    
    
    # pytest.main([__file__, 
    #             "--reportPath = ../reports",     # 報告生成路徑 Tips：当前文件的上一级目录同级目录reports文件夹下
    #             "--report = Test_report(br-02)全功能巡測測試報告.html",   # 報告名稱
    #             '--title = PS後台(br-02)全功能巡測',    # 報告標題
    #             '--tester = Ivan_Li',   # 測試人員
    #             '--desc = 報告描述訊息',   # 報告項目描述
    #             '--template = 2',    # 報告模板(1 or 2)
    #             '-W', "ignore:Module already imported:pytest.PytestWarning"
    #             ])
    
    # pytest.main([__file__,
    #             "--report = Test_report(br-02)全功能巡測測試報告.html",   # 報告名稱
    #             '--title = PS後台(br-02)全功能巡測',    # 報告標題
    #             '--tester = Ivan_Li',   # 測試人員
    #             '--desc = 報告描述訊息',   # 報告項目描述
    #             '--template = 1',    # 報告模板(1 or 2)
    #             '-W', "ignore:Module already imported:pytest.PytestWarning"
    #             ])

